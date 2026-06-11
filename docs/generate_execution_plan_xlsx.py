import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors (hex, no '#' for openpyxl)
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E5FA3"
LIGHT_BLUE  = "D6E4F0"
ACCENT_GOLD = "C9A84C"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
MID_GRAY    = "CCCCCC"

STATUS_UNREAD   = "F5F5F5"
STATUS_READING  = "FFF3CD"
STATUS_READ     = "D6E4F0"
STATUS_REVIEWED = "D4EDDA"

def side(color=MID_GRAY, style="thin"):
    return Side(border_style=style, color=color)

def border(all_sides=MID_GRAY):
    s = side(all_sides)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Data structures for the 3 implementation approaches & scientific research ──

approach1_data = [
    {
        "id": "1.M1",
        "name": "Research & Setup",
        "type": "MILESTONE",
        "start": "2026-06-09",
        "end": "2026-06-09",
        "o": 2.3, "ml": 4.7, "p": 10.5, "e": 5.25,
        "status": "Complete",
        "notes": "Install Semgrep CLI; scaffold repo; write one toy rule end-to-end.",
        "tasks": [
            {"id": "1.M1.T1", "name": "Install Semgrep CLI and verify installation", "o": 0.3, "ml": 0.5, "p": 1.5, "e": 0.55, "status": "Complete", "notes": ""},
            {"id": "1.M1.T2", "name": "Read Semgrep operator docs and YAML rule syntax documentation", "o": 1.0, "ml": 2.0, "p": 4.5, "e": 2.08, "status": "Complete", "notes": ""},
            {"id": "1.M1.T3", "name": "Write one end-to-end toy rule to confirm setup works", "o": 0.5, "ml": 1.5, "p": 3.5, "e": 1.67, "status": "Complete", "notes": ""},
            {"id": "1.M1.T4", "name": "Scaffold repo structure: rules/, tests/, scripts/ directories with README stubs", "o": 0.2, "ml": 0.5, "p": 1.0, "e": 0.53, "status": "Complete", "notes": ""}
        ]
    },
    {
        "id": "1.M2",
        "name": "Python Custom Rules",
        "type": "MILESTONE",
        "start": "2026-06-10",
        "end": "2026-06-11",
        "o": 4.2, "ml": 9.4, "p": 27.0, "e": 11.98,
        "status": "In Progress",
        "notes": "Write 10 Python rules PY-001 to PY-010 with bad.py / ok.py test pairs. Save hardest rules (bypass, f-string SQL) for last.",
        "tasks": [
            {"id": "1.M2.T1", "name": "Write PY-001: pickle.loads() without type validation (insecure deserialization)", "o": 0.3, "ml": 0.8, "p": 2.0, "e": 0.88, "status": "In Progress", "notes": ""},
            {"id": "1.M2.T2", "name": "Write PY-002: subprocess.run(shell=True) with dynamic input (OS command injection)", "o": 0.3, "ml": 0.8, "p": 2.0, "e": 0.88, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T3", "name": "Write PY-003: eval() / exec() on user-controlled input (code injection)", "o": 0.3, "ml": 0.7, "p": 1.5, "e": 0.75, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T4", "name": "Write PY-004: requests SSL verification bypass (verify=False)", "o": 0.2, "ml": 0.5, "p": 1.5, "e": 0.57, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T5", "name": "Write PY-005: Hardcoded credentials — password, api_key, token string literals", "o": 0.3, "ml": 0.8, "p": 2.0, "e": 0.88, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T6", "name": "Write PY-006: Hardcoded AI service API keys (sk-, sk-ant-, hf_ prefixes)", "o": 0.3, "ml": 0.7, "p": 2.0, "e": 0.80, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T7", "name": "Write PY-007: Path traversal with open() without sanitization", "o": 0.3, "ml": 0.8, "p": 2.5, "e": 0.97, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T8", "name": "Write PY-008: yaml.load() instead of yaml.safe_load() (unsafe YAML parsing)", "o": 0.2, "ml": 0.5, "p": 1.5, "e": 0.57, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T9", "name": "Write PY-009: SQL injection via f-string or % formatting (hardest — save for last)", "o": 0.5, "ml": 1.5, "p": 5.0, "e": 1.92, "status": "Not Started", "notes": ""},
            {"id": "1.M2.T10", "name": "Write PY-010: Unsanitized user input injected into LLM system prompts (AI-specific, hardest)", "o": 0.5, "ml": 2.0, "p": 7.0, "e": 2.58, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "1.M3",
        "name": "Java Custom Rules",
        "type": "MILESTONE",
        "start": "2026-06-12",
        "end": "2026-06-13",
        "o": 4.3, "ml": 10.3, "p": 29.5, "e": 12.84,
        "status": "Not Started",
        "notes": "Write 9 Java rules JV-001 to JV-009. Validate AST shapes with semgrep --dump-ast before authoring.",
        "tasks": [
            {"id": "1.M3.T1", "name": "Run semgrep --dump-ast --lang java on sample file; confirm AST node shapes before writing any rule", "o": 0.2, "ml": 0.5, "p": 1.5, "e": 0.57, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T2", "name": "Write JV-001: Runtime.getRuntime().exec(userInput) — OS command injection", "o": 0.4, "ml": 1.0, "p": 3.0, "e": 1.23, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T3", "name": "Write JV-002: JDBC string concatenation SQL injection (Statement.execute + string concat)", "o": 0.4, "ml": 1.2, "p": 4.0, "e": 1.57, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T4", "name": "Write JV-003: ObjectInputStream.readObject() without class filtering (insecure deserialization)", "o": 0.3, "ml": 1.0, "p": 3.5, "e": 1.30, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T5", "name": "Write JV-004: Hardcoded credentials in Java source (String password, String apiKey literals)", "o": 0.3, "ml": 0.8, "p": 2.0, "e": 0.88, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T6", "name": "Write JV-005: Path traversal with new File(baseDir + userInput) without normalization", "o": 0.3, "ml": 1.0, "p": 3.0, "e": 1.22, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T7", "name": "Write JV-006: Empty X509TrustManager (TLS certificate bypass — checkServerTrusted no-op)", "o": 0.4, "ml": 1.2, "p": 4.0, "e": 1.57, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T8", "name": "Write JV-007: MD5 or SHA1 used for password hashing (weak crypto)", "o": 0.2, "ml": 0.6, "p": 2.0, "e": 0.70, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T9", "name": "Write JV-008: Sensitive data in log statements (logger.info with password/token variable)", "o": 0.3, "ml": 0.8, "p": 2.5, "e": 0.97, "status": "Not Started", "notes": ""},
            {"id": "1.M3.T10", "name": "Write JV-009: AI-agent security bypass comment pattern (SECURITY_BYPASS / disabled for testing)", "o": 0.5, "ml": 1.5, "p": 6.0, "e": 2.08, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "1.M4",
        "name": "Test Codebase",
        "type": "MILESTONE",
        "start": "2026-06-16",
        "end": "2026-06-16",
        "o": 2.4, "ml": 5.7, "p": 16.0, "e": 6.75,
        "status": "Not Started",
        "notes": "AI-generate fake Spring Boot REST API (10-15 files, 800-1200 LOC). Embed >=8 intentional vulnerabilities. Success: Semgrep detects >=6 of 8.",
        "tasks": [
            {"id": "1.M4.T1", "name": "Prompt AI to generate Spring Boot REST API skeleton (controller/service/repository layers, Java 8, 10-15 files)", "o": 0.3, "ml": 0.8, "p": 2.0, "e": 0.88, "status": "Not Started", "notes": ""},
            {"id": "1.M4.T2", "name": "Embed >=8 intentional vulnerabilities into the codebase (mapped to JV-001 to JV-009 rule targets, realistic CVE-pattern placement)", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""},
            {"id": "1.M4.T3", "name": "Run semgrep --config rules/ against the test codebase; record raw findings JSON", "o": 0.2, "ml": 0.5, "p": 1.5, "e": 0.57, "status": "Not Started", "notes": ""},
            {"id": "1.M4.T4", "name": "Triage findings: confirm true positives, document false positives, verify >=6 of 8 embedded vulnerabilities caught", "o": 0.5, "ml": 1.5, "p": 5.0, "e": 1.92, "status": "Not Started", "notes": ""},
            {"id": "1.M4.T5", "name": "Run rule set against one clean open-source Java repo; count and document false positives to establish FP rate baseline", "o": 0.3, "ml": 1.0, "p": 3.5, "e": 1.30, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "1.M5",
        "name": "Demo Preparation",
        "type": "MILESTONE",
        "start": "2026-06-17",
        "end": "2026-06-17",
        "o": 1.9, "ml": 4.6, "p": 13.0, "e": 5.40,
        "status": "Not Started",
        "notes": "Write demo/run_demo.sh with pinned Semgrep version and hardcoded paths. Full dry-run in fresh terminal. Record 3-minute fallback video.",
        "tasks": [
            {"id": "1.M5.T1", "name": "Write demo/run_demo.sh with pinned Semgrep version, hardcoded absolute paths, and annotated output", "o": 0.3, "ml": 1.0, "p": 3.0, "e": 1.22, "status": "Not Started", "notes": ""},
            {"id": "1.M5.T2", "name": "Full dry-run in a fresh terminal (close and reopen shell, no venv activated, execute script cold)", "o": 0.2, "ml": 0.8, "p": 2.5, "e": 0.95, "status": "Not Started", "notes": ""},
            {"id": "1.M5.T3", "name": "Fix any path, version, or environment issues discovered in dry-run", "o": 0.2, "ml": 1.0, "p": 4.0, "e": 1.37, "status": "Not Started", "notes": ""},
            {"id": "1.M5.T4", "name": "Record 3-minute fallback screen recording of the full demo running successfully", "o": 0.3, "ml": 0.8, "p": 2.0, "e": 0.88, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "1.M6",
        "name": "Presentation Narrative",
        "type": "MILESTONE",
        "start": "2026-06-18",
        "end": "2026-06-18",
        "o": 2.0, "ml": 4.4, "p": 12.0, "e": 5.00,
        "status": "Not Started",
        "notes": "Write pros/cons with >=2 real limitations stated honestly. Draft Approach 2 next-step argument. Add speaker notes per slide.",
        "tasks": [
            {"id": "1.M6.T1", "name": "Draft Approach 1 pros/cons section — minimum 2 real limitations stated without apology (no semantic analysis, no interprocedural taint)", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""},
            {"id": "1.M6.T2", "name": "Draft the Approach 2 next-step argument: why Path B (independent LLM semantic scan) follows naturally from Approach 1 limitations", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""},
            {"id": "1.M6.T3", "name": "Add speaker notes to each presentation section; review for tech lead audience register (no fluff, no over-selling)", "o": 0.3, "ml": 1.0, "p": 3.0, "e": 1.22, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "1.M7",
        "name": "Jupyter Notebook (Bonus)",
        "type": "MILESTONE",
        "start": "2026-06-19",
        "end": "2026-06-19",
        "o": 2.3, "ml": 5.6, "p": 16.0, "e": 6.55,
        "status": "Not Started",
        "notes": "[BONUS] Produce five core metrics with charts. Only start if M4 is complete by Jun 17 EOD.",
        "tasks": [
            {"id": "1.M7.T1", "name": "Set up Jupyter notebook with Semgrep JSON output parser; load findings from M4 validation scan", "o": 0.3, "ml": 0.8, "p": 2.5, "e": 0.97, "status": "Not Started", "notes": ""},
            {"id": "1.M7.T2", "name": "Compute and chart precision and recall per rule against the intentional-vulnerability test codebase", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""},
            {"id": "1.M7.T3", "name": "Benchmark scan speed: measure lines/second on test codebase; compare against a cloud SAST round-trip baseline", "o": 0.3, "ml": 0.8, "p": 2.5, "e": 0.97, "status": "Not Started", "notes": ""},
            {"id": "1.M7.T4", "name": "Compute AI-specific detection rate: proportion of AI-pattern rules (PY-010, JV-009) that fire correctly vs. community-rule detections", "o": 0.3, "ml": 1.0, "p": 3.0, "e": 1.22, "status": "Not Started", "notes": ""},
            {"id": "1.M7.T5", "name": "Compute and chart false positive rate on clean codebase (data from M4.T5); add executive summary cell at notebook top", "o": 0.3, "ml": 1.0, "p": 3.5, "e": 1.30, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "1.BUFFER",
        "name": "Explicit Buffer",
        "type": "BUFFER",
        "start": "2026-06-09",
        "end": "2026-06-19",
        "o": 5.0, "ml": 5.0, "p": 5.0, "e": 5.0,
        "status": "Not Started",
        "notes": "Reserved for: bypass rule debugging (M2/M3), Java AST grammar mismatch (M3), detection rate shortfall requiring rule tuning (M4), demo environment polish (M5). Absorbed into whichever milestone overruns first.",
        "tasks": []
    }
]

approach2_data = [
    {
        "id": "2.M1",
        "name": "Core Engine Setup",
        "type": "MILESTONE",
        "start": "2026-06-23",
        "end": "2026-06-26",
        "o": 8.0, "ml": 21.6, "p": 32.0, "e": 21.10,
        "status": "Not Started",
        "notes": "Go learning curve applied (+20% ML). Initialize Go module, wire Cobra CLI, implement directory walker and ZIP ingestion, define core domain types.",
        "tasks": [
            {"id": "2.M1.T1", "name": "Initialize Go module, directory layout, and .gitignore", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.22, "status": "Not Started", "notes": ""},
            {"id": "2.M1.T2", "name": "Wire Cobra CLI framework: root command, --input flag, --output flag, --verbose flag", "o": 1.0, "ml": 3.6, "p": 6.0, "e": 3.43, "status": "Not Started", "notes": ""},
            {"id": "2.M1.T3", "name": "Implement directory walker: recursively enumerate files, respect .gitignore patterns, return file metadata structs", "o": 1.0, "ml": 3.6, "p": 6.0, "e": 3.43, "status": "Not Started", "notes": ""},
            {"id": "2.M1.T4", "name": "Implement ZIP ingestion: detect ZIP input, extract to temp directory, reuse directory walker", "o": 1.0, "ml": 3.6, "p": 6.0, "e": 3.43, "status": "Not Started", "notes": ""},
            {"id": "2.M1.T5", "name": "Define core domain types: FileRecord, Finding, Severity, ConfidenceTier, ScanConfig structs", "o": 1.0, "ml": 3.6, "p": 6.0, "e": 3.43, "status": "Not Started", "notes": ""},
            {"id": "2.M1.T6", "name": "Write unit tests for directory walker and ZIP ingestion; add Makefile targets: build, test, lint", "o": 1.5, "ml": 4.8, "p": 8.0, "e": 4.72, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M2",
        "name": "Differential Indexer",
        "type": "MILESTONE",
        "start": "2026-06-29",
        "end": "2026-06-30",
        "o": 4.0, "ml": 12.0, "p": 18.0, "e": 11.70,
        "status": "Not Started",
        "notes": "Hash-compares input vs previous scan cache. Only changed/new files enter the pipeline. ~80-95% cost reduction on repeat scans.",
        "tasks": [
            {"id": "2.M2.T1", "name": "Design and implement scan cache schema: SHA-256 per file, last-scanned timestamp, stored as JSON on disk", "o": 1.0, "ml": 3.0, "p": 5.0, "e": 2.83, "status": "Not Started", "notes": ""},
            {"id": "2.M2.T2", "name": "Implement hash-compare logic: load previous cache, diff against current file tree, return new/changed/deleted sets", "o": 1.0, "ml": 3.0, "p": 5.0, "e": 2.83, "status": "Not Started", "notes": ""},
            {"id": "2.M2.T3", "name": "Integrate differential indexer into CLI pipeline: skip unchanged files, persist updated cache after scan", "o": 0.5, "ml": 2.4, "p": 4.0, "e": 2.32, "status": "Not Started", "notes": ""},
            {"id": "2.M2.T4", "name": "Write unit tests for hash-compare logic covering new, changed, deleted, and no-change scenarios", "o": 0.5, "ml": 1.8, "p": 3.0, "e": 1.80, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M3",
        "name": "Path A — Semgrep + LLM Verifier",
        "type": "MILESTONE",
        "start": "2026-07-01",
        "end": "2026-07-04",
        "o": 10.0, "ml": 24.0, "p": 36.0, "e": 23.70,
        "status": "Not Started",
        "notes": "Go learning curve applied (+20% ML). LLM Verifier only sees structured findings — not raw code. Targets 88-93% FP reduction.",
        "tasks": [
            {"id": "2.M3.T1", "name": "Install Semgrep binary as subprocess dependency; implement SemgrepRunner: invoke semgrep --json, parse output into []Finding", "o": 1.0, "ml": 3.6, "p": 6.0, "e": 3.43, "status": "Not Started", "notes": ""},
            {"id": "2.M3.T2", "name": "Author 5+ custom Semgrep YAML rules for AI-specific threats: hallucinated package imports, security control bypasses, prompt injection in comments", "o": 2.0, "ml": 6.0, "p": 10.0, "e": 6.00, "status": "Not Started", "notes": ""},
            {"id": "2.M3.T3", "name": "Integrate community Semgrep rulesets (p/owasp-top-ten, p/secrets) into the scan config", "o": 0.5, "ml": 1.8, "p": 3.0, "e": 1.80, "status": "Not Started", "notes": ""},
            {"id": "2.M3.T4", "name": "Implement LLM Verifier: serialize each Finding into structured prompt (taint flow path + sink type + reachability condition); send to local Ollama; parse true/false/uncertain response", "o": 3.0, "ml": 7.2, "p": 12.0, "e": 7.20, "status": "Not Started", "notes": ""},
            {"id": "2.M3.T5", "name": "Wire LLM Verifier into Path A output: filter findings flagged false by LLM, retain true/uncertain; log suppression count", "o": 0.5, "ml": 1.8, "p": 3.0, "e": 1.80, "status": "Not Started", "notes": ""},
            {"id": "2.M3.T6", "name": "Write integration tests for Path A end-to-end using a synthetic vulnerable Go/Java file; assert findings contain expected CWEs", "o": 1.0, "ml": 3.6, "p": 6.0, "e": 3.43, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M4",
        "name": "Path B Tier 1 — Heuristic Targeting + Call Graph",
        "type": "MILESTONE",
        "start": "2026-07-07",
        "end": "2026-07-09",
        "o": 8.0, "ml": 18.0, "p": 30.0, "e": 18.30,
        "status": "Not Started",
        "notes": "Endpoints, auth functions, AI-modified regions — ~95% of files eliminated at zero LLM cost. CVE exact match → auto-flag, skip all further analysis.",
        "tasks": [
            {"id": "2.M4.T1", "name": "Implement heuristic surface selector: regex + AST heuristics to tag endpoint handlers, auth functions, and AI-modified regions (// AI-generated comments, recent git-blame dates)", "o": 2.0, "ml": 4.8, "p": 8.0, "e": 4.80, "status": "Not Started", "notes": ""},
            {"id": "2.M4.T2", "name": "Implement dependency extractor: parse import statements and package.json / go.mod / pom.xml / requirements.txt into a flat dependency list", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M4.T3", "name": "Implement NVD CVE cache: fetch NVD JSON feed weekly, store locally as SQLite, expose query-by-package-name API", "o": 2.0, "ml": 4.8, "p": 8.0, "e": 4.80, "status": "Not Started", "notes": ""},
            {"id": "2.M4.T4", "name": "Wire CVE cross-reference: for each dependency in extracted list, query local CVE cache; auto-flag exact matches as BLOCK-tier findings before LLM", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M4.T5", "name": "Write unit tests for surface selector (verify endpoint/auth/AI-region tagging on synthetic files) and CVE lookup (mock SQLite fixture)", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M5",
        "name": "Path B Tier 2 — UniXcoder Classifier Gate",
        "type": "MILESTONE",
        "start": "2026-07-10",
        "end": "2026-07-13",
        "o": 6.0, "ml": 16.0, "p": 28.0, "e": 16.30,
        "status": "Not Started",
        "notes": "UniXcoder-Base-Nine, ~125M params, F1=94.73% on BigVul. Runs locally on CPU, milliseconds per function, zero API cost. Only ~15-25% of surfaces escalate to Tier 3.",
        "tasks": [
            {"id": "2.M5.T1", "name": "Download and validate UniXcoder-Base-Nine ONNX export; write Python inference wrapper callable from Go via subprocess or gRPC", "o": 1.5, "ml": 4.0, "p": 7.0, "e": 3.92, "status": "Not Started", "notes": ""},
            {"id": "2.M5.T2", "name": "Implement classifier gate logic: confidence >= 0.85 → flag/dismiss directly; 0.55-0.84 → escalate to Tier 3; < 0.55 → dismiss", "o": 1.0, "ml": 2.8, "p": 5.0, "e": 2.80, "status": "Not Started", "notes": ""},
            {"id": "2.M5.T3", "name": "Integrate classifier gate into Path B pipeline after Tier 1 surface selection; pass only uncertain surfaces to Tier 3", "o": 0.5, "ml": 2.0, "p": 4.0, "e": 1.92, "status": "Not Started", "notes": ""},
            {"id": "2.M5.T4", "name": "Write unit tests for classifier gate routing logic using mock inference responses covering all three confidence bands", "o": 0.5, "ml": 2.0, "p": 4.0, "e": 1.92, "status": "Not Started", "notes": ""},
            {"id": "2.M5.T5", "name": "Benchmark classifier latency on CPU for 50 code surfaces; document p50/p95 latency in performance notes", "o": 0.5, "ml": 2.0, "p": 4.0, "e": 1.92, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M6",
        "name": "Path B Tier 3 — Token Budget + LLM Semantic Scan",
        "type": "MILESTONE",
        "start": "2026-07-14",
        "end": "2026-07-16",
        "o": 8.0, "ml": 18.0, "p": 30.0, "e": 18.30,
        "status": "Not Started",
        "notes": "Hard token cap per scan. CFG-based chunking at function boundaries. Sensitive data (credentials/PII) routed to secure model only. LLM Semantic Scan never sees Path A results.",
        "tasks": [
            {"id": "2.M6.T1", "name": "Implement Token Budget Controller: set hard token cap per scan (default 32k tokens); CFG-based chunker splits large surfaces at function boundaries", "o": 2.0, "ml": 4.8, "p": 8.0, "e": 4.80, "status": "Not Started", "notes": ""},
            {"id": "2.M6.T2", "name": "Implement surface prioritizer: rank uncertain surfaces by (CVE base score × classifier uncertainty score) descending; truncate to budget", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M6.T3", "name": "Implement LLM Semantic Scan via Ollama: craft system prompt for IDOR / missing auth guard / business logic flaw detection; never include Path A results in context", "o": 2.0, "ml": 4.8, "p": 8.0, "e": 4.80, "status": "Not Started", "notes": ""},
            {"id": "2.M6.T4", "name": "Parse LLM Semantic Scan response into []Finding with CWE ID, severity, affected lines, and explanation", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M6.T5", "name": "Write integration test for full Path B Tier 3 flow using a synthetic vulnerable endpoint with a deliberate IDOR vulnerability; assert finding is returned", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M7",
        "name": "Dedup + Confidence Scoring",
        "type": "MILESTONE",
        "start": "2026-07-17",
        "end": "2026-07-20",
        "o": 6.0, "ml": 14.0, "p": 24.0, "e": 14.30,
        "status": "Not Started",
        "notes": "Triple-path fusion: AST edit distance + LLM semantic similarity + CWE pattern hash. 5 tiers: BLOCK >=0.92, HIGH 0.75-0.91, MEDIUM 0.60-0.74, LOW 0.30-0.59, SUPPRESSED <0.30. Dual-path confirmation → +15% score boost.",
        "tasks": [
            {"id": "2.M7.T1", "name": "Implement AST edit distance deduplicator: compare code snippet AST hashes across Path A and Path B findings; group findings within edit distance threshold 2", "o": 1.5, "ml": 3.6, "p": 6.0, "e": 3.60, "status": "Not Started", "notes": ""},
            {"id": "2.M7.T2", "name": "Implement LLM semantic similarity deduplicator: embed finding explanations via local embedding model; cluster cosine similarity >= 0.85 as duplicates", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M7.T3", "name": "Implement CWE pattern hash deduplicator: same CWE + same file + overlapping line range → merge", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.20, "status": "Not Started", "notes": ""},
            {"id": "2.M7.T4", "name": "Implement confidence scoring engine: assign base score by tier; apply +15% boost for dual-path confirmation; suppress test-file and framework-safe findings", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M7.T5", "name": "Write unit tests for dedup and scoring: verify dual-path boost, SUPPRESSED findings filtered, correct tier assignment", "o": 0.5, "ml": 1.8, "p": 3.0, "e": 1.80, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M8",
        "name": "HTML Report + Patch Suggestions",
        "type": "MILESTONE",
        "start": "2026-07-21",
        "end": "2026-07-23",
        "o": 8.0, "ml": 16.0, "p": 28.0, "e": 16.70,
        "status": "Not Started",
        "notes": "Self-contained single-file HTML output (all CSS/JS inlined). Filterable findings table, confidence tier badge, expandable code diff, one-click patch copy.",
        "tasks": [
            {"id": "2.M8.T1", "name": "Design HTML report template (Go html/template): severity summary bar, filterable findings table, expandable code diff per finding, confidence tier badge", "o": 2.0, "ml": 4.0, "p": 7.0, "e": 4.00, "status": "Not Started", "notes": ""},
            {"id": "2.M8.T2", "name": "Implement report renderer: consume []Finding from dedup engine, embed all assets inline (CSS, JS) for fully self-contained single-file output", "o": 1.5, "ml": 3.0, "p": 5.0, "e": 3.00, "status": "Not Started", "notes": ""},
            {"id": "2.M8.T3", "name": "Implement patch suggestion generator: for each finding, produce a unified Git diff patch with a minimal secure replacement using LLM or template rules", "o": 2.0, "ml": 4.0, "p": 7.0, "e": 4.00, "status": "Not Started", "notes": ""},
            {"id": "2.M8.T4", "name": "Embed patch diffs into the HTML report as collapsible sections; add one-click copy-to-clipboard button per patch", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.20, "status": "Not Started", "notes": ""},
            {"id": "2.M8.T5", "name": "Write snapshot tests for the HTML renderer: render a fixture []Finding and assert key HTML elements are present (severity badge, CWE ID, patch block)", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M9",
        "name": "Integration Buffer + Demo Prep",
        "type": "MILESTONE",
        "start": "2026-07-24",
        "end": "2026-07-25",
        "o": 6.0, "ml": 12.0, "p": 20.0, "e": 12.30,
        "status": "Not Started",
        "notes": "Run full end-to-end on real-world vulnerable-by-design repo. Fix top-3 integration bugs. Record terminal demo GIF/MP4.",
        "tasks": [
            {"id": "2.M9.T1", "name": "Run full end-to-end scan on a real-world open-source project; triage any integration failures", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M9.T2", "name": "Fix top-3 integration bugs found during end-to-end run; re-run and confirm clean pass", "o": 1.0, "ml": 3.0, "p": 6.0, "e": 3.00, "status": "Not Started", "notes": ""},
            {"id": "2.M9.T3", "name": "Prepare demo codebase: synthetic vulnerable project with 5+ seeded vulnerabilities covering Path A and Path B findings", "o": 1.0, "ml": 2.4, "p": 4.0, "e": 2.40, "status": "Not Started", "notes": ""},
            {"id": "2.M9.T4", "name": "Record terminal demo: scan the demo codebase, show HTML report in browser, walk through one patch suggestion; export as GIF or MP4", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.20, "status": "Not Started", "notes": ""},
            {"id": "2.M9.T5", "name": "Update README with installation steps, usage examples, architecture diagram, and known limitations", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.20, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "2.M10",
        "name": "Presentation",
        "type": "MILESTONE",
        "start": "2026-07-28",
        "end": "2026-07-28",
        "o": 2.0, "ml": 4.0, "p": 8.0, "e": 4.30,
        "status": "Not Started",
        "notes": "Slide deck: problem statement, architecture diagram, live demo, benchmark numbers, roadmap to Approach 3.",
        "tasks": [
            {"id": "2.M10.T1", "name": "Build slide deck: problem statement, architecture diagram, live demo script, benchmark numbers (FP reduction rate, scan latency), roadmap to Approach 3", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.00, "status": "Not Started", "notes": ""},
            {"id": "2.M10.T2", "name": "Dry-run presentation with tech lead; incorporate feedback; commit final slide deck to repository", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.20, "status": "Not Started", "notes": ""},
            {"id": "2.M10.T3", "name": "Deliver Approach 2 demo presentation to stakeholders", "o": 0.5, "ml": 1.2, "p": 2.0, "e": 1.20, "status": "Not Started", "notes": ""}
        ]
    }
]

approach3_data = [
    {
        "id": "3.M1",
        "name": "CodeQL + Joern Integration into Path A",
        "type": "MILESTONE",
        "start": "2026-07-21",
        "end": "2026-07-22",
        "o": 5.5, "ml": 11.5, "p": 20.0, "e": 7.92,
        "status": "Not Started",
        "notes": "Runs in parallel with Semgrep. LLM Verifier extended to deduplicate and score across all three Path A sources.",
        "tasks": [
            {"id": "3.M1.T1", "name": "Install and configure CodeQL CLI; write one end-to-end QL taint-flow query that fires on the Approach 2 Java test codebase", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.17, "status": "Not Started", "notes": ""},
            {"id": "3.M1.T2", "name": "Install Joern; write CPG query script detecting SQL injection and command injection taint paths in the Java test codebase", "o": 1.5, "ml": 3.0, "p": 5.0, "e": 3.08, "status": "Not Started", "notes": ""},
            {"id": "3.M1.T3", "name": "Implement Go parallel runner that fans out Semgrep, CodeQL, and Joern concurrently and merges raw findings into the shared JSON schema from Approach 2", "o": 2.0, "ml": 4.0, "p": 7.0, "e": 4.17, "status": "Not Started", "notes": ""},
            {"id": "3.M1.T4", "name": "Extend LLM Verifier to deduplicate and score false-positive probability across all three Path A sources; output unified findings list", "o": 1.0, "ml": 2.5, "p": 4.0, "e": 2.50, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M2",
        "name": "Path B — Call Graph + CVE Enrichment (Full Integration)",
        "type": "MILESTONE",
        "start": "2026-07-23",
        "end": "2026-07-27",
        "o": 6.5, "ml": 13.5, "p": 21.0, "e": 13.58,
        "status": "Not Started",
        "notes": "Joern call-graph output feeds directly into Tier-1 surface selection. CVE enrichment matches high-risk surfaces to NVD/OSV identifiers by package name and function name.",
        "tasks": [
            {"id": "3.M2.T1", "name": "Integrate Joern call-graph output as Tier-1 input to Path B Heuristic Targeting: map function call edges onto the high-risk surface list from Approach 2", "o": 1.5, "ml": 3.0, "p": 5.0, "e": 3.08, "status": "Not Started", "notes": ""},
            {"id": "3.M2.T2", "name": "Implement CVE enrichment module: fetch and locally cache NVD/OSV feeds; match high-risk surfaces to CVE identifiers by package name and function name", "o": 2.0, "ml": 4.0, "p": 6.0, "e": 4.00, "status": "Not Started", "notes": ""},
            {"id": "3.M2.T3", "name": "Wire existing UniXcoder Tier-2 classifier to consume call-graph-enriched surfaces and output ranked risk scores with calibrated confidence values", "o": 1.0, "ml": 2.5, "p": 4.0, "e": 2.50, "status": "Not Started", "notes": ""},
            {"id": "3.M2.T4", "name": "Implement Tier-3 LLM Semantic Scan: prompt template includes call-graph context and CVE cross-reference; parse structured JSON vulnerability output from local Ollama model", "o": 2.0, "ml": 4.0, "p": 6.0, "e": 4.00, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M3",
        "name": "Red Team Agent (LangGraph Orchestration)",
        "type": "MILESTONE",
        "start": "2026-07-27",
        "end": "2026-07-28",
        "o": 4.5, "ml": 10.5, "p": 18.5, "e": 10.83,
        "status": "Not Started",
        "notes": "LangGraph orientation spike required before design (T1). Graph nodes: triage → exploit-gen → sandbox-dispatch → result-collect. StateAnnotation schema typed in Python.",
        "tasks": [
            {"id": "3.M3.T1", "name": "LangGraph orientation spike: install langgraph, build a two-node graph (planner → executor) against a toy task; confirm StateAnnotation wiring and local Python environment work", "o": 0.5, "ml": 1.5, "p": 3.0, "e": 1.58, "status": "Not Started", "notes": ""},
            {"id": "3.M3.T2", "name": "Design Red Team Agent graph: define nodes (triage, exploit-gen, sandbox-dispatch, result-collect), directed edges, and the StateAnnotation schema with typed fields in Python", "o": 1.0, "ml": 2.5, "p": 4.0, "e": 2.50, "status": "Not Started", "notes": ""},
            {"id": "3.M3.T3", "name": "Implement triage node: selects top-N candidates from dedup+scored findings list by confidence × severity product; emits exploit-plan structs to the exploit-gen node", "o": 1.0, "ml": 2.0, "p": 3.5, "e": 2.08, "status": "Not Started", "notes": ""},
            {"id": "3.M3.T4", "name": "Implement exploit-gen node: calls local Ollama LLM with structured vulnerability context to produce a minimal exploit script or curl payload per finding", "o": 1.5, "ml": 3.0, "p": 5.0, "e": 3.08, "status": "Not Started", "notes": ""},
            {"id": "3.M3.T5", "name": "Implement result-collect node: aggregates container exit codes, stdout/stderr, and timing from sandbox-dispatch into a structured PoE evidence struct; passes to output layer", "o": 0.5, "ml": 1.5, "p": 3.0, "e": 1.58, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M4",
        "name": "Docker Sandbox",
        "type": "MILESTONE",
        "start": "2026-07-29",
        "end": "2026-07-29",
        "o": 4.0, "ml": 7.5, "p": 13.0, "e": 7.83,
        "status": "Not Started",
        "notes": "Alpine base image: Java + Python runtimes, no outbound network, read-only rootfs except /tmp. 30-second timeout + OOM kill limit enforced per container run.",
        "tasks": [
            {"id": "3.M4.T1", "name": "Write Dockerfile for sandbox image: Alpine base with Java and Python runtimes, no outbound network, read-only rootfs except /tmp; build locally and smoke-test with a hello-world exploit script", "o": 1.0, "ml": 2.0, "p": 3.5, "e": 2.08, "status": "Not Started", "notes": ""},
            {"id": "3.M4.T2", "name": "Implement sandbox-dispatch node: spawns sandbox container with exploit payload via Docker SDK, captures stdout/stderr, enforces 30-second timeout and OOM kill limit", "o": 2.0, "ml": 3.5, "p": 6.0, "e": 3.67, "status": "Not Started", "notes": ""},
            {"id": "3.M4.T3", "name": "Implement exploit outcome classifier: parses container output to assign CONFIRMED, NOT_TRIGGERED, or ENVIRONMENT_ERROR status; writes structured result to agent state", "o": 1.0, "ml": 2.0, "p": 3.5, "e": 2.08, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M5",
        "name": "Two-Layer PoE Output Generation",
        "type": "MILESTONE",
        "start": "2026-07-30",
        "end": "2026-07-30",
        "o": 2.5, "ml": 5.0, "p": 9.0, "e": 5.25,
        "status": "Not Started",
        "notes": "Layer 1: technical trace (exploit script + container stdout + stack trace + vulnerable code). Layer 2: executive summary (plain-English, business impact, no code). Serialized to poe_report.json keyed by finding ID.",
        "tasks": [
            {"id": "3.M5.T1", "name": "Implement technical trace renderer: for each CONFIRMED finding produce a structured section containing exploit script, container stdout, stack trace excerpt, and vulnerable code snippet with line numbers", "o": 1.0, "ml": 2.0, "p": 3.5, "e": 2.08, "status": "Not Started", "notes": ""},
            {"id": "3.M5.T2", "name": "Implement executive summary renderer: one-paragraph plain-English description per confirmed finding with business impact, reproduction steps (no code), and severity badge", "o": 1.0, "ml": 2.0, "p": 3.5, "e": 2.08, "status": "Not Started", "notes": ""},
            {"id": "3.M5.T3", "name": "Write PoE serializer: bundle technical trace and executive summary into poe_report.json keyed by finding ID for downstream HTML template embedding", "o": 0.5, "ml": 1.0, "p": 2.0, "e": 1.08, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M6",
        "name": "HTML Report with PoE Evidence",
        "type": "MILESTONE",
        "start": "2026-07-30",
        "end": "2026-07-31",
        "o": 3.0, "ml": 6.5, "p": 11.0, "e": 6.67,
        "status": "Not Started",
        "notes": "Extends Approach 2 HTML template. Adds collapsible PoE evidence panel per finding. Report-level summary: CONFIRMED / NOT_TRIGGERED / UNVERIFIED counts + export button for standalone executive report.",
        "tasks": [
            {"id": "3.M6.T1", "name": "Extend Approach 2 HTML template: add a collapsible PoE evidence panel per finding that embeds the technical trace and executive summary rendered from poe_report.json", "o": 1.5, "ml": 3.0, "p": 5.0, "e": 3.08, "status": "Not Started", "notes": ""},
            {"id": "3.M6.T2", "name": "Add report-level PoE summary section: CONFIRMED / NOT_TRIGGERED / UNVERIFIED counts, risk heat-map, and export button that generates a standalone plain-text executive report", "o": 1.0, "ml": 2.5, "p": 4.0, "e": 2.50, "status": "Not Started", "notes": ""},
            {"id": "3.M6.T3", "name": "End-to-end smoke test: run full pipeline on fake Java test codebase; verify HTML renders in browser, all PoE panels load, and patch diffs display without errors", "o": 0.5, "ml": 1.0, "p": 2.0, "e": 1.08, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M7",
        "name": "Integration Buffer + Final Demo Prep",
        "type": "MILESTONE",
        "start": "2026-07-31",
        "end": "2026-07-31",
        "o": 3.0, "ml": 6.0, "p": 10.5, "e": 6.25,
        "status": "Not Started",
        "notes": "Full pipeline integration test: Path A → Path B → Red Team Agent → Docker Sandbox → PoE Output → HTML Report. Single-command reproducibility via demo/run_demo_approach3.sh. 3-minute fallback recording.",
        "tasks": [
            {"id": "3.M7.T1", "name": "Full pipeline integration test: execute Path A → Path B → Red Team Agent → Docker Sandbox → PoE Output → HTML Report end-to-end on the fake Java codebase; document and fix all breakages", "o": 2.0, "ml": 4.0, "p": 7.0, "e": 4.17, "status": "Not Started", "notes": ""},
            {"id": "3.M7.T2", "name": "Write demo/run_demo_approach3.sh with pinned dependency versions and hardcoded paths; execute a dry-run in a fresh terminal to confirm single-command reproducibility", "o": 0.5, "ml": 1.0, "p": 2.0, "e": 1.08, "status": "Not Started", "notes": ""},
            {"id": "3.M7.T3", "name": "Record a 3-minute fallback video of the full pipeline run including the HTML report PoE evidence panel and executive summary export", "o": 0.5, "ml": 1.0, "p": 1.5, "e": 1.00, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "3.M8",
        "name": "Final Presentation",
        "type": "MILESTONE",
        "start": "2026-07-31",
        "end": "2026-08-01",
        "o": 1.5, "ml": 3.0, "p": 5.0, "e": 3.08,
        "status": "Not Started",
        "notes": "Slides: achievements, architecture walk-through, live demo, honest limitations (false PoE rate, Docker setup complexity), future roadmap. Two full rehearsal runs targeting 5 min each.",
        "tasks": [
            {"id": "3.M8.T1", "name": "Write presentation slides: Approach 3 achievements, architecture diagram walk-through, live demo script, honest limitations, and future roadmap", "o": 1.0, "ml": 2.0, "p": 3.0, "e": 2.00, "status": "Not Started", "notes": ""},
            {"id": "3.M8.T2", "name": "Rehearse live demo walkthrough end-to-end twice: scan → PoE confirmation → HTML report → executive summary export; target 5 minutes per run", "o": 0.5, "ml": 1.0, "p": 2.0, "e": 1.08, "status": "Not Started", "notes": ""},
            {"id": "3.M8.T3", "name": "Deliver final presentation to tech lead (2026-08-01)", "o": 0.0, "ml": 0.0, "p": 0.0, "e": 0.00, "status": "Not Started", "notes": ""}
        ]
    }
]

research_data = [
    {
        "id": "R.M1",
        "name": "Literature Foundation",
        "type": "MILESTONE",
        "start": "2026-06-09",
        "end": "2026-06-20",
        "o": 20.0, "ml": 41.0, "p": 82.0, "e": 44.33,
        "status": "In Progress",
        "notes": "Read and annotate all 40 catalogued papers across 7 areas. Build a paper-to-architecture-component linkage map assigning each paper to >= 1 ZeroTrust.sh component as primary evidence.",
        "tasks": [
            {"id": "R.M1.T1", "name": "Read and annotate Area 1 (Deep Learning & ML for Vulnerability Detection, ~4 papers) and Area 2 (GNN for Vulnerability Detection, ~6 papers): record model architectures, training datasets, F1/precision benchmarks, and which ZeroTrust.sh component each paper most directly supports", "o": 4.0, "ml": 8.0, "p": 16.0, "e": 8.67, "status": "In Progress", "notes": ""},
            {"id": "R.M1.T2", "name": "Read and annotate Area 3 (LLM for Code Security Analysis, ~6 papers): note LLM prompting strategies, hybrid pipeline designs, and accuracy figures relative to ZeroTrust.sh Path B design", "o": 6.0, "ml": 12.0, "p": 22.0, "e": 12.67, "status": "Not Started", "notes": ""},
            {"id": "R.M1.T3", "name": "Read and annotate Area 4 (Hybrid Static Analysis + LLM, ~5 papers) and Area 5 (AI-Generated Code Security & Prompt Injection, ~5 papers): extract FP reduction rates and AI-specific threat taxonomies relevant to slopsquatting and prompt injection detection", "o": 5.0, "ml": 10.0, "p": 20.0, "e": 10.83, "status": "Not Started", "notes": ""},
            {"id": "R.M1.T4", "name": "Read and annotate Area 6 (Token Cost Optimization, ~4 papers) and Area 7 (Call Graph, Taint Analysis & Code Representations, ~10 papers): record cost reduction percentages, uncertainty-based routing approaches, and taint tracking accuracy figures", "o": 4.0, "ml": 8.0, "p": 18.0, "e": 9.00, "status": "Not Started", "notes": ""},
            {"id": "R.M1.T5", "name": "Build paper-to-architecture-component linkage map: assign each of the 40 papers to >= 1 ZeroTrust.sh component (Differential Indexer, Path A Semgrep, Path A CodeQL/Joern, Path A LLM Verifier, Path B Tier 1/2/3, Dedup/Confidence Scoring, PoE Layer) as primary evidence", "o": 1.0, "ml": 3.0, "p": 6.0, "e": 3.17, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "R.M2",
        "name": "Architecture Validation Matrix",
        "type": "MILESTONE",
        "start": "2026-06-16",
        "end": "2026-06-27",
        "o": 4.0, "ml": 8.5, "p": 21.0, "e": 9.83,
        "status": "Not Started",
        "notes": "Map every architecture component to >= 1 supporting paper. Flag any component with no academic backing as a validation gap with a defined mitigation (benchmark proposal, practitioner reference, or novel design claim).",
        "tasks": [
            {"id": "R.M2.T1", "name": "Map Differential Indexer and all Path A sub-components (Semgrep rules, CodeQL/Joern taint analysis, LLM Verifier CoT framework) to >= 1 supporting paper each — populate matrix with columns: component name, paper ID, claimed metric, page/section reference", "o": 1.0, "ml": 2.0, "p": 5.0, "e": 2.33, "status": "Not Started", "notes": ""},
            {"id": "R.M2.T2", "name": "Map Path B three-tier cost funnel (Heuristic Targeting, UniXcoder-Base-Nine Classifier Gate F1=94.73% on BigVul, Token Budget Controller UCCI-style calibration, LLM Semantic Scan) to supporting papers with specific claimed performance figures recorded", "o": 1.0, "ml": 2.0, "p": 5.0, "e": 2.33, "status": "Not Started", "notes": ""},
            {"id": "R.M2.T3", "name": "Map Dedup layer and 5-tier Confidence Scoring (BLOCK >=0.92 / HIGH 0.75-0.91 / MEDIUM 0.60-0.74 / LOW 0.30-0.59 / SUPPRESSED <0.30; +15% dual-path boost) to supporting papers — verify triple-path fusion approach (AST edit distance + LLM semantic similarity + CWE pattern hash)", "o": 0.5, "ml": 1.0, "p": 3.0, "e": 1.25, "status": "Not Started", "notes": ""},
            {"id": "R.M2.T4", "name": "Map Proof-of-Exploit Layer (Red Team Agent orchestration, Docker sandbox exploit execution, two-layer PoE output) to supporting papers on automated vulnerability verification, sandbox-based exploit confirmation, and agentic security frameworks", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""},
            {"id": "R.M2.T5", "name": "Identify every architecture component with zero academic backing and document each as an open validation gap — for each gap specify mitigation: (a) empirical benchmark proposal, (b) industry practitioner reference, or (c) novel design claim requiring ablation study", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.17, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "R.M3",
        "name": "Benchmark & Performance Claims Verification",
        "type": "MILESTONE",
        "start": "2026-06-23",
        "end": "2026-07-04",
        "o": 3.0, "ml": 6.5, "p": 18.0, "e": 7.83,
        "status": "Not Started",
        "notes": "Verify every quantitative claim in the architecture docs against its source paper. Produce a benchmarks reference table: claim text, source, confidence level (paper-cited / internally-estimated / unverified), and follow-up action.",
        "tasks": [
            {"id": "R.M3.T1", "name": "Verify UniXcoder-Base-Nine F1=94.73% on BigVul: locate original paper, confirm exact evaluation setup (dataset split ratio, positive/negative class balance, metric definition), and record any scope caveats (language coverage, vulnerability categories included)", "o": 0.5, "ml": 1.0, "p": 3.0, "e": 1.25, "status": "Not Started", "notes": ""},
            {"id": "R.M3.T2", "name": "Verify 88-93% false positive reduction claim for the LLM Verifier: confirm paper title, publication venue, year, experimental conditions, and whether the task is comparable to ZeroTrust.sh's SAST FP filtering use case", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""},
            {"id": "R.M3.T3", "name": "Verify ~80-95% cost reduction claim for the Differential Indexer: determine whether this is from a cited paper or is an internal engineering estimate; if internal, document the derivation assumptions (file change frequency, typical codebase churn rate)", "o": 0.5, "ml": 1.0, "p": 3.0, "e": 1.25, "status": "Not Started", "notes": ""},
            {"id": "R.M3.T4", "name": "Verify ~15-25% LLM escalation rate (uncertain surfaces reaching Tier 3) and ~95% file elimination rate (Tier 1 heuristic targeting): locate source paper or empirical basis; if internally estimated, document the heuristic targeting precision assumption", "o": 0.5, "ml": 1.0, "p": 3.0, "e": 1.25, "status": "Not Started", "notes": ""},
            {"id": "R.M3.T5", "name": "Compile benchmarks reference table with columns: claim text as stated in architecture docs, source paper or derivation method, confidence level (paper-cited / internally-estimated / unverified), and follow-up action — this becomes the primary fact-check appendix for tech lead presentations", "o": 1.0, "ml": 2.0, "p": 5.0, "e": 2.33, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "R.M4",
        "name": "Competitive Landscape Research",
        "type": "MILESTONE",
        "start": "2026-06-30",
        "end": "2026-07-11",
        "o": 6.0, "ml": 12.0, "p": 26.0, "e": 13.33,
        "status": "Not Started",
        "notes": "Research Semgrep, Snyk Code, and CodeRabbit structural gaps vs. ZeroTrust.sh. Find academic and practitioner evidence for ZeroTrust.sh's local-only + AI-specific threat detection differentiator. Compile a differentiation evidence table.",
        "tasks": [
            {"id": "R.M4.T1", "name": "Research publicly available accuracy and precision data for Semgrep OSS and Semgrep Pro: check official docs, blog posts, academic evaluations that include Semgrep as a baseline, and SAST benchmark studies (OWASP Benchmark, NIST SATE)", "o": 1.0, "ml": 2.0, "p": 5.0, "e": 2.33, "status": "Not Started", "notes": ""},
            {"id": "R.M4.T2", "name": "Research Snyk Code (DeepCode AI engine) architecture with focus on cloud dependency and source-code-upload requirements: find documented regulatory blockers (GDPR, SOC 2, air-gapped environments), published privacy incidents, or enterprise objections that ZeroTrust.sh's local-only model directly addresses", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.17, "status": "Not Started", "notes": ""},
            {"id": "R.M4.T3", "name": "Research CodeRabbit's PR-gated workflow architecture: document what CodeRabbit structurally cannot do (pre-commit local scan, offline execution, ZIP archive input, CI-free developer loop, non-GitHub VCS) and map each architectural gap to the corresponding ZeroTrust.sh capability", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.17, "status": "Not Started", "notes": ""},
            {"id": "R.M4.T4", "name": "Search academic databases (ACM Digital Library, arXiv cs.CR, IEEE Xplore) and NIST guidelines for evidence that local-only execution with AI-specific threat detection (slopsquatting, prompt injection in comments, safety-gate bypass patterns) represents an unoccupied or under-served position in the security tooling landscape", "o": 2.0, "ml": 4.0, "p": 9.0, "e": 4.50, "status": "Not Started", "notes": ""},
            {"id": "R.M4.T5", "name": "Compile differentiation evidence table with columns: ZeroTrust.sh differentiator, evidence source, competitor it distinguishes against, and whether the gap is architectural (cannot be closed without redesign) or feature-level (could be added by competitor)", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.17, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "R.M5",
        "name": "Architecture Justification Document",
        "type": "MILESTONE",
        "start": "2026-07-07",
        "end": "2026-07-25",
        "o": 6.5, "ml": 13.0, "p": 30.0, "e": 14.75,
        "status": "Not Started",
        "notes": "Formal document linking every major design decision to research evidence. Written for tech lead + stakeholders. Consumes outputs from R.M2 (validation matrix) and R.M3 (benchmarks reference table) as appendices.",
        "tasks": [
            {"id": "R.M5.T1", "name": "Write Path A justification section: why Semgrep is the correct pattern-detection engine (speed, community rules, extensibility), why CodeQL/Joern adds necessary taint-aware cross-file coverage, and why the LLM Verifier targeting 88-93% FP reduction is needed — cite supporting papers from R.M1 verified in R.M3", "o": 1.0, "ml": 2.0, "p": 5.0, "e": 2.33, "status": "Not Started", "notes": ""},
            {"id": "R.M5.T2", "name": "Write Path B justification section: explain the cost-funnel rationale for each tier — ~95% file elimination, UniXcoder-Base-Nine F1=94.73%, UCCI-style uncertainty calibration, CFG-based chunking, and why only ~15-25% of surfaces reach the LLM — cite papers for each sub-decision", "o": 2.0, "ml": 4.0, "p": 9.0, "e": 4.50, "status": "Not Started", "notes": ""},
            {"id": "R.M5.T3", "name": "Write Dedup and 5-tier Confidence Scoring justification: explain triple-path fusion (AST edit distance + LLM semantic similarity + CWE pattern hash), cite research for multi-tier scoring, document +15% dual-path confidence boost rationale and test-file suppression rule", "o": 1.0, "ml": 2.0, "p": 5.0, "e": 2.33, "status": "Not Started", "notes": ""},
            {"id": "R.M5.T4", "name": "Write Differential Indexer justification: explain hash-based incremental scanning design, cite evidence basis for 80-95% cost reduction claim, articulate why pre-commit developer-loop integration requires this optimization", "o": 0.5, "ml": 1.0, "p": 3.0, "e": 1.25, "status": "Not Started", "notes": ""},
            {"id": "R.M5.T5", "name": "Write PoE Layer justification and compile tech-lead narrative summary: explain phased deferral to Approach 3, summarize all major design decisions and evidence in <= 2 pages for non-technical stakeholders, attach validation matrix and benchmarks reference table as appendices", "o": 2.0, "ml": 4.0, "p": 8.0, "e": 4.33, "status": "Not Started", "notes": ""}
        ]
    },
    {
        "id": "R.M6",
        "name": "Ongoing Research Monitoring",
        "type": "MILESTONE",
        "start": "2026-07-14",
        "end": "2026-08-01",
        "o": 3.5, "ml": 8.5, "p": 19.0, "e": 9.42,
        "status": "Not Started",
        "notes": "Set up alert infrastructure for post-June-2026 papers. Perform forward-citation pass on top 5 most-cited papers in the catalogue. Write monitoring runbook for ongoing use beyond the internship.",
        "tasks": [
            {"id": "R.M6.T1", "name": "Configure arXiv email digest alerts for all 7 research areas using category filters cs.CR (cryptography and security), cs.SE (software engineering), and cs.LG (machine learning) combined with keyword filters covering Cascading Intelligence Pipeline components", "o": 0.5, "ml": 1.0, "p": 2.0, "e": 1.08, "status": "Not Started", "notes": ""},
            {"id": "R.M6.T2", "name": "Configure Google Scholar keyword alerts for: 'slopsquatting', 'UniXcoder vulnerability detection', 'LLM false positive reduction SAST', 'AI-generated code security', 'token budget LLM code analysis', 'differential program analysis incremental', 'proof of exploit automated'", "o": 0.5, "ml": 1.0, "p": 2.0, "e": 1.08, "status": "Not Started", "notes": ""},
            {"id": "R.M6.T3", "name": "Perform forward-citation pass on top 5 most-cited papers in the 40-paper catalogue using Google Scholar 'Cited by' — identify any 2025-2026 follow-up papers that may supersede original findings, introduce higher benchmarks, or cover gaps flagged in R.M2.T5", "o": 1.0, "ml": 3.0, "p": 7.0, "e": 3.33, "status": "Not Started", "notes": ""},
            {"id": "R.M6.T4", "name": "Write monitoring runbook documenting: exact search queries per area, weekly arXiv digest review steps, monthly deep-search protocol, and standard checklist for integrating a newly found paper into the validation matrix (R.M2) and updating affected benchmark claims (R.M3)", "o": 1.0, "ml": 2.0, "p": 4.0, "e": 2.17, "status": "Not Started", "notes": ""},
            {"id": "R.M6.T5", "name": "Integrate newly discovered papers from R.M6.T3 into the validation matrix and architecture-to-paper linkage map; update benchmarks reference table if new figures supersede existing claims; flag remaining gaps for post-August-1 follow-up", "o": 0.5, "ml": 1.5, "p": 4.0, "e": 1.75, "status": "Not Started", "notes": ""}
        ]
    }
]

constraints_data = [
    {
        "id": "C-01",
        "constraint": "Developer has no prior Go experience entering Approach 2",
        "category": "Experience",
        "impact": "High",
        "applied_to": "Approach 2 (M1–M3); Approach 3 (all milestones)",
        "buffer_added": "+20% applied to ML estimate on all Go-first tasks in M1–M3",
        "notes": "AI assistance (Copilot / Claude Code) estimated to reduce effective learning time by 40-60% vs. unaided learning. Buffer not reduced — AI reduces blocked-state duration, not probability."
    },
    {
        "id": "C-02",
        "constraint": "8-week internship window (Jun 9 – Aug 1, 2026) at approximately 6 productive hours per day",
        "category": "Time",
        "impact": "High",
        "applied_to": "All approaches",
        "buffer_added": "Total capacity: ~240 h. Allocated: ~315 h PERT E across all goals. Overlap between Research and Approach tasks allows parallel work within the same day.",
        "notes": "Research milestones (R.M1–R.M6) are designed to be interleaved with implementation work, not sequential. Research reading occurs during low-energy periods; coding during peak focus hours."
    },
    {
        "id": "C-03",
        "constraint": "Approach 1 tech lead presentation is fixed at 2026-06-20 — this date cannot slip",
        "category": "Time",
        "impact": "High",
        "applied_to": "Approach 1 (all milestones)",
        "buffer_added": "None — date is immovable. Bonus milestone 1.M7 is explicitly cut first if M4 runs late.",
        "notes": "If M4 (Test Codebase) is not complete by Jun 17 EOD, 1.M7 (Jupyter Notebook) is automatically cancelled. Demo recording fallback (1.M5.T4) must exist before Jun 17."
    },
    {
        "id": "C-04",
        "constraint": "Approach 2 presentation targeted for 2026-07-28",
        "category": "Time",
        "impact": "Medium",
        "applied_to": "Approach 2 (M10)",
        "buffer_added": "2.M9 (Integration Buffer + Demo Prep, Jun 24-25) provides 2-day absorption window.",
        "notes": "Date is soft — can slip 1-2 days if integration testing reveals critical bugs. Must not slip past Aug 1 (intern's final day)."
    },
    {
        "id": "C-05",
        "constraint": "Approach 3 final presentation is the intern's last deliverable — deadline is 2026-08-01",
        "category": "Time",
        "impact": "High",
        "applied_to": "Approach 3 (all milestones)",
        "buffer_added": "3.M7 (Integration Buffer + Final Demo Prep, Jul 31) is the only buffer. No slack beyond that date.",
        "notes": "Approach 3 is 61.4h of PERT E in 54h of available capacity. Intern must escalate blockers by day 3 (Jul 23) if M2 is behind."
    },
    {
        "id": "C-06",
        "constraint": "Ollama LLM runtime requires a 4-8 GB quantized GGUF model download before any LLM task in Approach 2 can begin",
        "category": "Tool",
        "impact": "Medium",
        "applied_to": "Approach 2 (M3: LLM Verifier; M6: LLM Semantic Scan)",
        "buffer_added": "0.5 day pre-setup buffer built into the Jun 23-26 sprint start week",
        "notes": "Model: CodeLlama-7B-Instruct or Mistral-7B-Instruct-v0.2 (recommended). Download must happen before Jun 30 (M3 start date). Corporate network proxy settings may add setup time."
    },
    {
        "id": "C-07",
        "constraint": "Approach 3 Docker sandbox requires Docker Desktop installed and running on the development machine",
        "category": "Environment",
        "impact": "Medium",
        "applied_to": "Approach 3 (M4: Docker Sandbox)",
        "buffer_added": "0.5 day installation buffer included in 3.M4 PERT P estimate",
        "notes": "Corporate machines may have Docker Desktop blocked by IT policy. Fallback: use Podman (drop-in replacement with same Docker SDK API). Verify Docker availability on Jul 14 — 1 week before Approach 3 begins."
    },
    {
        "id": "C-08",
        "constraint": "Semgrep CLI version must be pinned in all demo scripts to ensure reproducible output across machines",
        "category": "Tool",
        "impact": "Low",
        "applied_to": "Approach 1 (M5: Demo Preparation); Approach 2 (M9: Demo Prep)",
        "buffer_added": "1 h included in demo prep tasks to handle version verification",
        "notes": "Semgrep YAML rule syntax can differ between minor versions. Pin to the version used during rule development. Document in demo/README.md."
    },
    {
        "id": "C-09",
        "constraint": "Mentor/tech lead review cycles take 2-3 business days per round",
        "category": "Dependency",
        "impact": "Medium",
        "applied_to": "All approaches — presentation milestones and architecture decisions",
        "buffer_added": "Explicit buffer rows in each approach absorb review wait time",
        "notes": "Architecture sign-off for Approach 2/3 is gated on Approach 1 tech lead approval (Jun 20). Submit Approach 1 materials at least 2 days before Jun 20 to allow feedback incorporation."
    },
    {
        "id": "C-10",
        "constraint": "Detailed planning for Approach 2 and Approach 3 is deferred until the Approach 1 tech lead presentation is approved (Jun 20)",
        "category": "Dependency",
        "impact": "Medium",
        "applied_to": "Approach 2 (start gate); Approach 3 (start gate)",
        "buffer_added": "Approach 2 officially starts Jun 23 — 3 days after the Approach 1 presentation, allowing for feedback incorporation and scope adjustment.",
        "notes": "If Approach 1 tech lead requests major scope changes, the Approach 2 milestones may need to be re-estimated. The 3-day gap (Jun 20-23) is intentionally left as a planning and adjustment window."
    }
]

def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    
    # Title row (Row 1)
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "ZeroTrust.sh  ·  Executive Dashboard"
    title.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    title.fill = fill(DARK_BLUE)
    title.alignment = align("center", "center")
    ws.row_dimensions[1].height = 50
    
    # Subtitle row (Row 2)
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Date range: 2026-06-09 to 2026-08-01"
    sub.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    sub.fill = fill(MID_BLUE)
    sub.alignment = align("center", "center")
    ws.row_dimensions[2].height = 22
    
    # Spacer (Row 3)
    ws.row_dimensions[3].height = 6
    
    # Zone 1 Header (Row 4)
    ws.merge_cells("A4:G4")
    z1 = ws["A4"]
    z1.value = "PROJECT KPIs"
    z1.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z1.fill = fill(DARK_BLUE)
    z1.alignment = align("left", "center")
    ws.row_dimensions[4].height = 30
    
    # KPI Labels (Row 5)
    kpi_labels = ["Total Tasks", "Completed", "In Progress", "Blocked", "Days to Deadline", "% Complete"]
    ws.row_dimensions[5].height = 22
    for col_idx, label in enumerate(kpi_labels, start=1):
        cell = ws.cell(row=5, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        
    # G5 border
    ws.cell(row=5, column=7).border = border()
    
    # KPI Values (Row 6)
    ws.row_dimensions[6].height = 60
    
    ws.cell(row=6, column=1, value="=COUNTIF('Approach 1 - Semgrep PoC'!C:C,\"TASK\")+COUNTIF('Approach 2 - Hybrid LLM'!C:C,\"TASK\")+COUNTIF('Approach 3 - Agentic Scanner'!C:C,\"TASK\")+COUNTIF('Research'!C:C,\"TASK\")")
    ws.cell(row=6, column=2, value="=COUNTIFS('Approach 1 - Semgrep PoC'!C:C,\"TASK\",'Approach 1 - Semgrep PoC'!K:K,\"Complete\")+COUNTIFS('Approach 2 - Hybrid LLM'!C:C,\"TASK\",'Approach 2 - Hybrid LLM'!K:K,\"Complete\")+COUNTIFS('Approach 3 - Agentic Scanner'!C:C,\"TASK\",'Approach 3 - Agentic Scanner'!K:K,\"Complete\")+COUNTIFS('Research'!C:C,\"TASK\",'Research'!K:K,\"Complete\")")
    ws.cell(row=6, column=3, value="=COUNTIFS('Approach 1 - Semgrep PoC'!C:C,\"TASK\",'Approach 1 - Semgrep PoC'!K:K,\"In Progress\")+COUNTIFS('Approach 2 - Hybrid LLM'!C:C,\"TASK\",'Approach 2 - Hybrid LLM'!K:K,\"In Progress\")+COUNTIFS('Approach 3 - Agentic Scanner'!C:C,\"TASK\",'Approach 3 - Agentic Scanner'!K:K,\"In Progress\")+COUNTIFS('Research'!C:C,\"TASK\",'Research'!K:K,\"In Progress\")")
    ws.cell(row=6, column=4, value="=COUNTIFS('Approach 1 - Semgrep PoC'!C:C,\"TASK\",'Approach 1 - Semgrep PoC'!K:K,\"Blocked\")+COUNTIFS('Approach 2 - Hybrid LLM'!C:C,\"TASK\",'Approach 2 - Hybrid LLM'!K:K,\"Blocked\")+COUNTIFS('Approach 3 - Agentic Scanner'!C:C,\"TASK\",'Approach 3 - Agentic Scanner'!K:K,\"Blocked\")+COUNTIFS('Research'!C:C,\"TASK\",'Research'!K:K,\"Blocked\")")
    ws.cell(row=6, column=5, value="=DATE(2026,8,1)-TODAY()")
    ws.cell(row=6, column=6, value="=B6/A6")
    
    for c_idx in range(1, 7):
        cell = ws.cell(row=6, column=c_idx)
        cell.font = Font(name="Calibri", size=20, bold=True, color="000000")
        cell.fill = fill(LIGHT_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        if c_idx == 6:
            cell.number_format = '0.0%'
        elif c_idx == 5:
            cell.number_format = '#,##0'
            
    # G6 border
    ws.cell(row=6, column=7).border = border()
    
    # Spacer (Row 7)
    ws.row_dimensions[7].height = 6
    
    # Zone 2 Header (Row 8)
    ws.merge_cells("A8:G8")
    z2 = ws["A8"]
    z2.value = "PROGRESS BY GOAL"
    z2.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z2.fill = fill(DARK_BLUE)
    z2.alignment = align("left", "center")
    ws.row_dimensions[8].height = 30
    
    # Column Headers (Row 9)
    z2_headers = ["Goal", "Not Started", "In Progress", "Complete", "Blocked", "Total Tasks", "% Done"]
    ws.row_dimensions[9].height = 22
    for col_idx, h in enumerate(z2_headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        
    goals_data = [
        ("Approach 1 — Semgrep PoC", "Approach 1 - Semgrep PoC"),
        ("Approach 2 — Hybrid LLM Scanner", "Approach 2 - Hybrid LLM"),
        ("Approach 3 — Agentic Scanner", "Approach 3 - Agentic Scanner"),
        ("Scientific Research", "Research")
    ]
    
    for idx, (label, sheet) in enumerate(goals_data, start=10):
        ws.row_dimensions[idx].height = 30
        
        # Col A
        cell_a = ws.cell(row=idx, column=1, value=label)
        cell_a.font = Font(name="Calibri", size=12, bold=True, color="000000")
        cell_a.fill = fill("F5F5F5")
        cell_a.alignment = align("left", "center")
        cell_a.border = border()
        
        formulas = [
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"Not Started\")",
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"In Progress\")",
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"Complete\")",
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"Blocked\")",
            f"=COUNTIF('{sheet}'!C:C,\"TASK\")",
            f"=D{idx}/F{idx}"
        ]
        
        fills_colors = [STATUS_UNREAD, STATUS_READING, STATUS_REVIEWED, "F8D7DA", "FFFFFF", "FFFFFF"]
        fonts_colors = ["666666", "B45309", "1E7B34", "842029", "000000", "000000"]
        
        for offset, (form, fl, fn) in enumerate(zip(formulas, fills_colors, fonts_colors), start=2):
            cell = ws.cell(row=idx, column=offset, value=form)
            cell.font = Font(name="Calibri", size=12, color=fn)
            cell.fill = fill(fl)
            cell.alignment = align("center", "center")
            cell.border = border()
            if offset == 7:
                cell.font = Font(name="Calibri", size=12, bold=True, color="000000")
                cell.number_format = '0.0%'
                
    # Total (Row 14)
    ws.row_dimensions[14].height = 30
    total_lbl = ws.cell(row=14, column=1, value="TOTAL")
    total_lbl.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    total_lbl.fill = fill(DARK_BLUE)
    total_lbl.alignment = align("left", "center")
    total_lbl.border = border()
    
    total_formulas = [
        "=SUM(B10:B13)",
        "=SUM(C10:C13)",
        "=SUM(D10:D13)",
        "=SUM(E10:E13)",
        "=SUM(F10:F13)",
        "=D14/F14"
    ]
    for offset, form in enumerate(total_formulas, start=2):
        cell = ws.cell(row=14, column=offset, value=form)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(DARK_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        if offset == 7:
            cell.number_format = '0.0%'
            
    # Conditional formatting for G10:G14
    from openpyxl.formatting.rule import CellIsRule
    rule_green = CellIsRule(operator='greaterThanOrEqual', formula=['0.8'], stopIfTrue=True,
                            fill=fill("D4EDDA"),
                            font=Font(name="Calibri", size=12, bold=True, color="1E7B34"))
    rule_amber = CellIsRule(operator='between', formula=['0.4', '0.79'], stopIfTrue=True,
                            fill=fill("FFF3CD"),
                            font=Font(name="Calibri", size=12, bold=True, color="B45309"))
    rule_red = CellIsRule(operator='lessThan', formula=['0.4'], stopIfTrue=True,
                          fill=fill("F8D7DA"),
                          font=Font(name="Calibri", size=12, bold=True, color="842029"))
    ws.conditional_formatting.add('G10:G14', rule_green)
    ws.conditional_formatting.add('G10:G14', rule_amber)
    ws.conditional_formatting.add('G10:G14', rule_red)
    
    # Spacer (Row 15)
    ws.row_dimensions[15].height = 6
    
    # Zone 3 Header (Row 16)
    ws.merge_cells("A16:G16")
    z3 = ws["A16"]
    z3.value = "UPCOMING MILESTONES (next 7 days)"
    z3.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z3.fill = fill(DARK_BLUE)
    z3.alignment = align("left", "center")
    ws.row_dimensions[16].height = 30
    
    # Zone 3 headers (Row 17)
    z3_headers = ["Goal", "Milestone ID", "Milestone Name", "Due Date", "Status", "", ""]
    ws.row_dimensions[17].height = 22
    for col_idx in range(1, 8):
        val = z3_headers[col_idx-1]
        cell = ws.cell(row=17, column=col_idx, value=val if val else None)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        
    # Static data rows for Upcoming Milestones
    milestones_data = [
        ("Approach 1", "1.M2", "Python Custom Rules", "2026-06-11", "In Progress"),
        ("Approach 1", "1.M3", "Java Custom Rules", "2026-06-13", "Not Started"),
        ("Approach 1", "1.M4", "Test Codebase", "2026-06-16", "Not Started"),
        ("Research", "R.M1", "Literature Foundation", "2026-06-20", "In Progress"),
        ("Approach 1", "1.M5", "Demo Preparation", "2026-06-17", "Not Started")
    ]
    for row_offset, (goal, m_id, name, due, status) in enumerate(milestones_data, start=18):
        ws.row_dimensions[row_offset].height = 30
        is_even = (row_offset % 2 == 0)
        base_fill = "D6E4F0" if is_even else "F5F5F5"
        
        row_vals = [goal, m_id, name, due, status]
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_offset, column=col_idx, value=row_vals[col_idx-1])
            cell.font = Font(name="Calibri", size=12, color="000000")
            cell.fill = fill(base_fill)
            cell.alignment = align("center", "center")
            cell.border = border()
            if col_idx == 3:
                cell.alignment = align("left", "center")
                
        # Status styling
        status_cell = ws.cell(row=row_offset, column=5)
        if status == "Complete":
            status_fill, status_color = "D4EDDA", "1E7B34"
        elif status == "In Progress":
            status_fill, status_color = "FFF3CD", "B45309"
        else:
            status_fill, status_color = "F5F5F5", "666666"
        status_cell.fill = fill(status_fill)
        status_cell.font = Font(name="Calibri", size=12, bold=True, color=status_color)
        
        # Empty cells for columns F-G
        for col_idx in [6, 7]:
            cell = ws.cell(row=row_offset, column=col_idx)
            cell.border = border()
            
    # Spacer (Row 23)
    ws.row_dimensions[23].height = 6
    
    # Zone 3b Header (Row 24)
    ws.merge_cells("A24:G24")
    z3b = ws["A24"]
    z3b.value = "KEY CONSTRAINTS"
    z3b.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z3b.fill = fill(DARK_BLUE)
    z3b.alignment = align("left", "center")
    ws.row_dimensions[24].height = 30
    
    # Zone 3b headers (Row 25)
    z3b_headers = ["ID", "Constraint", "Impact", "Buffer Added", "", "", ""]
    ws.row_dimensions[25].height = 22
    for col_idx in range(1, 8):
        val = z3b_headers[col_idx-1]
        cell = ws.cell(row=25, column=col_idx, value=val if val else None)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        
    summary_constraints = [
        ("C-01", "No prior Go experience", "High", "+20% ML on Approach 2/3 Go tasks"),
        ("C-03", "Approach 1 deadline Jun 20 — cannot slip", "High", "No buffer; fixed date"),
        ("C-06", "Ollama model download 4–8 GB", "Medium", "0.5–1 day setup before M6 LLM tasks"),
        ("C-07", "Docker Desktop required for Approach 3", "Medium", "0.5 day setup buffer in 3.M4"),
        ("C-09", "Mentor review cycle 2–3 days", "Medium", "Explicit buffer rows in each approach")
    ]
    for row_offset, (c_id, desc, impact, buf) in enumerate(summary_constraints, start=26):
        ws.row_dimensions[row_offset].height = 30
        
        if impact == "High":
            row_fill, row_color, is_bold = "F8D7DA", "842029", True
        elif impact == "Medium":
            row_fill, row_color, is_bold = "FFF3CD", "B45309", False
        else:
            row_fill, row_color, is_bold = "EAF4EA", "1E5924", False
            
        row_vals = [c_id, desc, impact, buf]
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_offset, column=col_idx, value=row_vals[col_idx-1])
            cell.font = Font(name="Calibri", size=12, bold=is_bold, color=row_color)
            cell.fill = fill(row_fill)
            cell.border = border()
            if col_idx == 2:
                cell.alignment = align("left", "center")
            else:
                cell.alignment = align("center", "center")
                
        # Empty cells for columns E-G
        for col_idx in [5, 6, 7]:
            cell = ws.cell(row=row_offset, column=col_idx)
            cell.border = border()
            
    # Set widths
    col_widths = [35, 45, 30, 35, 16, 16, 16]
    for c_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w
        
    ws.print_area = "A1:G30"
    ws.freeze_panes = "A7"

def build_approach_sheet(wb, title, subtitle, sheet_name, goal_id, milestones):
    ws = wb.create_sheet(title=sheet_name)
    
    # Title Row (Row 1)
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    title_cell.fill = fill(DARK_BLUE)
    title_cell.alignment = align("center", "center")
    ws.row_dimensions[1].height = 50
    
    # Subtitle Row (Row 2)
    ws.merge_cells("A2:M2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitle
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    sub_cell.fill = fill(MID_BLUE)
    sub_cell.alignment = align("center", "center")
    ws.row_dimensions[2].height = 22
    
    # Spacer Row (Row 3)
    ws.row_dimensions[3].height = 6
    
    # Headers Row (Row 4)
    headers = [
        "ID", "Name", "Type", "Start Date", "End Date",
        "PERT O (h)", "PERT ML (h)", "PERT P (h)", "PERT E (h)",
        "Actual Hrs", "Status", "Owner", "Notes"
    ]
    col_widths = [14, 52, 13, 13, 13, 11, 11, 11, 11, 12, 17, 10, 38]
    
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[4].height = 50
    
    row_num = 5
    for m_idx, m in enumerate(milestones):
        is_bonus = "Bonus" in m["name"]
        
        # Write milestone row
        ws.row_dimensions[row_num].height = 40
        m_vals = [
            m["id"], m["name"], m["type"], m["start"], m["end"],
            m["o"], m["ml"], m["p"], m["e"],
            "", m["status"], "", m["notes"]
        ]
        
        # Style row
        if m["type"] == "BUFFER":
            m_fill = "FFF9E6"
            m_font = Font(name="Calibri", size=12, italic=True, color="5C4B00")
        elif is_bonus:
            m_fill = "EAF4EA"
            m_font = Font(name="Calibri", size=12, bold=True, color="1E5924")
        else:
            m_fill = MID_BLUE
            m_font = Font(name="Calibri", size=12, bold=True, color=WHITE)
            
        for c_idx, val in enumerate(m_vals, start=1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.fill = fill(m_fill)
            cell.font = m_font
            cell.border = border()
            
            # Alignments
            if c_idx == 2 or c_idx == 13:
                cell.alignment = align("left", "center")
            else:
                cell.alignment = align("center", "center")
                
            # Number formats
            if c_idx in [6, 7, 8]:
                cell.number_format = '0.0'
            elif c_idx == 9:
                cell.number_format = '0.00'
                
        row_num += 1
        
        # Write tasks
        for t in m.get("tasks", []):
            ws.row_dimensions[row_num].height = 45
            t_vals = [
                t["id"], "    " + t["name"], "TASK", "", "",
                t["o"], t["ml"], t["p"], t["e"],
                "", t["status"], "Hoang", t["notes"]
            ]
            
            is_even = (row_num % 2 == 0)
            
            # Row coloring based on task status
            if t["status"] == "Complete":
                t_fill = "D4EDDA"
                t_font = Font(name="Calibri", size=12, color="1E7B34")
            elif t["status"] == "In Progress":
                t_fill = "FFF3CD"
                t_font = Font(name="Calibri", size=12, color="B45309")
            else:
                t_fill = "D6E4F0" if is_even else "F5F5F5"
                t_font = Font(name="Calibri", size=12, color="000000")
                
            for c_idx, val in enumerate(t_vals, start=1):
                cell = ws.cell(row=row_num, column=c_idx, value=val)
                cell.fill = fill(t_fill)
                cell.font = t_font
                cell.border = border()
                
                # Alignments
                if c_idx == 2 or c_idx == 13:
                    cell.alignment = align("left", "center")
                else:
                    cell.alignment = align("center", "center")
                    
                # Number formats
                if c_idx in [6, 7, 8]:
                    cell.number_format = '0.0'
                elif c_idx == 9:
                    cell.number_format = '0.00'
            
            # Status cell specific styling for Blocked, At Risk, Not Started, Complete, In Progress
            status_cell = ws.cell(row=row_num, column=11)
            if t["status"] == "Complete":
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="1E7B34")
            elif t["status"] == "In Progress":
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="B45309")
            elif t["status"] == "Blocked":
                status_cell.fill = fill("F8D7DA")
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="842029")
            elif t["status"] == "At Risk":
                status_cell.fill = fill("FFE5B4")
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="8B4513")
            elif t["status"] == "Not Started":
                status_cell.fill = fill("F5F5F5")
                status_cell.font = Font(name="Calibri", size=12, color="666666")
            
            # ID is bold and centered
            ws.cell(row=row_num, column=1).font = Font(name="Calibri", size=12, bold=True, color=t_font.color)
            
            row_num += 1
            
        # Add between-milestone spacer (row height 5px) if it is not the last milestone
        if m_idx < len(milestones) - 1:
            ws.row_dimensions[row_num].height = 5
            row_num += 1
            
    ws.print_area = f"A1:M{row_num-1}"
    ws.freeze_panes = "A5"

def build_constraints(wb, data):
    ws = wb.create_sheet(title="Constraints")
    
    # Header logic
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "ZeroTrust.sh  ·  Planning Constraints Register"
    title_cell.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    title_cell.fill = fill(DARK_BLUE)
    title_cell.alignment = align("center", "center")
    ws.row_dimensions[1].height = 50
    
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Last updated: 2026-06-11"
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    sub_cell.fill = fill(MID_BLUE)
    sub_cell.alignment = align("center", "center")
    ws.row_dimensions[2].height = 22
    
    ws.row_dimensions[3].height = 6 # Spacer
    
    headers = ["ID", "Constraint", "Category", "Impact", "Applied To", "Buffer Added", "Notes"]
    col_widths = [8, 55, 14, 10, 30, 20, 50]
    
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[4].height = 50
    
    row_num = 5
    for c in data:
        ws.row_dimensions[row_num].height = 60
        
        impact = c["impact"]
        if impact == "High":
            row_fill = "F8D7DA"
            row_font = Font(name="Calibri", size=12, bold=True, color="842029")
        elif impact == "Medium":
            row_fill = "FFF3CD"
            row_font = Font(name="Calibri", size=12, color="B45309")
        else: # Low
            row_fill = "EAF4EA"
            row_font = Font(name="Calibri", size=12, color="1E5924")
            
        values = [c["id"], c["constraint"], c["category"], c["impact"], c["applied_to"], c["buffer_added"], c["notes"]]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill(row_fill)
            cell.font = row_font
            cell.border = border()
            
            if col_idx in [1, 3, 4]:
                cell.alignment = align("center", "center")
            else:
                cell.alignment = align("left", "center")
        row_num += 1
        
    ws.print_area = f"A1:G{row_num-1}"
    ws.freeze_panes = "A5"

def build_research_papers(wb, papers):
    ws2 = wb.create_sheet(title="Research Papers")
    NUM_COLS = 11
    LAST_COL = get_column_letter(NUM_COLS)   # "K"
    
    ws2.merge_cells(f"A1:{LAST_COL}1")
    t2 = ws2["A1"]
    t2.value = "ZeroTrust.sh  ·  Research Paper Manager"
    t2.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t2.fill = fill(DARK_BLUE)
    t2.alignment = align("center", "center")
    t2.border = border(DARK_BLUE)
    ws2.row_dimensions[1].height = 30
    
    ws2.merge_cells(f"A2:{LAST_COL}2")
    s2 = ws2["A2"]
    s2.value = (
        "Smart Paper Manager  ·  Use column dropdowns (▾) to filter by Category, Tags, Read Status  "
        "·  Click any column header to sort  ·  Write your notes in column K"
    )
    s2.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
    s2.fill = fill(MID_BLUE)
    s2.alignment = align("center", "center")
    s2.border = border(MID_BLUE)
    ws2.row_dimensions[2].height = 18
    
    ws2.row_dimensions[3].height = 6
    
    headers2 = [
        "#", "Title", "Authors", "Year", "Venue",
        "Category", "Tags", "Read Status",
        "Relevance to ZeroTrust.sh", "URL", "Literature Review Notes"
    ]
    col_widths2 = [5, 56, 22, 6, 24, 26, 38, 14, 50, 40, 62]
    
    for ci, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        cell = ws2.cell(row=4, column=ci, value=h)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center", wrap=False)
        cell.border = border(MID_BLUE)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[4].height = 22
    
    current_row = 5
    last_paper_row = current_row
    
    for entry in papers:
        (num, title, authors, year, venue, category, tags,
         read_status, relevance, url, lit_notes) = entry
         
        is_even = (current_row % 2 == 0)
        base_fill_color = LIGHT_BLUE if is_even else LIGHT_GRAY
        
        status_color_map = {
            "Unread":   STATUS_UNREAD,
            "Reading":  STATUS_READING,
            "Read":     STATUS_READ,
            "Reviewed": STATUS_REVIEWED,
        }
        
        values = [num, title, authors, year, venue, category, tags,
                  read_status, relevance, url, lit_notes]
                  
        for ci, value in enumerate(values, start=1):
            cell = ws2.cell(row=current_row, column=ci, value=value)
            cell_fill = base_fill_color
            if ci == 8:
                cell_fill = status_color_map.get(read_status, STATUS_UNREAD)
            if ci == 11:
                cell_fill = "FFFEF5" if not is_even else "F0F8FF"
            cell.fill = fill(cell_fill)
            cell.border = border()
            cell.alignment = align("left", "center", wrap=True)
            cell.font = Font(size=10, name="Calibri")
            
            if ci == 1:
                cell.alignment = align("center", "center", wrap=False)
                cell.font = Font(bold=True, size=10, name="Calibri")
            if ci == 4:
                cell.alignment = align("center", "center", wrap=False)
            if ci == 8:
                cell.alignment = align("center", "center", wrap=False)
                cell.font = Font(bold=True, size=9, name="Calibri")
            if ci == 10:
                cell.font = Font(size=9, color="2E5FA3", name="Calibri", underline="single")
            if ci == 11:
                if not value:
                    cell.font = Font(size=9, color="AAAAAA", italic=True, name="Calibri")
                    cell.value = "Write your notes here…"
                    
        ws2.row_dimensions[current_row].height = 60
        last_paper_row = current_row
        current_row += 1
        
    ws2.auto_filter.ref = f"A4:{LAST_COL}{last_paper_row}"
    
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Unread,Reading,Read,Reviewed"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error='Choose: Unread, Reading, Read, or Reviewed',
    )
    ws2.add_data_validation(dv)
    dv.add(f"H5:H{last_paper_row}")
    
    legend2_row = last_paper_row + 2
    ws2.merge_cells(f"A{legend2_row}:{LAST_COL}{legend2_row}")
    leg2 = ws2[f"A{legend2_row}"]
    leg2.value = (
        "Read Status:  Unread (gray) — not yet read  ·  Reading (amber) — in progress  "
        "·  Read (blue) — finished  ·  Reviewed (green) — literature review written  "
        "·  Filter tip: click ▾ on Category or Tags header to filter by research area"
    )
    leg2.font = Font(italic=True, color="666666", size=9, name="Calibri")
    leg2.alignment = align("left", "center")
    
    ws2.print_area = f"A1:{LAST_COL}{last_paper_row}"
    ws2.freeze_panes = "B5"


papers = [
    # ── AREA 1 — Deep Learning & ML for Vulnerability Detection ───────────────
    (1,
     "Automated Vulnerability Detection in Source Code Using Deep Representation Learning",
     "Feng et al.", "2026", "arXiv",
     "ML/DL Detection",
     "CNN, deep representation learning, source code, vulnerability classifier",
     "Unread",
     "CNN-based deep representation learning for vulnerability detection — validates ML classifier gate in Path B Tier 2",
     "https://arxiv.org/abs/2602.23121", ""),
    (2,
     "DiverseVul: A New Vulnerable Source Code Dataset for Deep Learning Based Vulnerability Detection",
     "Jia et al.", "2023", "RAID 2023 / ACM",
     "ML/DL Detection",
     "dataset, C/C++, CWE, BigVul alternative, training data, benchmark, 18945 functions",
     "Unread",
     "Largest diverse C/C++ vulnerability dataset (18,945 functions, 150 CWEs) — training data source for Code Vulnerability Classifier",
     "https://arxiv.org/abs/2304.00409", ""),
    (3,
     "Vulnerability Detection in C/C++ Code with Deep Learning",
     "Multiple authors", "2024", "arXiv",
     "ML/DL Detection",
     "neural network, program slicing, C/C++, deep learning",
     "Unread",
     "Neural networks with program slices for vulnerability detection — informs Tier 2 classifier design",
     "https://arxiv.org/abs/2405.12384", ""),
    (4,
     "Deep Learning Aided Software Vulnerability Detection: A Survey",
     "Survey authors", "2025", "arXiv",
     "ML/DL Detection",
     "survey, deep learning, vulnerability detection, baseline comparison",
     "Unread",
     "Comprehensive DL survey for vulnerability detection — baseline reference for Tier 2 classifier selection",
     "https://arxiv.org/abs/2503.04002", ""),

    # ── AREA 2 — Graph Neural Networks for Vulnerability Detection ────────────
    (5,
     "Software Vulnerability Detection Using a Lightweight Graph Neural Network (VulGNN)",
     "Zhu et al.", "2026", "arXiv",
     "Graph Neural Networks",
     "GNN, lightweight, LLM-parity, local inference, classifier, small model",
     "Unread",
     "Lightweight GNN achieving LLM-parity at 100x smaller size — validates cheap local classifier concept in new architecture",
     "https://arxiv.org/abs/2603.29216", ""),
    (6,
     "Vul-LMGNNs: Fusing Language Models and Graph Neural Networks for Code Vulnerability Detection",
     "Rong et al.", "2024", "arXiv",
     "Graph Neural Networks",
     "GNN, language model, hybrid, code representation, code LM",
     "Unread",
     "Hybrid code LM + GNN — informs Call Graph + Classifier integration in Path B Tier 2",
     "https://arxiv.org/abs/2404.14719", ""),
    (7,
     "Structure-Aware Code Vulnerability Analysis With Graph Neural Networks",
     "Allamanis et al.", "2024", "arXiv",
     "Graph Neural Networks",
     "GNN, Java, structure-aware, vulnerability fixing, commit-level",
     "Unread",
     "GNN-based vulnerability analysis using Java vulnerability-fixing commits — informs structure-aware detection",
     "https://arxiv.org/abs/2307.11454", ""),
    (8,
     "Graph Neural Networks for Vulnerability Detection: A Counterfactual Explanation",
     "Li et al.", "2024", "arXiv",
     "Graph Neural Networks",
     "GNN, explainability, counterfactual, confidence scoring, interpretability",
     "Unread",
     "Explainability analysis of GNN detection — informs confidence scoring in Dedup layer",
     "https://arxiv.org/abs/2404.15687", ""),
    (9,
     "ReGVD: Revisiting Graph Neural Networks for Vulnerability Detection",
     "Nguyen et al.", "2022", "ACM/IEEE",
     "Graph Neural Networks",
     "GNN, token sequence, baseline, benchmark, foundational",
     "Unread",
     "Foundational GNN model treating source code as flat token sequences — baseline for Tier 2 classifier",
     "https://dl.acm.org/doi/abs/10.1145/3510454.3516865", ""),
    (10,
     "LineVD: Statement-level Vulnerability Detection using Graph Neural Networks",
     "Chen et al.", "2022", "arXiv",
     "Graph Neural Networks",
     "GNN, statement-level, line localization, fine-grained, HTML report",
     "Unread",
     "Fine-grained GNN vulnerability localization — informs line-level finding output in HTML report",
     "https://arxiv.org/abs/2203.05181", ""),

    # ── AREA 3 — LLM for Code Security Analysis ───────────────────────────────
    (11,
     "LLMs in Code Vulnerability Analysis: A Proof of Concept",
     "Kochling et al.", "2026", "arXiv",
     "LLM Code Analysis",
     "LLM, empirical, proof-of-concept, zero-shot, Path B validation",
     "Unread",
     "Empirical PoC for LLM vulnerability analysis — validates LLM Semantic Scan role in Path B Tier 3",
     "https://arxiv.org/abs/2601.08691", ""),
    (12,
     "IRIS: LLM-Assisted Static Analysis for Detecting Security Vulnerabilities",
     "Scanlon et al.", "2024", "arXiv",
     "LLM Code Analysis",
     "hybrid SAST+LLM, false positive, 80% FP reduction, IRIS, key paper",
     "Unread",
     "Hybrid SAST+LLM detecting 55/120 vulns + 6 new, reducing FP by 80% — directly validates ZeroTrust.sh hybrid design",
     "https://arxiv.org/abs/2405.17238", ""),
    (13,
     "Large Language Model for Vulnerability Detection and Repair: Literature Review and the Road Ahead",
     "Zhang et al.", "2025", "ACM TOSEM",
     "LLM Code Analysis",
     "survey, LLM, vulnerability repair, patch generation, comprehensive review",
     "Unread",
     "Comprehensive LLM vulnerability + repair survey — informs patch generation design across all approaches",
     "https://dl.acm.org/doi/10.1145/3708522", ""),
    (14,
     "Understanding the Effectiveness of LLMs in Detecting Security Vulnerabilities",
     "Steenhoek et al.", "2023", "arXiv",
     "LLM Code Analysis",
     "LLM, prompting strategy, systematic evaluation, false positive, prompt design",
     "Unread",
     "Systematic LLM evaluation with prompting strategy analysis — informs LLM Verifier prompt design in Path A",
     "https://arxiv.org/abs/2311.16169", ""),
    (15,
     "Large Language Models for Source Code Analysis: Applications, Models and Datasets",
     "Sharma et al.", "2025", "arXiv",
     "LLM Code Analysis",
     "survey, LLM architectures, code analysis, model selection, datasets",
     "Unread",
     "Survey of LLM architectures for code analysis — model selection reference for LLM components",
     "https://arxiv.org/abs/2503.17502", ""),
    (16,
     "Can Large Language Models Find And Fix Vulnerable Software?",
     "Pearce et al.", "2023", "arXiv",
     "LLM Code Analysis",
     "LLM, detection, repair, empirical evaluation, fix generation",
     "Unread",
     "Empirical evaluation of LLM detection + repair — validates dual-role LLM use in Approaches 2 and 3",
     "https://arxiv.org/abs/2308.10345", ""),

    # ── AREA 4 — Hybrid Static Analysis + LLM ─────────────────────────────────
    (17,
     "LLM-Driven SAST-Genius: A Hybrid Static Analysis Framework for Comprehensive and Actionable Security",
     "Multiple authors", "2024", "arXiv",
     "Hybrid SAST+LLM",
     "hybrid, SAST+LLM, false positive, 91% FP reduction, Semgrep, key paper",
     "Unread",
     "Hybrid SAST+LLM reducing FP by 91% (225→20 alerts) vs Semgrep alone — strongest validation of ZeroTrust.sh architecture",
     "https://arxiv.org/abs/2509.15433", ""),
    (18,
     "ZeroFalse: Improving Precision in Static Analysis with LLMs",
     "Scanlon et al.", "2024", "arXiv",
     "Hybrid SAST+LLM",
     "LLM, false positive reduction, static analysis, precision, ZeroFalse",
     "Unread",
     "LLM false positive reduction in static analysis — validates LLM Verifier design in Path A",
     "https://arxiv.org/abs/2510.02534", ""),
    (19,
     "Combining Large Language Models with Static Analyzers for Code Review Generation",
     "Jaoua et al.", "2025", "arXiv",
     "Hybrid SAST+LLM",
     "LLM, static analysis, code review, patch suggestion, output format",
     "Unread",
     "LLM + static analysis for code review — informs patch suggestion output format",
     "https://arxiv.org/abs/2502.06633", ""),
    (20,
     "A Contemporary Survey of LLM-Assisted Program Analysis",
     "Survey authors", "2025", "arXiv",
     "Hybrid SAST+LLM",
     "survey, program analysis, LLM, comprehensive, all approaches reference",
     "Unread",
     "Comprehensive survey of LLM program analysis techniques — architecture reference for all three approaches",
     "https://arxiv.org/abs/2502.18474", ""),
    (21,
     "RepoAudit: An Autonomous LLM-Agent for Repository-Level Code Auditing",
     "Li et al.", "2025", "arXiv",
     "Hybrid SAST+LLM",
     "LLM agent, repository-level, multi-agent, orchestration, Approach 3",
     "Unread",
     "LLM-agent for repo-level auditing — informs Approach 3 multi-agent orchestration design",
     "https://arxiv.org/abs/2501.18160", ""),

    # ── AREA 5 — AI-Generated Code Security & Prompt Injection ────────────────
    (22,
     "Security Vulnerabilities in AI-Generated Code: A Large-Scale Analysis of Public GitHub Repositories",
     "Zhao et al.", "2024", "IEEE/ACM",
     "AI-Generated Code Security",
     "AI-generated code, CWE, GitHub, empirical, 4241 instances, market validation",
     "Unread",
     "4,241 CWE instances across AI-generated code from 4 tools — empirical validation of the core ZeroTrust.sh problem statement",
     "https://arxiv.org/abs/2510.26103", ""),
    (23,
     "Prompt Injection Attacks on Agentic Coding Assistants: A Systematic Analysis",
     "Yang et al.", "2026", "arXiv",
     "AI Agent Security",
     "prompt injection, agentic, coding assistant, attack, 85% success rate",
     "Unread",
     "85%+ attack success rates for prompt injection in agentic assistants — validates ZeroTrust.sh AI-specific threat detection",
     "https://arxiv.org/abs/2601.17548", ""),
    (24,
     "Security Degradation in Iterative AI Code Generation: A Systematic Analysis of the Paradox",
     "Multiple authors", "2025", "IEEE-ISTAS",
     "AI-Generated Code Security",
     "iterative generation, vibe-coding, security degradation, LLM, threat model",
     "Unread",
     "Iterative LLM interactions without human review introduce new vulnerabilities — validates vibe-coding threat model",
     "https://arxiv.org/abs/2506.11022", ""),
    (25,
     "Assessing the Quality and Security of AI-Generated Code: A Quantitative Analysis",
     "Multiple authors", "2024", "arXiv",
     "AI-Generated Code Security",
     "AI-generated code, quantitative, security assessment, quality metrics",
     "Unread",
     "Quantitative security analysis of AI-generated code — market validation and threat taxonomy reference",
     "https://arxiv.org/abs/2508.14727", ""),
    (26,
     "You Still Have to Study: On the Security of LLM Generated Code",
     "Ferrara et al.", "2024", "arXiv",
     "AI-Generated Code Security",
     "Copilot, CWE, 36-40% vulnerability rate, LLM-generated code, pitch statistic",
     "Unread",
     "36–40% of Copilot code contains CWE vulnerabilities — key statistic for product positioning and pitch",
     "https://arxiv.org/abs/2408.07106", ""),

    # ── AREA 6 — Token Cost Optimization for LLM Pipelines ────────────────────
    (27,
     "FrugalGPT: How to Use Large Language Models While Reducing Cost and Improving Performance",
     "Eisingerich et al.", "2023", "arXiv",
     "Token Cost Optimization",
     "cascade routing, cost optimization, 98% savings, LLM pipeline, FrugalGPT",
     "Unread",
     "Cascade routing achieving 98% cost savings — theoretical foundation for Cascading Intelligence architecture",
     "https://arxiv.org/abs/2305.05176", ""),
    (28,
     "Batch Prompting: Efficient Inference with Large Language Model APIs",
     "Rajkumar et al.", "2023", "arXiv",
     "Token Cost Optimization",
     "batch processing, token cost, 5x reduction, inference efficiency",
     "Unread",
     "Batch processing reducing LLM token costs up to 5x — informs Token Budget Controller batching strategy",
     "https://arxiv.org/abs/2301.08721", ""),
    (29,
     "Token Sugar: Making Source Code Sweeter for LLMs through Token-Efficient Shorthand",
     "Multiple authors", "2025", "arXiv",
     "Token Cost Optimization",
     "token optimization, code representation, shorthand, context compression",
     "Unread",
     "Token optimization for code representation — informs context chunking in Token Budget Controller",
     "https://arxiv.org/abs/2512.08266", ""),
    (30,
     "Learning to Focus: Context Extraction for Efficient Code Vulnerability Detection with Language Models",
     "Dittmann et al.", "2025", "arXiv",
     "Token Cost Optimization",
     "context filtering, token budget, vulnerability detection, efficiency, focus extraction",
     "Unread",
     "Context filtering for reducing LLM token consumption in vulnerability detection — directly validates Token Budget Controller",
     "https://arxiv.org/abs/2505.17460", ""),

    # ── AREA 7 — Call Graph, Taint Analysis & Code Representations ────────────
    (31,
     "Vulnerability Detection with Interprocedural Context in Multiple Languages: Assessing Effectiveness and Cost",
     "Gharibi et al.", "2024", "arXiv",
     "Call Graph & Taint Analysis",
     "interprocedural, multi-language, call graph, LLM detection, effectiveness",
     "Unread",
     "Interprocedural analysis impact on LLM detection across languages — validates Call Graph + CVE Enrichment in Path B",
     "https://arxiv.org/abs/2604.08417", ""),
    (32,
     "Multi-Agent Taint Specification Extraction for Vulnerability Detection",
     "Zhang et al.", "2026", "arXiv",
     "Call Graph & Taint Analysis",
     "taint analysis, multi-agent, CodeQL, LLM, Approach 3",
     "Unread",
     "Multi-agent LLM + taint analysis — informs Approach 3 multi-agent architecture with CodeQL/Joern integration",
     "https://arxiv.org/abs/2601.10865", ""),
    (33,
     "LLMxCPG: Context-Aware Vulnerability Detection Through Code Property Graph-Guided LLMs",
     "Pan et al.", "2024", "arXiv",
     "Call Graph & Taint Analysis",
     "CPG, code property graph, context-aware, LLM, Path B integration",
     "Unread",
     "CPG-guided LLM for context-aware detection — informs CPG integration in Path B call graph analysis",
     "https://arxiv.org/abs/2507.16585", ""),
    (34,
     "Bridging Code Property Graphs and Language Models for Program Analysis",
     "Mahfouz et al.", "2026", "arXiv",
     "Call Graph & Taint Analysis",
     "CPG, LLM, program analysis, hybrid CPG+LLM, framework",
     "Unread",
     "Framework bridging CPG and LLMs — validates hybrid CPG+LLM design in Path B",
     "https://arxiv.org/abs/2603.24837", ""),
    (35,
     "Enhancing Software Vulnerability Detection Using Code Property Graphs and Convolutional Neural Networks",
     "Multiple authors", "2025", "arXiv",
     "Call Graph & Taint Analysis",
     "CPG, CNN, local structure, global structure, code classifier training",
     "Unread",
     "CPG+CNN for local and global code structure — informs Code Classifier training in new architecture",
     "https://arxiv.org/abs/2503.18175", ""),
    (36,
     "VulTrLM: LLM-Assisted Vulnerability Detection via AST Decomposition and Comment Enhancement",
     "Liu et al.", "2025", "Empirical Software Engineering",
     "Call Graph & Taint Analysis",
     "LLM, AST decomposition, semantic enhancement, comment, AST preprocessing",
     "Unread",
     "LLM-assisted AST decomposition for semantic enhancement — informs AST preprocessing in Path B Tier 1",
     "https://dl.acm.org/doi/10.1007/s10664-025-10738-7", ""),
    (37,
     "Dataflow Analysis-Inspired Deep Learning for Efficient Vulnerability Detection",
     "Cheng et al.", "2022", "arXiv",
     "Call Graph & Taint Analysis",
     "dataflow analysis, deep learning, CodeQL integration, Path A validation",
     "Unread",
     "Dataflow analysis-inspired DL approach — validates dataflow integration in Path A CodeQL/Joern",
     "https://arxiv.org/abs/2212.08108", ""),
    (38,
     "Reducing False Positives in Static Bug Detection with LLMs: An Empirical Study in Industry",
     "Dittmann et al.", "2026", "arXiv",
     "Hybrid SAST+LLM",
     "false positive, LLM, industrial study, production scale, empirical",
     "Unread",
     "Industrial study on LLM false positive reduction — validates LLM Verifier design in Path A at production scale",
     "https://arxiv.org/abs/2601.18844", ""),
    (39,
     "LSAST: Enhancing Cybersecurity through LLM-supported Static Application Security Testing",
     "Multiple authors", "2024", "arXiv",
     "Local LLM Deployment",
     "local LLM, privacy-first, offline, SAST, no cloud API, on-device",
     "Unread",
     "Locally-hostable LLM for SAST without cloud APIs — validates privacy-first local LLM deployment approach",
     "https://arxiv.org/abs/2409.15735", ""),
    (40,
     "Software Vulnerability Analysis Across Programming Language and Program Representation Landscapes: A Survey",
     "Multiple authors", "2025", "arXiv",
     "Call Graph & Taint Analysis",
     "survey, AST, CFG, PDG, CPG, program representation, multi-language",
     "Unread",
     "Survey of AST, CFG, PDG, CPG representations — reference for program representation selection across all approaches",
     "https://arxiv.org/abs/2503.20244", ""),

    # ── AREA 8 — Package Hallucination & Supply Chain ─────────────────────────
    (41,
     "We Have a Package for You! A Comprehensive Analysis of Package Hallucinations by Code Generating LLMs",
     "Spracklen et al.", "2024", "USENIX Security 2025",
     "AI Agent Security",
     "slopsquatting, package hallucination, supply chain, npm, pip, 205474 hallucinations, foundational",
     "Unread",
     "Foundational study: 205,474 hallucinated package names across 16 LLMs — commercial at 5.2%, open-source at 21.7%. Baseline threat ZeroTrust.sh must detect.",
     "https://arxiv.org/abs/2406.10279", ""),
    (42,
     "Importing Phantoms: Measuring LLM Package Hallucination Vulnerabilities",
     "Krishna et al.", "2025", "arXiv:2501.19012",
     "AI Agent Security",
     "slopsquatting, package hallucination, language comparison, model size, Pareto frontier",
     "Unread",
     "Hallucination rates across languages, model sizes, and task specificity — identifies highest-risk agent configurations for slopsquatting detection.",
     "https://arxiv.org/abs/2501.19012", ""),
    (43,
     "The Range Shrinks, the Threat Remains: Re-evaluating LLM Package Hallucinations on the 2026 Frontier-Model Cohort",
     "Multiple authors", "2026", "arXiv:2605.17062",
     "AI Agent Security",
     "slopsquatting, frontier models, Claude, GPT-5, 127 common hallucinations, 2026 benchmark",
     "Unread",
     "Frontier models (Claude Haiku 4.5 at 4.62%, GPT-5.4-mini at 6.10%) still hallucinate; 127 package names all models invent — rich target set for detection rules.",
     "https://arxiv.org/abs/2605.17062", ""),
    (44,
     "PackMonitor: Enabling Zero Package Hallucinations Through Decoding-Time Monitoring",
     "Liu et al.", "2026", "arXiv:2602.20717",
     "AI Agent Security",
     "slopsquatting, decoding-time defense, PyPI, npm, authoritative package list, prevention",
     "Unread",
     "Defense via decoding-time intervention constraining generation to authoritative package lists — informs ZeroTrust.sh's detection and verification strategy.",
     "https://arxiv.org/abs/2602.20717", ""),
    (45,
     "Secure or Suspect? Investigating Package Hallucinations of Shell Command in Original and Quantized LLMs",
     "Haque et al.", "2025", "arXiv:2512.08213",
     "AI Agent Security",
     "slopsquatting, quantized models, GGUF, shell command, realistic URL mimicry",
     "Unread",
     "Quantized models have significantly higher hallucination rates; fabricated packages mimic GitHub/golang.org URLs — relevant for sophisticated slopsquatting detection.",
     "https://arxiv.org/abs/2512.08213", ""),

    # ── AREA 9 — AI Agent Trust, Privilege Escalation & Prompt Injection ──────
    (46,
     "'Your AI, My Shell': Demystifying Prompt Injection Attacks on Agentic AI Coding Editors",
     "Liu et al.", "2025", "arXiv:2509.22040",
     "AI Agent Security",
     "prompt injection, Cursor, repository poisoning, agentic editor, shell execution hijack",
     "Unread",
     "First empirical analysis of Cursor via repository poisoning — attackers inject malicious instructions into dev resources to hijack shell execution. Direct ZeroTrust.sh threat model.",
     "https://arxiv.org/abs/2509.22040", ""),
    (47,
     "The Dark Side of LLMs: Agent-based Attacks for Complete Computer Takeover",
     "Lupinacci et al.", "2025", "arXiv:2507.06850",
     "AI Agent Security",
     "privilege escalation, inter-agent trust, 94.4% vulnerable, agent-to-agent attack, safety bypass",
     "Unread",
     "94.4% of models vulnerable to direct injection; 83.3% to inter-agent trust exploitation — key data for detecting agent-to-agent attack chains in ZeroTrust.sh.",
     "https://arxiv.org/abs/2507.06850", ""),
    (48,
     "VIGIL: Defending LLM Agents Against Tool Stream Injection via Verify-Before-Commit",
     "Lin et al.", "2026", "arXiv:2601.05755",
     "AI Agent Security",
     "tool stream injection, SIREN benchmark, verify-before-commit, runtime poisoning, 959 attack cases",
     "Unread",
     "Defense against tool stream injection; SIREN benchmark (959 attack cases) provides test corpus for ZeroTrust.sh detection validation.",
     "https://arxiv.org/abs/2601.05755", ""),
    (49,
     "Agentic AI Security: Threats, Defenses, Evaluation, and Open Challenges",
     "Datta et al.", "2025", "arXiv:2510.23883",
     "AI Agent Security",
     "survey, agentic AI, threat taxonomy, tool use, memory, autonomy, planning",
     "Unread",
     "Comprehensive survey of agentic AI threat taxonomy distinguishing autonomous execution risks from traditional LLM safety — essential scope reference for ZeroTrust.sh.",
     "https://arxiv.org/abs/2510.23883", ""),
    (50,
     "Are AI-assisted Development Tools Immune to Prompt Injection?",
     "Multiple authors", "2026", "arXiv:2603.21642",
     "AI Agent Security",
     "MCP, tool poisoning, Claude Code, Cursor, Cline, Continue, Gemini CLI, coding tools benchmark",
     "Unread",
     "First empirical analysis of MCP clients (Claude Code, Cursor, Cline, Continue, Gemini CLI) against tool-poisoning attacks — directly relevant to detecting injection in AI tool contexts.",
     "https://arxiv.org/abs/2603.21642", ""),
    (51,
     "Taming Various Privilege Escalation in LLM-Based Agent Systems: A Mandatory Access Control Framework",
     "Multiple authors", "2026", "arXiv:2601.11893",
     "AI Agent Security",
     "privilege escalation, MAC framework, SEAgent, formal model, agent security policy",
     "Unread",
     "Formal model of privilege escalation in LLM agents with SEAgent MAC framework — informs detection rules for trust boundary violations in agentic codebases.",
     "https://arxiv.org/abs/2601.11893", ""),

    # ── AREA 10 — Logic Vulnerability Detection (IDOR, Access Control) ────────
    (52,
     "BacAlarm: Mining and Simulating Composite API Traffic to Prevent Broken Access Control Violations",
     "Multiple authors", "2025", "arXiv:2512.19997",
     "Logic Vuln Detection",
     "broken access control, REST API, OWASP API, LLM agent, RAG, anomaly detection",
     "Unread",
     "Broken access control detection in REST APIs via LLM-based agent traffic simulation and ensemble anomaly detection — core to detecting authorization gaps in Path B.",
     "https://arxiv.org/abs/2512.19997", ""),
    (53,
     "Rethinking Broken Object Level Authorization Attacks Under Zero Trust Principle",
     "Wu et al.", "2025", "arXiv:2507.02309",
     "Logic Vuln Detection",
     "BOLA, IDOR, zero trust, resource ID dataflow, authorization interval, 35 new CVEs",
     "Unread",
     "BOLAZ: zero-trust defense for BOLA/IDOR via resource ID data flow analysis; discovered 35 new vulnerabilities — directly validates Path B's IDOR detection design.",
     "https://arxiv.org/abs/2507.02309", ""),
    (54,
     "VULSOLVER: Vulnerability Detection via LLM-Driven Constraint Solving",
     "Li et al.", "2025", "arXiv:2509.00882",
     "Logic Vuln Detection",
     "LLM, constraint solving, call-chain analysis, 96.29% accuracy, semantic reasoning",
     "Unread",
     "SAST + LLM semantic reasoning with progressive constraint modeling achieves 96.29% accuracy — applicable to detecting contextually-wrong authorization logic.",
     "https://arxiv.org/abs/2509.00882", ""),
    (55,
     "SAVANT: Vulnerability Detection in Application Dependencies through Semantic-Guided Reachability Analysis",
     "Multiple authors", "2025", "arXiv:2506.17798",
     "Logic Vuln Detection",
     "semantic preprocessing, LLM, reachability analysis, library API, 83.8% precision",
     "Unread",
     "Semantic preprocessing + LLM context analysis for vulnerable API patterns; 83.8% precision — relevant to detecting missing authorization checks in library calls.",
     "https://arxiv.org/abs/2506.17798", ""),
    (56,
     "Argus: Reorchestrating Static Analysis via a Multi-Agent Ensemble for Full-Chain Security Vulnerability Detection",
     "Multiple authors", "2025", "arXiv:2604.06633",
     "Logic Vuln Detection",
     "multi-agent, SAST, false positive reduction, multi-hop reasoning, zero-day, full-chain",
     "Unread",
     "First multi-agent LLM framework for SAST with multi-hop reasoning — discovers zero-day vulnerabilities spanning multiple functions; applicable to detecting business logic flaws.",
     "https://arxiv.org/abs/2604.06633", ""),
    (57,
     "Benchmarking LLMs and LLM-based Agents in Practical Vulnerability Detection for Code Repositories",
     "Multiple authors", "2025", "ACL 2025 / arXiv:2503.03586",
     "Logic Vuln Detection",
     "JitVul benchmark, 879 CVEs, interprocedural, ReAct agents, authorization flaws, ACL",
     "Unread",
     "JitVul benchmark (879 CVEs) requires interprocedural analysis; ReAct agents outperform LLMs on auth flaws — suggests agentic approaches for Path B's multi-function context.",
     "https://arxiv.org/abs/2503.03586", ""),
    (58,
     "Prompting the Priorities: A First Look at Evaluating LLMs for Vulnerability Triage and Prioritization",
     "Multiple authors", "2025", "arXiv:2510.18508",
     "Logic Vuln Detection",
     "SSVC, triage, prioritization, 384 real-world CVEs, risk surface selection, LLM evaluation",
     "Unread",
     "LLM evaluation on SSVC triage framework with 384 real-world vulnerabilities — methodology for prioritizing high-risk code surfaces before deep Path B analysis.",
     "https://arxiv.org/abs/2510.18508", ""),
    (59,
     "I Can't Believe It's Not a Valid Exploit",
     "Multiple authors", "2026", "arXiv:2602.04165",
     "Logic Vuln Detection",
     "PoC-Gym, exploit generation, static analysis guidance, 21% improvement, validation limits",
     "Unread",
     "PoC-Gym framework shows static analysis guidance improves LLM exploit generation by 21% — informs heuristics for filtering false-positive authorization scenarios in Path B.",
     "https://arxiv.org/abs/2602.04165", ""),

    # ── AREA 11 — Local LLM Deployment & On-Device Security ──────────────────
    (60,
     "Mind the Gap: A Practical Attack on GGUF Quantization",
     "Egashira et al.", "2025", "ICML 2025 / arXiv:2505.23786",
     "Local LLM Deployment",
     "GGUF, quantization, backdoor injection, Ollama, llama.cpp, supply chain, ICML",
     "Unread",
     "Backdoor injection via GGUF quantization errors — essential threat model for ZeroTrust.sh's local LLM deployment: quantized models can exhibit hidden malicious behaviors.",
     "https://arxiv.org/abs/2505.23786", ""),
    (61,
     "Widening the Gap: Exploiting LLM Quantization via Outlier Injection",
     "Shi et al.", "2025", "arXiv:2605.15152",
     "Local LLM Deployment",
     "quantization attack, AWQ, GPTQ, GGUF, outlier injection, adversarial, supply chain",
     "Unread",
     "First quantization-conditioned attack affecting AWQ/GPTQ/GGUF — demonstrates supply-chain risks in locally-deployed quantized models; complements GGUF threat model.",
     "https://arxiv.org/abs/2605.15152", ""),
    (62,
     "A First Look At Efficient And Secure On-Device LLM Inference Against KV Leakage",
     "Yang et al.", "2024", "arXiv:2409.04040",
     "Local LLM Deployment",
     "on-device LLM, KV cache leakage, FHE, TEE, privacy, local inference security",
     "Unread",
     "Privacy attacks on on-device LLM inference via KV cache leakage — demonstrates how local inference can leak conversation history; relevant to ZeroTrust.sh's privacy-first design.",
     "https://arxiv.org/abs/2409.04040", ""),
    (63,
     "A Cost-Benefit Analysis of On-Premise Large Language Model Deployment: Breaking Even with Commercial LLM Services",
     "Pan et al.", "2025", "arXiv:2509.18101",
     "Local LLM Deployment",
     "on-premise, cost analysis, breakeven, Qwen, Llama, Mistral, cloud vs local, value proposition",
     "Unread",
     "Quantifies breakeven for on-premise LLM vs cloud APIs — supports ZeroTrust.sh's value proposition: local deployment becomes cost-competitive at scale.",
     "https://arxiv.org/abs/2509.18101", ""),

    # ── AREA 13 — Prompt Compression & Context Reduction ─────────────────────
    (69,
     "LLMLingua: Compressing Prompts for Accelerated Inference of Large Language Models",
     "Jiang et al.", "2023", "arXiv:2310.05736",
     "Prompt Compression",
     "prompt compression, token pruning, 20x compression, budget control, coarse-to-fine",
     "Unread",
     "Iterative token-level compression with budget control; up to 20x compression with minimal loss — directly applicable to compressing code context before LLM security scans.",
     "https://arxiv.org/abs/2310.05736", ""),
    (70,
     "LLMLingua-2: Data Distillation for Efficient and Faithful Task-Agnostic Prompt Compression",
     "Jiang et al.", "2024", "arXiv:2403.12968",
     "Prompt Compression",
     "prompt compression, data distillation, token classification, 3-6x faster, 1.6-2.9x latency",
     "Unread",
     "Compression as token classification via GPT-4 distillation; 3–6x faster than LLMLingua with 2–5x compression ratios — ideal for repeated scans of similar code patterns.",
     "https://arxiv.org/abs/2403.12968", ""),
    (71,
     "LongLLMLingua: Accelerating and Enhancing LLMs in Long Context Scenarios via Prompt Compression",
     "Jiang et al.", "2023", "arXiv:2310.06839",
     "Prompt Compression",
     "long context, question-aware compression, document reordering, 94% cost reduction, 4x fewer tokens",
     "Unread",
     "Question-aware compression for long-context scenarios; 94% cost reduction on LooGLE — critical for scanning large codebases and enriching prompts with vulnerability databases.",
     "https://arxiv.org/abs/2310.06839", ""),
    (72,
     "RECOMP: Improving Retrieval-Augmented LMs with Compression and Selective Augmentation",
     "Xu et al.", "2023", "arXiv:2310.04408",
     "Prompt Compression",
     "RAG, retrieval augmentation, extractive compression, abstractive compression, 6% compression rate",
     "Unread",
     "Compresses retrieved documents in RAG pipelines via extractive and abstractive methods — relevant when ZeroTrust.sh augments prompts with retrieved CVE or security knowledge.",
     "https://arxiv.org/abs/2310.04408", ""),
    (73,
     "Pruning the Unsurprising: Efficient LLM Reasoning via First-Token Surprisal",
     "Multiple authors", "2025", "arXiv:2508.05988",
     "Prompt Compression",
     "CoT compression, first-token surprisal, 23.5% token reduction, 43.5% latency reduction, ASAP",
     "Unread",
     "ASAP framework compresses chain-of-thought by 23.5% tokens and 43.5% latency — optimizes reasoning chains in vulnerability analysis without sacrificing accuracy.",
     "https://arxiv.org/abs/2508.05988", ""),

    # ── AREA 14 — Prompt Engineering & Automatic Optimization ─────────────────
    (74,
     "Large Language Models as Optimizers (OPRO)",
     "Yang et al.", "2023", "arXiv:2309.03409",
     "Prompt Engineering",
     "automatic prompt optimization, OPRO, LLM-based refinement, in-context learning, 50% improvement",
     "Unread",
     "Automatic prompt optimization via LLM-based iterative refinement; best prompts beat human-designed by up to 50% — applicable to auto-tuning ZeroTrust.sh security scanning prompts.",
     "https://arxiv.org/abs/2309.03409", ""),
    (75,
     "Chain of Draft: Thinking Faster by Writing Less",
     "Xu et al.", "2025", "arXiv:2502.18600",
     "Prompt Engineering",
     "chain of draft, CoT efficiency, 7.6% tokens, reasoning cost reduction, latency",
     "Unread",
     "Reduces CoT verbosity to 7.6% of tokens while matching CoT accuracy — practical cost/latency reduction for reasoning-heavy vulnerability analysis in Path B.",
     "https://arxiv.org/abs/2502.18600", ""),
    (76,
     "Which Examples to Annotate for In-Context Learning? Towards Effective and Efficient Selection",
     "Mavromatis et al.", "2023", "arXiv:2310.20046",
     "Prompt Engineering",
     "few-shot selection, in-context learning, budget constraint, uncertainty sampling, diversity, AdaICL",
     "Unread",
     "AdaICL selects maximally informative few-shot examples under token budget constraints — applicable to selecting the best security examples per scan without exceeding token limits.",
     "https://arxiv.org/abs/2310.20046", ""),
    (77,
     "DecoPrompt: Decoding Prompts Reduces Hallucinations when Large Language Models Meet False Premises",
     "Xu and Ma", "2024", "arXiv:2411.07457",
     "Prompt Engineering",
     "hallucination reduction, false premises, decoding prompts, inference-time, no retraining",
     "Unread",
     "Inference-time hallucination reduction by decoding the prompt itself — cost-effective false positive reduction for ZeroTrust.sh without model retraining.",
     "https://arxiv.org/abs/2411.07457", ""),

    # ── AREA 15 — Inference Speed & KV-Cache Optimization ─────────────────────
    (78,
     "Unlocking Efficiency in Large Language Model Inference: A Comprehensive Survey of Speculative Decoding",
     "Multiple authors", "2024", "arXiv:2401.07851",
     "Inference Speed",
     "speculative decoding, draft model, parallel verification, inference acceleration, survey",
     "Unread",
     "Comprehensive survey of speculative decoding (small draft model + large verifier) for LLM inference acceleration — core technique for optimizing local Ollama/llama.cpp execution speed.",
     "https://arxiv.org/abs/2401.07851", ""),
    (79,
     "Keep the Cost Down: A Review on Methods to Optimize LLM's KV-Cache Consumption",
     "Multiple authors", "2024", "arXiv:2407.18003",
     "Inference Speed",
     "KV-cache, PagedAttention, vLLM, token pruning, VRAM, memory optimization, survey",
     "Unread",
     "Survey of KV-cache optimization strategies (PagedAttention, distributed KV-cache, token pruning) — reduces VRAM footprint for quantized GGUF models in ZeroTrust.sh local inference.",
     "https://arxiv.org/abs/2407.18003", ""),
    (80,
     "FlashAttention: Fast and Memory-Efficient Exact Attention with IO-Awareness",
     "Dao et al.", "2022", "arXiv:2205.14135",
     "Inference Speed",
     "flash attention, IO-aware, O(N) memory, foundational, attention efficiency, transformer",
     "Unread",
     "Foundational IO-aware attention reducing memory from O(N²) to O(N) — enables processing longer code snippets with lower memory overhead during semantic analysis.",
     "https://arxiv.org/abs/2205.14135", ""),
    (81,
     "Why Low-Precision Transformer Training Fails: An Analysis on Flash Attention",
     "Multiple authors", "2025", "arXiv:2510.04212",
     "Inference Speed",
     "low-precision, GGUF, flash attention, quantization, performance analysis, precision limits",
     "Unread",
     "Analyzes Flash Attention behavior under low-precision quantization — critical for understanding performance characteristics of quantized models used in ZeroTrust.sh's local deployment.",
     "https://arxiv.org/abs/2510.04212", ""),
    (82,
     "Optimizing LLM Inference Throughput via Memory-aware and SLA-constrained Dynamic Batching",
     "Multiple authors", "2025", "arXiv:2503.05248",
     "Inference Speed",
     "dynamic batching, throughput, SLA, memory-aware, continuous batching, parallel scanning",
     "Unread",
     "Memory-aware dynamic batching with SLA constraints — optimizes continuous batching for ZeroTrust.sh's parallel Path A/B analysis phases to maximize throughput.",
     "https://arxiv.org/abs/2503.05248", ""),

    # ── AREA 16 — Structured Generation & Constrained Decoding ───────────────
    (83,
     "XGrammar: Flexible and Efficient Structured Generation Engine for Large Language Models",
     "Multiple authors", "2024", "arXiv:2411.15100",
     "Structured Generation",
     "grammar-constrained decoding, BNF, CFG, JSON, zero overhead, structured output, XGrammar",
     "Unread",
     "Grammar-constrained decoding supporting BNF/CFG with zero overhead — enables Path B LLM to output structured JSON vulnerability findings without post-processing verification tokens.",
     "https://arxiv.org/abs/2411.15100", ""),
    (84,
     "JSONSchemaBench: A Rigorous Benchmark of Structured Outputs for Language Models",
     "Multiple authors", "2025", "arXiv:2501.10868",
     "Structured Generation",
     "JSON schema, structured output, Guidance, Outlines, XGrammar, benchmark, compliance",
     "Unread",
     "Benchmarks constrained-decoding frameworks (Guidance, Outlines, XGrammar) on JSON Schema compliance — validates structured generation reliability for ZeroTrust.sh vulnerability report output.",
     "https://arxiv.org/abs/2501.10868", ""),
    (85,
     "SynCode: LLM Generation with Grammar Augmentation",
     "Multiple authors", "2024", "arXiv:2403.01632",
     "Structured Generation",
     "grammar augmentation, syntactic constraints, output schema, structured generation, alternative to XGrammar",
     "Unread",
     "Grammar-augmented generation enforcing output structure via syntactic constraints — alternative to XGrammar for ensuring ZeroTrust.sh semantic output conforms to fixed schema.",
     "https://arxiv.org/abs/2403.01632", ""),

    # ── AREA 17 — Context Window Management ──────────────────────────────────
    (86,
     "Beyond RAG vs. Long-Context: Learning Distraction-Aware Retrieval for Efficient Knowledge Grounding",
     "Multiple authors", "2024", "arXiv:2509.21865",
     "Context Management",
     "RAG, long context, distraction-aware, lost in the middle, retrieval, code snippet selection",
     "Unread",
     "Distraction-aware retrieval mitigating 'lost in the middle' effect — addresses how ZeroTrust.sh should select and contextualize code snippets rather than passing entire files to the LLM.",
     "https://arxiv.org/abs/2509.21865", ""),

    # ── AREA 12 — Exploit Automation & Proof-of-Concept Generation ────────────
    (64,
     "Automated Vulnerability Validation and Verification: A Large Language Model Approach",
     "Lotfi et al.", "2025", "arXiv:2509.24037",
     "Exploit Automation",
     "CVE exploitation, Docker, automated validation, 70% reproduction rate, offline LLM, PoE layer",
     "Unread",
     "End-to-end LLM pipeline for automated CVE exploitation with Docker orchestration; 70% CVE reproduction rate — directly applicable to ZeroTrust.sh's Approach 3 PoE layer design.",
     "https://arxiv.org/abs/2509.24037", ""),
    (65,
     "Patch-to-PoC: A Systematic Study of Agentic LLM Systems for Linux Kernel N-Day Reproduction",
     "Pu et al.", "2026", "arXiv:2602.07287",
     "Exploit Automation",
     "PoC generation, kernel exploit, KernelCTF, VM sandbox, 50%+ success, K-REPRO agent",
     "Unread",
     "Autonomous LLM-based PoC generation from kernel patches with 50%+ success on KernelCTF — K-REPRO agent architecture with VM management informs sandbox PoE design.",
     "https://arxiv.org/abs/2602.07287", ""),
    (66,
     "PentestAgent: Incorporating LLM Agents to Automated Penetration Testing",
     "Feng et al.", "2024", "arXiv:2411.05185",
     "Exploit Automation",
     "penetration testing, multi-agent, VulHub, Docker, reconnaissance, exploitation, red team",
     "Unread",
     "Multi-agent framework for autonomous pentest using Docker VulHub environments — demonstrates agent architecture for reconnaissance and exploitation applicable to ZeroTrust.sh red-team layer.",
     "https://arxiv.org/abs/2411.05185", ""),
    (67,
     "Directed Greybox Fuzzing via Large Language Model",
     "Xu et al.", "2025", "arXiv:2505.03425",
     "Exploit Automation",
     "LLM fuzzing, path constraint, code generation, 9 new CVEs, 17/20 trigger rate, HGFuzzer",
     "Unread",
     "HGFuzzer uses LLMs to transform path constraints into code generation tasks; discovered 9 new CVEs with 17/20 trigger rate — alternative exploit discovery layer for Approach 3.",
     "https://arxiv.org/abs/2505.03425", ""),
    (68,
     "PoCGen: Generating Proof-of-Concept Exploits for Vulnerabilities in NPM Packages",
     "Shen et al.", "2025", "arXiv:2506.04962",
     "Exploit Automation",
     "PoC generation, npm, JavaScript, Node.js, 77% success rate, CVE disclosure, ecosystem-specific",
     "Unread",
     "Autonomous LLM-based PoC for npm vulnerabilities (77% on SecBench.js); 6 PoCs in official CVE disclosures — demonstrates ecosystem-specific PoE applicability for JavaScript targets.",
     "https://arxiv.org/abs/2506.04962", ""),
]

def main():
    wb = openpyxl.Workbook()
    
    # 1. Build Dashboard
    build_dashboard(wb)
    
    # 2. Build Approach 1
    build_approach_sheet(
        wb,
        "ZeroTrust.sh  ·  Approach 1 (Semgrep PoC)",
        "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Goal: Semgrep-based custom rule engine  ·  Deadline: 2026-06-20",
        "Approach 1 - Semgrep PoC",
        "1",
        approach1_data
    )
    
    # 3. Build Approach 2
    build_approach_sheet(
        wb,
        "ZeroTrust.sh  ·  Approach 2 (Hybrid LLM)",
        "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Goal: Go engine + Two-Path detection  ·  Deadline: 2026-07-28",
        "Approach 2 - Hybrid LLM",
        "2",
        approach2_data
    )
    
    # 4. Build Approach 3
    build_approach_sheet(
        wb,
        "ZeroTrust.sh  ·  Approach 3 (Agentic Scanner)",
        "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Goal: Full PoE verification pipeline  ·  Deadline: 2026-08-01",
        "Approach 3 - Agentic Scanner",
        "3",
        approach3_data
    )
    
    # 5. Build Research
    build_approach_sheet(
        wb,
        "ZeroTrust.sh  ·  Scientific Research & Architecture Validation",
        "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Goal: Evidence-backed architecture validation  ·  Runs Jun 9 – Aug 1",
        "Research",
        "R",
        research_data
    )
    
    # 6. Build Constraints
    build_constraints(wb, constraints_data)
    
    # 7. Build Research Papers
    build_research_papers(wb, papers)
    
    # Save Workbook
    out = "docs/ZeroTrust_Internship_Roadmap.xlsx"
    wb.save(out)
    print(f"Roadmap Excel workbook successfully generated at: {out}")

if __name__ == "__main__":
    main()
