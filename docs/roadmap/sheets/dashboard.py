from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
from .styles import DARK_BLUE, MID_BLUE, LIGHT_BLUE, fill, align, border

# Sheet names used in COUNTIF formulas — must match what main() passes to build_goal_sheet
G1 = "G1 - Foundation"
G2 = "G2 - Path A"
G3 = "G3 - Path B"
G4 = "G4 - Integration"
RS = "Research"


def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"

    # Row 1 — title
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "ZeroTrust.sh  ·  Executive Dashboard"
    title.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    title.fill = fill(DARK_BLUE)
    title.alignment = align("center", "center")
    ws.row_dimensions[1].height = 50

    # Row 2 — subtitle
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Date range: 2026-06-11 to 2026-08-06"
    sub.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    sub.fill = fill(MID_BLUE)
    sub.alignment = align("center", "center")
    ws.row_dimensions[2].height = 22

    # Row 3 — spacer
    ws.row_dimensions[3].height = 6

    # Row 4 — Zone 1 header
    ws.merge_cells("A4:G4")
    z1 = ws["A4"]
    z1.value = "PROJECT KPIs"
    z1.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z1.fill = fill(DARK_BLUE)
    z1.alignment = align("left", "center")
    ws.row_dimensions[4].height = 30

    # Row 5 — KPI labels
    kpi_labels = ["Total Tasks", "Completed", "In Progress", "Blocked", "Days to Deadline", "% Complete"]
    ws.row_dimensions[5].height = 22
    for col_idx, label in enumerate(kpi_labels, start=1):
        cell = ws.cell(row=5, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
    ws.cell(row=5, column=7).border = border()

    # Row 6 — KPI values
    ws.row_dimensions[6].height = 60

    def _sheets_countif(criteria):
        parts = [f"COUNTIF('{s}'!C:C,\"{criteria}\")" for s in (G1, G2, G3, G4, RS)]
        return "=" + "+".join(parts)

    def _sheets_countifs(status):
        parts = [
            f"COUNTIFS('{s}'!C:C,\"TASK\",'{s}'!K:K,\"{status}\")"
            for s in (G1, G2, G3, G4, RS)
        ]
        return "=" + "+".join(parts)

    ws.cell(row=6, column=1, value=_sheets_countif("TASK"))
    ws.cell(row=6, column=2, value=_sheets_countifs("Complete"))
    ws.cell(row=6, column=3, value=_sheets_countifs("In Progress"))
    ws.cell(row=6, column=4, value=_sheets_countifs("Blocked"))
    ws.cell(row=6, column=5, value="=DATE(2026,8,6)-TODAY()")
    ws.cell(row=6, column=6, value="=B6/A6")

    for c_idx in range(1, 7):
        cell = ws.cell(row=6, column=c_idx)
        cell.font = Font(name="Calibri", size=20, bold=True, color="000000")
        cell.fill = fill(LIGHT_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        if c_idx == 6:
            cell.number_format = "0.0%"
        elif c_idx == 5:
            cell.number_format = "#,##0"
    ws.cell(row=6, column=7).border = border()

    # Row 7 — spacer
    ws.row_dimensions[7].height = 6

    # Row 8 — Zone 2 header
    ws.merge_cells("A8:G8")
    z2 = ws["A8"]
    z2.value = "PROGRESS BY GOAL"
    z2.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z2.fill = fill(DARK_BLUE)
    z2.alignment = align("left", "center")
    ws.row_dimensions[8].height = 30

    # Row 9 — column headers
    z2_headers = ["Goal", "Not Started", "In Progress", "Complete", "Blocked", "Total Tasks", "% Done"]
    ws.row_dimensions[9].height = 22
    for col_idx, h in enumerate(z2_headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()

    # Rows 10-14 — per-goal data (5 goals)
    goals_data = [
        ("G1 — Foundation & Detection Scaffold", G1),
        ("G2 — Path A: CPG + Taint + LLM Verifier",  G2),
        ("G3 — Path B: Semantic Cost Funnel",          G3),
        ("G4 — Integration, Report & Demo",            G4),
        ("Scientific Research",                        RS),
    ]
    STATUS_UNREAD   = "F5F5F5"
    STATUS_READING  = "FFF3CD"
    STATUS_REVIEWED = "D4EDDA"

    for idx, (label, sheet) in enumerate(goals_data, start=10):
        ws.row_dimensions[idx].height = 30

        cell_a = ws.cell(row=idx, column=1, value=label)
        cell_a.font = Font(name="Calibri", size=12, bold=True, color="000000")
        cell_a.fill = fill("F5F5F5")
        cell_a.alignment = align("left", "center")
        cell_a.border = border()

        formulas = [
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"Not Started\")",
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"In Progress\")",
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"Complete\")",
            f"=COUNTIFS('{sheet}'!C:C,\"TASK\",'{sheet}'!K:K,\"Blocked\")",
            f"=COUNTIF('{sheet}'!C:C,\"TASK\")",
            f"=D{idx}/F{idx}",
        ]
        fills_colors = [STATUS_UNREAD, STATUS_READING, STATUS_REVIEWED, "F8D7DA", "FFFFFF", "FFFFFF"]
        fonts_colors = ["666666", "B45309", "1E7B34", "842029", "000000", "000000"]

        for offset, (form, fl, fn) in enumerate(zip(formulas, fills_colors, fonts_colors), start=2):
            cell = ws.cell(row=idx, column=offset, value=form)
            cell.font = Font(name="Calibri", size=12, color=fn)
            cell.fill = fill(fl)
            cell.alignment = align("center", "center")
            cell.border = border()
            if offset == 7:
                cell.font = Font(name="Calibri", size=12, bold=True, color="000000")
                cell.number_format = "0.0%"

    # Row 15 — totals
    ws.row_dimensions[15].height = 30
    total_lbl = ws.cell(row=15, column=1, value="TOTAL")
    total_lbl.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    total_lbl.fill = fill(DARK_BLUE)
    total_lbl.alignment = align("left", "center")
    total_lbl.border = border()

    total_formulas = [
        "=SUM(B10:B14)", "=SUM(C10:C14)", "=SUM(D10:D14)",
        "=SUM(E10:E14)", "=SUM(F10:F14)", "=D15/F15",
    ]
    for offset, form in enumerate(total_formulas, start=2):
        cell = ws.cell(row=15, column=offset, value=form)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(DARK_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        if offset == 7:
            cell.number_format = "0.0%"

    ws.conditional_formatting.add("G10:G15",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.8"], stopIfTrue=True,
                   fill=fill("D4EDDA"),
                   font=Font(name="Calibri", size=12, bold=True, color="1E7B34")))
    ws.conditional_formatting.add("G10:G15",
        CellIsRule(operator="between", formula=["0.4", "0.79"], stopIfTrue=True,
                   fill=fill("FFF3CD"),
                   font=Font(name="Calibri", size=12, bold=True, color="B45309")))
    ws.conditional_formatting.add("G10:G15",
        CellIsRule(operator="lessThan", formula=["0.4"], stopIfTrue=True,
                   fill=fill("F8D7DA"),
                   font=Font(name="Calibri", size=12, bold=True, color="842029")))

    # Row 16 — spacer
    ws.row_dimensions[16].height = 6

    # Row 17 — Zone 3 header
    ws.merge_cells("A17:G17")
    z3 = ws["A17"]
    z3.value = "UPCOMING MILESTONES (next 7 days)"
    z3.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    z3.fill = fill(DARK_BLUE)
    z3.alignment = align("left", "center")
    ws.row_dimensions[17].height = 30

    # Row 18 — Zone 3 column headers
    z3_headers = ["Goal", "Milestone ID", "Milestone Name", "Due Date", "Status", "", ""]
    ws.row_dimensions[18].height = 22
    for col_idx in range(1, 8):
        val = z3_headers[col_idx - 1]
        cell = ws.cell(row=18, column=col_idx, value=val if val else None)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()

    # Rows 19-23 — milestone data
    milestones_data = [
        ("G1", "1.M1", "Semgrep + ast-grep Rule Suite",                  "2026-06-16", "In Progress"),
        ("G1", "1.M2", "Go CLI Core",                                    "2026-06-20", "Not Started"),
        ("G1", "1.M3", "Ingestion Layer — MIV + Differential Indexer",   "2026-06-25", "Not Started"),
        ("Research", "R.M1", "Literature Foundation",                     "2026-06-20", "In Progress"),
        ("G1", "1.M4", "Canonical Finding Schema + CLI Output",           "2026-06-27", "Not Started"),
    ]
    for row_offset, (goal, m_id, name, due, status) in enumerate(milestones_data, start=19):
        ws.row_dimensions[row_offset].height = 30
        is_even = row_offset % 2 == 0
        base_fill = "D6E4F0" if is_even else "F5F5F5"

        row_vals = [goal, m_id, name, due, status]
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_offset, column=col_idx, value=row_vals[col_idx - 1])
            cell.font = Font(name="Calibri", size=12, color="000000")
            cell.fill = fill(base_fill)
            cell.alignment = align("center", "center")
            cell.border = border()
            if col_idx == 3:
                cell.alignment = align("left", "center")

        status_cell = ws.cell(row=row_offset, column=5)
        if status == "Complete":
            status_fill, status_color = "D4EDDA", "1E7B34"
        elif status == "In Progress":
            status_fill, status_color = "FFF3CD", "B45309"
        else:
            status_fill, status_color = "F5F5F5", "666666"
        status_cell.fill = fill(status_fill)
        status_cell.font = Font(name="Calibri", size=12, bold=True, color=status_color)

        for col_idx in (6, 7):
            ws.cell(row=row_offset, column=col_idx).border = border()

    # Row 24 — spacer
    ws.row_dimensions[24].height = 6

    # Row 25 — Internship Timeline header (spans 8 cols: A=label + B-H=7 weeks)
    ws.merge_cells("A25:H25")
    ztl = ws["A25"]
    ztl.value = "INTERNSHIP TIMELINE"
    ztl.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ztl.fill = fill(DARK_BLUE)
    ztl.alignment = align("left", "center")
    ws.row_dimensions[25].height = 30

    # Row 26 — week column headers
    weeks = ["Jun W3\nJun 16–22", "Jun W4\nJun 23–29", "Jul W1\nJun 30–Jul 6",
             "Jul W2\nJul 7–13",  "Jul W3\nJul 14–20", "Jul W4\nJul 21–27",
             "Aug W1\nJul 28–Aug 6"]
    ws.row_dimensions[26].height = 36
    cell = ws.cell(row=26, column=1, value="Goal")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = fill(MID_BLUE)
    cell.alignment = align("center", "center")
    cell.border = border()
    for ci, w in enumerate(weeks, start=2):
        cell = ws.cell(row=26, column=ci, value=w)
        cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center", wrap=True)
        cell.border = border()

    # Rows 27-31 — one row per goal; active weeks shaded, inactive gray
    # active[goal][week_idx] = True/False
    gantt = [
        ("G1 — Foundation",          "D6E4F0", "1F3864", [1, 1, 0, 0, 0, 0, 0]),
        ("G2 — Path A",              "C8E6C9", "1B5E20", [0, 0, 1, 1, 1, 0, 0]),
        ("G3 — Path B",              "FFE0B2", "E65100", [0, 0, 0, 0, 0, 1, 1]),
        ("G4 — Integration",         "E1BEE7", "4A148C", [0, 0, 0, 0, 0, 0, 1]),
        ("Research (all goals)",     "E0F7FA", "006064", [1, 1, 1, 1, 1, 1, 1]),
    ]
    INACTIVE_FILL = "F0F0F0"

    for row_offset, (label, act_fill, act_font, active) in enumerate(gantt, start=27):
        ws.row_dimensions[row_offset].height = 28
        cell = ws.cell(row=row_offset, column=1, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
        cell.fill = fill("FAFAFA")
        cell.alignment = align("left", "center")
        cell.border = border()
        for ci, is_active in enumerate(active, start=2):
            cell = ws.cell(row=row_offset, column=ci)
            cell.fill = fill(act_fill if is_active else INACTIVE_FILL)
            cell.border = border()
            if is_active:
                cell.value = "▓"
                cell.font = Font(name="Calibri", size=14, bold=True, color=act_font)
                cell.alignment = align("center", "center")

    # Column widths: A=label, B-H=week columns
    col_widths = [28, 13, 13, 13, 13, 13, 13, 13]
    for c_idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    ws.print_area = "A1:H32"
    ws.freeze_panes = "A7"
