from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .styles import DARK_BLUE, MID_BLUE, fill, align, border


def build_constraints(wb, data):
    ws = wb.create_sheet(title="Constraints")

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "ZeroTrust.sh  ·  Planning Constraints Register"
    title_cell.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    title_cell.fill = fill(DARK_BLUE)
    title_cell.alignment = align("center", "center")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Last updated: 2026-06-12"
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    sub_cell.fill = fill(MID_BLUE)
    sub_cell.alignment = align("center", "center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 6

    headers = ["ID", "Constraint", "Category", "Impact", "Applied To", "Buffer Added", "Notes"]
    col_widths = [8, 55, 14, 10, 30, 20, 50]

    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[4].height = 50

    row_num = 5
    for c in data:
        ws.row_dimensions[row_num].height = 60

        impact = c["impact"]
        if impact == "High":
            row_fill = "F8D7DA"
            row_font = Font(name="Calibri", size=12, bold=True, color="842029")
        elif impact == "Medium":
            row_fill = "FFF3CD"
            row_font = Font(name="Calibri", size=12, color="B45309")
        else:
            row_fill = "EAF4EA"
            row_font = Font(name="Calibri", size=12, color="1E5924")

        values = [
            c["id"], c["constraint"], c["category"], c["impact"],
            c["applied_to"], c["buffer_added"], c["notes"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill(row_fill)
            cell.font = row_font
            cell.border = border()
            if col_idx in (1, 3, 4):
                cell.alignment = align("center", "center")
            else:
                cell.alignment = align("left", "center")
        row_num += 1

    ws.print_area = f"A1:G{row_num - 1}"
    ws.freeze_panes = "A5"
