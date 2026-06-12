from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .styles import DARK_BLUE, MID_BLUE, WHITE, fill, align, border


def build_goal_sheet(wb, title, subtitle, sheet_name, milestones):
    ws = wb.create_sheet(title=sheet_name)

    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    title_cell.fill = fill(DARK_BLUE)
    title_cell.alignment = align("center", "center")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:M2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitle
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
    sub_cell.fill = fill(MID_BLUE)
    sub_cell.alignment = align("center", "center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 6

    headers = [
        "ID", "Name", "Type", "Start Date", "End Date",
        "PERT O (h)", "PERT ML (h)", "PERT P (h)", "PERT E (h)",
        "Actual Hrs", "Status", "Owner", "Notes",
    ]
    col_widths = [14, 52, 13, 13, 13, 11, 11, 11, 11, 12, 17, 10, 38]

    for c_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center")
        cell.border = border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[4].height = 50

    row_num = 5
    for m_idx, m in enumerate(milestones):
        ws.row_dimensions[row_num].height = 40
        m_vals = [
            m["id"], m["name"], m["type"], m.get("start", ""), m.get("end", ""),
            m.get("o"), m.get("ml"), m.get("p"), m["e"],
            "", m.get("status", ""), "", m.get("notes", ""),
        ]

        if m["type"] == "BUFFER":
            m_fill = "FFF9E6"
            m_font = Font(name="Calibri", size=12, italic=True, color="5C4B00")
        elif m["type"] == "STRETCH":
            m_fill = "EDE7F6"
            m_font = Font(name="Calibri", size=12, bold=True, italic=True, color="4527A0")
        elif "Bonus" in m["name"]:
            m_fill = "EAF4EA"
            m_font = Font(name="Calibri", size=12, bold=True, color="1E5924")
        else:
            m_fill = MID_BLUE
            m_font = Font(name="Calibri", size=12, bold=True, color=WHITE)

        for c_idx, val in enumerate(m_vals, start=1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.fill = fill(m_fill)
            cell.font = m_font
            cell.border = border()
            if c_idx in (2, 13):
                cell.alignment = align("left", "center")
            else:
                cell.alignment = align("center", "center")
            if c_idx in (6, 7, 8):
                cell.number_format = "0.0"
            elif c_idx == 9:
                cell.number_format = "0.00"

        row_num += 1

        for t in m.get("tasks", []):
            ws.row_dimensions[row_num].height = 45
            t_vals = [
                t["id"], "    " + t["name"], "TASK", "", "",
                t.get("o"), t.get("ml"), t.get("p"), t["e"],
                "", t["status"], "Hoang", t.get("notes", ""),
            ]

            is_even = row_num % 2 == 0

            if t["status"] == "Complete":
                t_fill = "D4EDDA"
                t_font = Font(name="Calibri", size=12, color="1E7B34")
            elif t["status"] == "In Progress":
                t_fill = "FFF3CD"
                t_font = Font(name="Calibri", size=12, color="B45309")
            else:
                t_fill = "D6E4F0" if is_even else "F5F5F5"
                t_font = Font(name="Calibri", size=12, color="000000")

            for c_idx, val in enumerate(t_vals, start=1):
                cell = ws.cell(row=row_num, column=c_idx, value=val)
                cell.fill = fill(t_fill)
                cell.font = t_font
                cell.border = border()
                if c_idx in (2, 13):
                    cell.alignment = align("left", "center")
                else:
                    cell.alignment = align("center", "center")
                if c_idx in (6, 7, 8):
                    cell.number_format = "0.0"
                elif c_idx == 9:
                    cell.number_format = "0.00"

            status_cell = ws.cell(row=row_num, column=11)
            if t["status"] == "Complete":
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="1E7B34")
            elif t["status"] == "In Progress":
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="B45309")
            elif t["status"] == "Blocked":
                status_cell.fill = fill("F8D7DA")
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="842029")
            elif t["status"] == "At Risk":
                status_cell.fill = fill("FFE5B4")
                status_cell.font = Font(name="Calibri", size=12, bold=True, color="8B4513")
            elif t["status"] == "Not Started":
                status_cell.fill = fill("F5F5F5")
                status_cell.font = Font(name="Calibri", size=12, color="666666")

            ws.cell(row=row_num, column=1).font = Font(
                name="Calibri", size=12, bold=True, color=t_font.color
            )

            row_num += 1

        if m_idx < len(milestones) - 1:
            ws.row_dimensions[row_num].height = 5
            row_num += 1

    ws.print_area = f"A1:M{row_num - 1}"
    ws.freeze_panes = "A5"
