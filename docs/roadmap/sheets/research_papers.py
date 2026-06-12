from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from .styles import (
    DARK_BLUE, MID_BLUE, WHITE, LIGHT_BLUE, LIGHT_GRAY,
    STATUS_UNREAD, STATUS_READING, STATUS_READ, STATUS_REVIEWED,
    fill, align, border,
)


def build_research_papers(wb, papers):
    ws2 = wb.create_sheet(title="Research Papers")
    NUM_COLS = 11
    LAST_COL = get_column_letter(NUM_COLS)

    ws2.merge_cells(f"A1:{LAST_COL}1")
    t2 = ws2["A1"]
    t2.value = "ZeroTrust.sh  ·  Research Paper Manager"
    t2.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    t2.fill = fill(DARK_BLUE)
    t2.alignment = align("center", "center")
    t2.border = border(DARK_BLUE)
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells(f"A2:{LAST_COL}2")
    s2 = ws2["A2"]
    s2.value = (
        "Smart Paper Manager  ·  Use column dropdowns (▾) to filter by Category, Tags, Read Status  "
        "·  Click any column header to sort  ·  Write your notes in column K"
    )
    s2.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
    s2.fill = fill(MID_BLUE)
    s2.alignment = align("center", "center")
    s2.border = border(MID_BLUE)
    ws2.row_dimensions[2].height = 18

    ws2.row_dimensions[3].height = 6

    headers2 = [
        "#", "Title", "Authors", "Year", "Venue",
        "Category", "Tags", "Read Status",
        "Relevance to ZeroTrust.sh", "URL", "Literature Review Notes",
    ]
    col_widths2 = [5, 56, 22, 6, 24, 26, 38, 14, 50, 40, 62]

    for ci, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        cell = ws2.cell(row=4, column=ci, value=h)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.fill = fill(MID_BLUE)
        cell.alignment = align("center", "center", wrap=False)
        cell.border = border(MID_BLUE)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[4].height = 22

    current_row = 5
    last_paper_row = current_row

    for entry in papers:
        (num, title, authors, year, venue, category, tags,
         read_status, relevance, url, lit_notes) = entry

        is_even = current_row % 2 == 0
        base_fill_color = LIGHT_BLUE if is_even else LIGHT_GRAY

        status_color_map = {
            "Unread":   STATUS_UNREAD,
            "Reading":  STATUS_READING,
            "Read":     STATUS_READ,
            "Reviewed": STATUS_REVIEWED,
        }

        values = [num, title, authors, year, venue, category, tags,
                  read_status, relevance, url, lit_notes]

        for ci, value in enumerate(values, start=1):
            cell = ws2.cell(row=current_row, column=ci, value=value)
            cell_fill = base_fill_color
            if ci == 8:
                cell_fill = status_color_map.get(read_status, STATUS_UNREAD)
            if ci == 11:
                cell_fill = "FFFEF5" if not is_even else "F0F8FF"
            cell.fill = fill(cell_fill)
            cell.border = border()
            cell.alignment = align("left", "center", wrap=True)
            cell.font = Font(size=10, name="Calibri")

            if ci == 1:
                cell.alignment = align("center", "center", wrap=False)
                cell.font = Font(bold=True, size=10, name="Calibri")
            if ci == 4:
                cell.alignment = align("center", "center", wrap=False)
            if ci == 8:
                cell.alignment = align("center", "center", wrap=False)
                cell.font = Font(bold=True, size=9, name="Calibri")
            if ci == 10:
                cell.font = Font(size=9, color="2E5FA3", name="Calibri", underline="single")
            if ci == 11:
                if not value:
                    cell.font = Font(size=9, color="AAAAAA", italic=True, name="Calibri")
                    cell.value = "Write your notes here…"

        ws2.row_dimensions[current_row].height = 60
        last_paper_row = current_row
        current_row += 1

    ws2.auto_filter.ref = f"A4:{LAST_COL}{last_paper_row}"

    dv = DataValidation(
        type="list",
        formula1='"Unread,Reading,Read,Reviewed"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Choose: Unread, Reading, Read, or Reviewed",
    )
    ws2.add_data_validation(dv)
    dv.add(f"H5:H{last_paper_row}")

    legend2_row = last_paper_row + 2
    ws2.merge_cells(f"A{legend2_row}:{LAST_COL}{legend2_row}")
    leg2 = ws2[f"A{legend2_row}"]
    leg2.value = (
        "Read Status:  Unread (gray) — not yet read  ·  Reading (amber) — in progress  "
        "·  Read (blue) — finished  ·  Reviewed (green) — literature review written  "
        "·  Filter tip: click ▾ on Category or Tags header to filter by research area"
    )
    leg2.font = Font(italic=True, color="666666", size=9, name="Calibri")
    leg2.alignment = align("left", "center")

    ws2.print_area = f"A1:{LAST_COL}{last_paper_row}"
    ws2.freeze_panes = "B5"
