import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Execution Overview"

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E5FA3"
LIGHT_BLUE  = "D6E4F0"
ACCENT_GOLD = "C9A84C"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
MID_GRAY    = "CCCCCC"

# Read Status badge colors
STATUS_UNREAD   = "F5F5F5"   # light gray
STATUS_READING  = "FFF3CD"   # warm amber
STATUS_READ     = "D6E4F0"   # light blue
STATUS_REVIEWED = "D4EDDA"   # soft green

def side(color=MID_GRAY, style="thin"):
    return Side(border_style=style, color=color)

def border(all_sides=MID_GRAY):
    s = side(all_sides)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=None, size=11, italic=False):
    return Font(bold=bold, color=color or "000000", size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Execution Overview Data ───────────────────────────────────────────────────
rows = [
    ("2026-06-09", "M-1  Research & Setup",
     "Install Semgrep CLI · Read YAML rule docs · Write one toy rule end-to-end · Scaffold repo (rules/ tests/ scripts/)",
     "Complete"),

    ("2026-06-10", "M-2  Python Custom Rules (Day 1 of 2)",
     "Write custom rules: LLM prompt injection (OpenAI / Anthropic / LangChain) · AI bypass comment detection · Set up bad.py / ok.py test pairs",
     "In Progress"),

    ("2026-06-11", "M-2  Python Custom Rules (Day 2 of 2)",
     "Write custom rule: hardcoded AI service API keys (sk- / sk-ant- / hf_) · Configure community rule packs (p/python, p/owasp-top-ten) · Tune false positive rate with pattern-not",
     ""),

    ("2026-06-12", "M-3  Java Custom Rules (Day 1 of 2)",
     "Validate Java AST shapes with semgrep --dump-ast · Write custom rule: AI bypass comments (Java) · Write custom rule: hardcoded AI service credentials in Java",
     ""),

    ("2026-06-13", "M-3  Java Custom Rules (Day 2 of 2)",
     "Configure community rule packs (p/java, p/java-security-audit) · Tune rules against bad.java / ok.java test pairs · Document precision-recall tradeoff per rule",
     ""),

    ("2026-06-16", "M-4  Test Codebase",
     "AI-generate fake Spring Boot REST API (10–15 files, 800–1 200 LOC) · Embed ≥8 intentional vulnerabilities · Run full rule set · Document detection rate and false positive count",
     ""),

    ("2026-06-17", "M-5  Demo Preparation",
     "Write demo/run_demo.sh with pinned Semgrep version · Full dry-run in fresh terminal · Record 3-minute fallback video",
     ""),

    ("2026-06-18", "M-6  Presentation Narrative",
     "Write pros / cons of Semgrep-only approach · Draft next-step argument for Approach 2 (LLM Verifier + Path B) · Add speaker notes",
     ""),

    ("2026-06-19", "M-7  Jupyter Notebook  (Bonus)",
     "Precision and recall per rule · Scan speed (lines/second) · AI-specific detection rate · False positive rate on clean codebase · Charts",
     ""),

    ("2026-06-20", "Presentation",
     "Live CLI demo + narrative delivery to tech lead · Collect approval or actionable feedback",
     ""),
]

# ── Execution Overview Sheet ──────────────────────────────────────────────────
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "ZeroTrust.sh  ·  Approach 1 (Semgrep PoC)  ·  Executive Timeline"
title_cell.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
title_cell.fill = fill(DARK_BLUE)
title_cell.alignment = align("center", "center")
title_cell.border = border(DARK_BLUE)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:E2")
sub = ws["A2"]
sub.value = "Intern: Ton Minh Hoang  ·  VNG ZingPlay Studio  ·  Deadline: 2026-06-20"
sub.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
sub.fill = fill(MID_BLUE)
sub.alignment = align("center", "center")
sub.border = border(MID_BLUE)
ws.row_dimensions[2].height = 18

ws.row_dimensions[3].height = 6

headers = ["Date", "Milestone", "Tasks", "Status", "Notes"]
col_widths = [14, 30, 70, 14, 22]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    cell.fill = fill(MID_BLUE)
    cell.alignment = align("center", "center", wrap=False)
    cell.border = border(MID_BLUE)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[4].height = 22

for row_num, (date, milestone, tasks, status) in enumerate(rows, start=5):
    is_even = (row_num % 2 == 0)
    row_fill = fill(LIGHT_BLUE if is_even else LIGHT_GRAY)

    is_presentation = "Presentation" in milestone and "M-" not in milestone
    if is_presentation:
        row_fill = fill("FFF3CD")

    is_bonus = "Bonus" in milestone
    if is_bonus:
        row_fill = fill("EAF4EA")

    values = [date, milestone, tasks, status, ""]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.fill = row_fill
        cell.border = border()
        cell.alignment = align("left", "center", wrap=True)
        cell.font = font(size=10)

        if col_idx == 1:
            cell.font = font(bold=True, size=10)
            cell.alignment = align("center", "center", wrap=False)

        if col_idx == 2:
            cell.font = font(bold=True, size=10,
                             color=ACCENT_GOLD if is_presentation else "000000")

        if col_idx == 4 and value == "Complete":
            cell.font = font(bold=True, color="1E7B34", size=10)
        elif col_idx == 4 and value == "In Progress":
            cell.font = font(bold=True, color="B45309", size=10)

    ws.row_dimensions[row_num].height = 42

legend_row = len(rows) + 6
ws.merge_cells(f"A{legend_row}:E{legend_row}")
legend = ws[f"A{legend_row}"]
legend.value = (
    "Status key:  Complete = delivered  ·  In Progress = active  "
    "·  (blank) = not started  ·  Bonus milestone is time-permitting only"
)
legend.font = Font(italic=True, color="666666", size=9, name="Calibri")
legend.alignment = align("left", "center")
ws.freeze_panes = "A5"

# ── Research Papers Sheet ─────────────────────────────────────────────────────
#
# Schema (11 columns):
#   A  #               — paper number
#   B  Title           — full paper title
#   C  Authors         — first author et al.
#   D  Year            — publication year
#   E  Venue           — conference / journal / arXiv ID
#   F  Category        — research area (filterable)
#   G  Tags            — comma-separated keywords (filterable)
#   H  Read Status     — Unread / Reading / Read / Reviewed  (dropdown)
#   I  Relevance       — how it applies to ZeroTrust.sh
#   J  URL             — direct link
#   K  Literature Review Notes — user-written review space
#
# Auto-filter is applied to row 4 (header) so every column is sortable
# and filterable natively in Excel / LibreOffice.
# Read Status has a dropdown data-validation list.
# ─────────────────────────────────────────────────────────────────────────────

ws2 = wb.create_sheet(title="Research Papers")

NUM_COLS = 11
LAST_COL = get_column_letter(NUM_COLS)   # "K"

ws2.merge_cells(f"A1:{LAST_COL}1")
t2 = ws2["A1"]
t2.value = "ZeroTrust.sh  ·  Research Paper Manager"
t2.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
t2.fill = fill(DARK_BLUE)
t2.alignment = align("center", "center")
t2.border = border(DARK_BLUE)
ws2.row_dimensions[1].height = 30

ws2.merge_cells(f"A2:{LAST_COL}2")
s2 = ws2["A2"]
s2.value = (
    "Smart Paper Manager  ·  Use column dropdowns (▾) to filter by Category, Tags, Read Status  "
    "·  Click any column header to sort  ·  Write your notes in column K"
)
s2.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
s2.fill = fill(MID_BLUE)
s2.alignment = align("center", "center")
s2.border = border(MID_BLUE)
ws2.row_dimensions[2].height = 18

ws2.row_dimensions[3].height = 6

headers2 = [
    "#", "Title", "Authors", "Year", "Venue",
    "Category", "Tags", "Read Status",
    "Relevance to ZeroTrust.sh", "URL", "Literature Review Notes"
]
col_widths2 = [5, 56, 22, 6, 24, 26, 38, 14, 50, 40, 62]

for ci, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
    cell = ws2.cell(row=4, column=ci, value=h)
    cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    cell.fill = fill(MID_BLUE)
    cell.alignment = align("center", "center", wrap=False)
    cell.border = border(MID_BLUE)
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[4].height = 22

# ── Paper data ────────────────────────────────────────────────────────────────
#
# Each tuple: (num, title, authors, year, venue, category, tags,
#              read_status, relevance, url, lit_review_notes)
#
# Categories (for filter column F):
#   ML/DL Detection · Graph Neural Networks · LLM Code Analysis
#   Hybrid SAST+LLM · AI-Generated Code Security · Token Cost Optimization
#   Call Graph & Taint Analysis · AI Agent Security · Logic Vuln Detection
#   Local LLM Deployment · Exploit Automation
#
# ─────────────────────────────────────────────────────────────────────────────

papers = [
    # ── AREA 1 — Deep Learning & ML for Vulnerability Detection ───────────────
    (1,
     "Automated Vulnerability Detection in Source Code Using Deep Representation Learning",
     "Feng et al.", "2026", "arXiv",
     "ML/DL Detection",
     "CNN, deep representation learning, source code, vulnerability classifier",
     "Unread",
     "CNN-based deep representation learning for vulnerability detection — validates ML classifier gate in Path B Tier 2",
     "https://arxiv.org/abs/2602.23121", ""),
    (2,
     "DiverseVul: A New Vulnerable Source Code Dataset for Deep Learning Based Vulnerability Detection",
     "Jia et al.", "2023", "RAID 2023 / ACM",
     "ML/DL Detection",
     "dataset, C/C++, CWE, BigVul alternative, training data, benchmark, 18945 functions",
     "Unread",
     "Largest diverse C/C++ vulnerability dataset (18,945 functions, 150 CWEs) — training data source for Code Vulnerability Classifier",
     "https://arxiv.org/abs/2304.00409", ""),
    (3,
     "Vulnerability Detection in C/C++ Code with Deep Learning",
     "Multiple authors", "2024", "arXiv",
     "ML/DL Detection",
     "neural network, program slicing, C/C++, deep learning",
     "Unread",
     "Neural networks with program slices for vulnerability detection — informs Tier 2 classifier design",
     "https://arxiv.org/abs/2405.12384", ""),
    (4,
     "Deep Learning Aided Software Vulnerability Detection: A Survey",
     "Survey authors", "2025", "arXiv",
     "ML/DL Detection",
     "survey, deep learning, vulnerability detection, baseline comparison",
     "Unread",
     "Comprehensive DL survey for vulnerability detection — baseline reference for Tier 2 classifier selection",
     "https://arxiv.org/abs/2503.04002", ""),

    # ── AREA 2 — Graph Neural Networks for Vulnerability Detection ────────────
    (5,
     "Software Vulnerability Detection Using a Lightweight Graph Neural Network (VulGNN)",
     "Zhu et al.", "2026", "arXiv",
     "Graph Neural Networks",
     "GNN, lightweight, LLM-parity, local inference, classifier, small model",
     "Unread",
     "Lightweight GNN achieving LLM-parity at 100x smaller size — validates cheap local classifier concept in new architecture",
     "https://arxiv.org/abs/2603.29216", ""),
    (6,
     "Vul-LMGNNs: Fusing Language Models and Graph Neural Networks for Code Vulnerability Detection",
     "Rong et al.", "2024", "arXiv",
     "Graph Neural Networks",
     "GNN, language model, hybrid, code representation, code LM",
     "Unread",
     "Hybrid code LM + GNN — informs Call Graph + Classifier integration in Path B Tier 2",
     "https://arxiv.org/abs/2404.14719", ""),
    (7,
     "Structure-Aware Code Vulnerability Analysis With Graph Neural Networks",
     "Allamanis et al.", "2024", "arXiv",
     "Graph Neural Networks",
     "GNN, Java, structure-aware, vulnerability fixing, commit-level",
     "Unread",
     "GNN-based vulnerability analysis using Java vulnerability-fixing commits — informs structure-aware detection",
     "https://arxiv.org/abs/2307.11454", ""),
    (8,
     "Graph Neural Networks for Vulnerability Detection: A Counterfactual Explanation",
     "Li et al.", "2024", "arXiv",
     "Graph Neural Networks",
     "GNN, explainability, counterfactual, confidence scoring, interpretability",
     "Unread",
     "Explainability analysis of GNN detection — informs confidence scoring in Dedup layer",
     "https://arxiv.org/abs/2404.15687", ""),
    (9,
     "ReGVD: Revisiting Graph Neural Networks for Vulnerability Detection",
     "Nguyen et al.", "2022", "ACM/IEEE",
     "Graph Neural Networks",
     "GNN, token sequence, baseline, benchmark, foundational",
     "Unread",
     "Foundational GNN model treating source code as flat token sequences — baseline for Tier 2 classifier",
     "https://dl.acm.org/doi/abs/10.1145/3510454.3516865", ""),
    (10,
     "LineVD: Statement-level Vulnerability Detection using Graph Neural Networks",
     "Chen et al.", "2022", "arXiv",
     "Graph Neural Networks",
     "GNN, statement-level, line localization, fine-grained, HTML report",
     "Unread",
     "Fine-grained GNN vulnerability localization — informs line-level finding output in HTML report",
     "https://arxiv.org/abs/2203.05181", ""),

    # ── AREA 3 — LLM for Code Security Analysis ───────────────────────────────
    (11,
     "LLMs in Code Vulnerability Analysis: A Proof of Concept",
     "Kochling et al.", "2026", "arXiv",
     "LLM Code Analysis",
     "LLM, empirical, proof-of-concept, zero-shot, Path B validation",
     "Unread",
     "Empirical PoC for LLM vulnerability analysis — validates LLM Semantic Scan role in Path B Tier 3",
     "https://arxiv.org/abs/2601.08691", ""),
    (12,
     "IRIS: LLM-Assisted Static Analysis for Detecting Security Vulnerabilities",
     "Scanlon et al.", "2024", "arXiv",
     "LLM Code Analysis",
     "hybrid SAST+LLM, false positive, 80% FP reduction, IRIS, key paper",
     "Unread",
     "Hybrid SAST+LLM detecting 55/120 vulns + 6 new, reducing FP by 80% — directly validates ZeroTrust.sh hybrid design",
     "https://arxiv.org/abs/2405.17238", ""),
    (13,
     "Large Language Model for Vulnerability Detection and Repair: Literature Review and the Road Ahead",
     "Zhang et al.", "2025", "ACM TOSEM",
     "LLM Code Analysis",
     "survey, LLM, vulnerability repair, patch generation, comprehensive review",
     "Unread",
     "Comprehensive LLM vulnerability + repair survey — informs patch generation design across all approaches",
     "https://dl.acm.org/doi/10.1145/3708522", ""),
    (14,
     "Understanding the Effectiveness of LLMs in Detecting Security Vulnerabilities",
     "Steenhoek et al.", "2023", "arXiv",
     "LLM Code Analysis",
     "LLM, prompting strategy, systematic evaluation, false positive, prompt design",
     "Unread",
     "Systematic LLM evaluation with prompting strategy analysis — informs LLM Verifier prompt design in Path A",
     "https://arxiv.org/abs/2311.16169", ""),
    (15,
     "Large Language Models for Source Code Analysis: Applications, Models and Datasets",
     "Sharma et al.", "2025", "arXiv",
     "LLM Code Analysis",
     "survey, LLM architectures, code analysis, model selection, datasets",
     "Unread",
     "Survey of LLM architectures for code analysis — model selection reference for LLM components",
     "https://arxiv.org/abs/2503.17502", ""),
    (16,
     "Can Large Language Models Find And Fix Vulnerable Software?",
     "Pearce et al.", "2023", "arXiv",
     "LLM Code Analysis",
     "LLM, detection, repair, empirical evaluation, fix generation",
     "Unread",
     "Empirical evaluation of LLM detection + repair — validates dual-role LLM use in Approaches 2 and 3",
     "https://arxiv.org/abs/2308.10345", ""),

    # ── AREA 4 — Hybrid Static Analysis + LLM ─────────────────────────────────
    (17,
     "LLM-Driven SAST-Genius: A Hybrid Static Analysis Framework for Comprehensive and Actionable Security",
     "Multiple authors", "2024", "arXiv",
     "Hybrid SAST+LLM",
     "hybrid, SAST+LLM, false positive, 91% FP reduction, Semgrep, key paper",
     "Unread",
     "Hybrid SAST+LLM reducing FP by 91% (225→20 alerts) vs Semgrep alone — strongest validation of ZeroTrust.sh architecture",
     "https://arxiv.org/abs/2509.15433", ""),
    (18,
     "ZeroFalse: Improving Precision in Static Analysis with LLMs",
     "Scanlon et al.", "2024", "arXiv",
     "Hybrid SAST+LLM",
     "LLM, false positive reduction, static analysis, precision, ZeroFalse",
     "Unread",
     "LLM false positive reduction in static analysis — validates LLM Verifier design in Path A",
     "https://arxiv.org/abs/2510.02534", ""),
    (19,
     "Combining Large Language Models with Static Analyzers for Code Review Generation",
     "Jaoua et al.", "2025", "arXiv",
     "Hybrid SAST+LLM",
     "LLM, static analysis, code review, patch suggestion, output format",
     "Unread",
     "LLM + static analysis for code review — informs patch suggestion output format",
     "https://arxiv.org/abs/2502.06633", ""),
    (20,
     "A Contemporary Survey of LLM-Assisted Program Analysis",
     "Survey authors", "2025", "arXiv",
     "Hybrid SAST+LLM",
     "survey, program analysis, LLM, comprehensive, all approaches reference",
     "Unread",
     "Comprehensive survey of LLM program analysis techniques — architecture reference for all three approaches",
     "https://arxiv.org/abs/2502.18474", ""),
    (21,
     "RepoAudit: An Autonomous LLM-Agent for Repository-Level Code Auditing",
     "Li et al.", "2025", "arXiv",
     "Hybrid SAST+LLM",
     "LLM agent, repository-level, multi-agent, orchestration, Approach 3",
     "Unread",
     "LLM-agent for repo-level auditing — informs Approach 3 multi-agent orchestration design",
     "https://arxiv.org/abs/2501.18160", ""),

    # ── AREA 5 — AI-Generated Code Security & Prompt Injection ────────────────
    (22,
     "Security Vulnerabilities in AI-Generated Code: A Large-Scale Analysis of Public GitHub Repositories",
     "Zhao et al.", "2024", "IEEE/ACM",
     "AI-Generated Code Security",
     "AI-generated code, CWE, GitHub, empirical, 4241 instances, market validation",
     "Unread",
     "4,241 CWE instances across AI-generated code from 4 tools — empirical validation of the core ZeroTrust.sh problem statement",
     "https://arxiv.org/abs/2510.26103", ""),
    (23,
     "Prompt Injection Attacks on Agentic Coding Assistants: A Systematic Analysis",
     "Yang et al.", "2026", "arXiv",
     "AI Agent Security",
     "prompt injection, agentic, coding assistant, attack, 85% success rate",
     "Unread",
     "85%+ attack success rates for prompt injection in agentic assistants — validates ZeroTrust.sh AI-specific threat detection",
     "https://arxiv.org/abs/2601.17548", ""),
    (24,
     "Security Degradation in Iterative AI Code Generation: A Systematic Analysis of the Paradox",
     "Multiple authors", "2025", "IEEE-ISTAS",
     "AI-Generated Code Security",
     "iterative generation, vibe-coding, security degradation, LLM, threat model",
     "Unread",
     "Iterative LLM interactions without human review introduce new vulnerabilities — validates vibe-coding threat model",
     "https://arxiv.org/abs/2506.11022", ""),
    (25,
     "Assessing the Quality and Security of AI-Generated Code: A Quantitative Analysis",
     "Multiple authors", "2024", "arXiv",
     "AI-Generated Code Security",
     "AI-generated code, quantitative, security assessment, quality metrics",
     "Unread",
     "Quantitative security analysis of AI-generated code — market validation and threat taxonomy reference",
     "https://arxiv.org/abs/2508.14727", ""),
    (26,
     "You Still Have to Study: On the Security of LLM Generated Code",
     "Ferrara et al.", "2024", "arXiv",
     "AI-Generated Code Security",
     "Copilot, CWE, 36-40% vulnerability rate, LLM-generated code, pitch statistic",
     "Unread",
     "36–40% of Copilot code contains CWE vulnerabilities — key statistic for product positioning and pitch",
     "https://arxiv.org/abs/2408.07106", ""),

    # ── AREA 6 — Token Cost Optimization for LLM Pipelines ────────────────────
    (27,
     "FrugalGPT: How to Use Large Language Models While Reducing Cost and Improving Performance",
     "Eisingerich et al.", "2023", "arXiv",
     "Token Cost Optimization",
     "cascade routing, cost optimization, 98% savings, LLM pipeline, FrugalGPT",
     "Unread",
     "Cascade routing achieving 98% cost savings — theoretical foundation for Cascading Intelligence architecture",
     "https://arxiv.org/abs/2305.05176", ""),
    (28,
     "Batch Prompting: Efficient Inference with Large Language Model APIs",
     "Rajkumar et al.", "2023", "arXiv",
     "Token Cost Optimization",
     "batch processing, token cost, 5x reduction, inference efficiency",
     "Unread",
     "Batch processing reducing LLM token costs up to 5x — informs Token Budget Controller batching strategy",
     "https://arxiv.org/abs/2301.08721", ""),
    (29,
     "Token Sugar: Making Source Code Sweeter for LLMs through Token-Efficient Shorthand",
     "Multiple authors", "2025", "arXiv",
     "Token Cost Optimization",
     "token optimization, code representation, shorthand, context compression",
     "Unread",
     "Token optimization for code representation — informs context chunking in Token Budget Controller",
     "https://arxiv.org/abs/2512.08266", ""),
    (30,
     "Learning to Focus: Context Extraction for Efficient Code Vulnerability Detection with Language Models",
     "Dittmann et al.", "2025", "arXiv",
     "Token Cost Optimization",
     "context filtering, token budget, vulnerability detection, efficiency, focus extraction",
     "Unread",
     "Context filtering for reducing LLM token consumption in vulnerability detection — directly validates Token Budget Controller",
     "https://arxiv.org/abs/2505.17460", ""),

    # ── AREA 7 — Call Graph, Taint Analysis & Code Representations ────────────
    (31,
     "Vulnerability Detection with Interprocedural Context in Multiple Languages: Assessing Effectiveness and Cost",
     "Gharibi et al.", "2024", "arXiv",
     "Call Graph & Taint Analysis",
     "interprocedural, multi-language, call graph, LLM detection, effectiveness",
     "Unread",
     "Interprocedural analysis impact on LLM detection across languages — validates Call Graph + CVE Enrichment in Path B",
     "https://arxiv.org/abs/2604.08417", ""),
    (32,
     "Multi-Agent Taint Specification Extraction for Vulnerability Detection",
     "Zhang et al.", "2026", "arXiv",
     "Call Graph & Taint Analysis",
     "taint analysis, multi-agent, CodeQL, LLM, Approach 3",
     "Unread",
     "Multi-agent LLM + taint analysis — informs Approach 3 multi-agent architecture with CodeQL/Joern integration",
     "https://arxiv.org/abs/2601.10865", ""),
    (33,
     "LLMxCPG: Context-Aware Vulnerability Detection Through Code Property Graph-Guided LLMs",
     "Pan et al.", "2024", "arXiv",
     "Call Graph & Taint Analysis",
     "CPG, code property graph, context-aware, LLM, Path B integration",
     "Unread",
     "CPG-guided LLM for context-aware detection — informs CPG integration in Path B call graph analysis",
     "https://arxiv.org/abs/2507.16585", ""),
    (34,
     "Bridging Code Property Graphs and Language Models for Program Analysis",
     "Mahfouz et al.", "2026", "arXiv",
     "Call Graph & Taint Analysis",
     "CPG, LLM, program analysis, hybrid CPG+LLM, framework",
     "Unread",
     "Framework bridging CPG and LLMs — validates hybrid CPG+LLM design in Path B",
     "https://arxiv.org/abs/2603.24837", ""),
    (35,
     "Enhancing Software Vulnerability Detection Using Code Property Graphs and Convolutional Neural Networks",
     "Multiple authors", "2025", "arXiv",
     "Call Graph & Taint Analysis",
     "CPG, CNN, local structure, global structure, code classifier training",
     "Unread",
     "CPG+CNN for local and global code structure — informs Code Classifier training in new architecture",
     "https://arxiv.org/abs/2503.18175", ""),
    (36,
     "VulTrLM: LLM-Assisted Vulnerability Detection via AST Decomposition and Comment Enhancement",
     "Liu et al.", "2025", "Empirical Software Engineering",
     "Call Graph & Taint Analysis",
     "LLM, AST decomposition, semantic enhancement, comment, AST preprocessing",
     "Unread",
     "LLM-assisted AST decomposition for semantic enhancement — informs AST preprocessing in Path B Tier 1",
     "https://dl.acm.org/doi/10.1007/s10664-025-10738-7", ""),
    (37,
     "Dataflow Analysis-Inspired Deep Learning for Efficient Vulnerability Detection",
     "Cheng et al.", "2022", "arXiv",
     "Call Graph & Taint Analysis",
     "dataflow analysis, deep learning, CodeQL integration, Path A validation",
     "Unread",
     "Dataflow analysis-inspired DL approach — validates dataflow integration in Path A CodeQL/Joern",
     "https://arxiv.org/abs/2212.08108", ""),
    (38,
     "Reducing False Positives in Static Bug Detection with LLMs: An Empirical Study in Industry",
     "Dittmann et al.", "2026", "arXiv",
     "Hybrid SAST+LLM",
     "false positive, LLM, industrial study, production scale, empirical",
     "Unread",
     "Industrial study on LLM false positive reduction — validates LLM Verifier design in Path A at production scale",
     "https://arxiv.org/abs/2601.18844", ""),
    (39,
     "LSAST: Enhancing Cybersecurity through LLM-supported Static Application Security Testing",
     "Multiple authors", "2024", "arXiv",
     "Local LLM Deployment",
     "local LLM, privacy-first, offline, SAST, no cloud API, on-device",
     "Unread",
     "Locally-hostable LLM for SAST without cloud APIs — validates privacy-first local LLM deployment approach",
     "https://arxiv.org/abs/2409.15735", ""),
    (40,
     "Software Vulnerability Analysis Across Programming Language and Program Representation Landscapes: A Survey",
     "Multiple authors", "2025", "arXiv",
     "Call Graph & Taint Analysis",
     "survey, AST, CFG, PDG, CPG, program representation, multi-language",
     "Unread",
     "Survey of AST, CFG, PDG, CPG representations — reference for program representation selection across all approaches",
     "https://arxiv.org/abs/2503.20244", ""),

    # ── AREA 8 — Package Hallucination & Supply Chain ─────────────────────────
    (41,
     "We Have a Package for You! A Comprehensive Analysis of Package Hallucinations by Code Generating LLMs",
     "Spracklen et al.", "2024", "USENIX Security 2025",
     "AI Agent Security",
     "slopsquatting, package hallucination, supply chain, npm, pip, 205474 hallucinations, foundational",
     "Unread",
     "Foundational study: 205,474 hallucinated package names across 16 LLMs — commercial at 5.2%, open-source at 21.7%. Baseline threat ZeroTrust.sh must detect.",
     "https://arxiv.org/abs/2406.10279", ""),
    (42,
     "Importing Phantoms: Measuring LLM Package Hallucination Vulnerabilities",
     "Krishna et al.", "2025", "arXiv:2501.19012",
     "AI Agent Security",
     "slopsquatting, package hallucination, language comparison, model size, Pareto frontier",
     "Unread",
     "Hallucination rates across languages, model sizes, and task specificity — identifies highest-risk agent configurations for slopsquatting detection.",
     "https://arxiv.org/abs/2501.19012", ""),
    (43,
     "The Range Shrinks, the Threat Remains: Re-evaluating LLM Package Hallucinations on the 2026 Frontier-Model Cohort",
     "Multiple authors", "2026", "arXiv:2605.17062",
     "AI Agent Security",
     "slopsquatting, frontier models, Claude, GPT-5, 127 common hallucinations, 2026 benchmark",
     "Unread",
     "Frontier models (Claude Haiku 4.5 at 4.62%, GPT-5.4-mini at 6.10%) still hallucinate; 127 package names all models invent — rich target set for detection rules.",
     "https://arxiv.org/abs/2605.17062", ""),
    (44,
     "PackMonitor: Enabling Zero Package Hallucinations Through Decoding-Time Monitoring",
     "Liu et al.", "2026", "arXiv:2602.20717",
     "AI Agent Security",
     "slopsquatting, decoding-time defense, PyPI, npm, authoritative package list, prevention",
     "Unread",
     "Defense via decoding-time intervention constraining generation to authoritative package lists — informs ZeroTrust.sh's detection and verification strategy.",
     "https://arxiv.org/abs/2602.20717", ""),
    (45,
     "Secure or Suspect? Investigating Package Hallucinations of Shell Command in Original and Quantized LLMs",
     "Haque et al.", "2025", "arXiv:2512.08213",
     "AI Agent Security",
     "slopsquatting, quantized models, GGUF, shell command, realistic URL mimicry",
     "Unread",
     "Quantized models have significantly higher hallucination rates; fabricated packages mimic GitHub/golang.org URLs — relevant for sophisticated slopsquatting detection.",
     "https://arxiv.org/abs/2512.08213", ""),

    # ── AREA 9 — AI Agent Trust, Privilege Escalation & Prompt Injection ──────
    (46,
     "'Your AI, My Shell': Demystifying Prompt Injection Attacks on Agentic AI Coding Editors",
     "Liu et al.", "2025", "arXiv:2509.22040",
     "AI Agent Security",
     "prompt injection, Cursor, repository poisoning, agentic editor, shell execution hijack",
     "Unread",
     "First empirical analysis of Cursor via repository poisoning — attackers inject malicious instructions into dev resources to hijack shell execution. Direct ZeroTrust.sh threat model.",
     "https://arxiv.org/abs/2509.22040", ""),
    (47,
     "The Dark Side of LLMs: Agent-based Attacks for Complete Computer Takeover",
     "Lupinacci et al.", "2025", "arXiv:2507.06850",
     "AI Agent Security",
     "privilege escalation, inter-agent trust, 94.4% vulnerable, agent-to-agent attack, safety bypass",
     "Unread",
     "94.4% of models vulnerable to direct injection; 83.3% to inter-agent trust exploitation — key data for detecting agent-to-agent attack chains in ZeroTrust.sh.",
     "https://arxiv.org/abs/2507.06850", ""),
    (48,
     "VIGIL: Defending LLM Agents Against Tool Stream Injection via Verify-Before-Commit",
     "Lin et al.", "2026", "arXiv:2601.05755",
     "AI Agent Security",
     "tool stream injection, SIREN benchmark, verify-before-commit, runtime poisoning, 959 attack cases",
     "Unread",
     "Defense against tool stream injection; SIREN benchmark (959 attack cases) provides test corpus for ZeroTrust.sh detection validation.",
     "https://arxiv.org/abs/2601.05755", ""),
    (49,
     "Agentic AI Security: Threats, Defenses, Evaluation, and Open Challenges",
     "Datta et al.", "2025", "arXiv:2510.23883",
     "AI Agent Security",
     "survey, agentic AI, threat taxonomy, tool use, memory, autonomy, planning",
     "Unread",
     "Comprehensive survey of agentic AI threat taxonomy distinguishing autonomous execution risks from traditional LLM safety — essential scope reference for ZeroTrust.sh.",
     "https://arxiv.org/abs/2510.23883", ""),
    (50,
     "Are AI-assisted Development Tools Immune to Prompt Injection?",
     "Multiple authors", "2026", "arXiv:2603.21642",
     "AI Agent Security",
     "MCP, tool poisoning, Claude Code, Cursor, Cline, Continue, Gemini CLI, coding tools benchmark",
     "Unread",
     "First empirical analysis of MCP clients (Claude Code, Cursor, Cline, Continue, Gemini CLI) against tool-poisoning attacks — directly relevant to detecting injection in AI tool contexts.",
     "https://arxiv.org/abs/2603.21642", ""),
    (51,
     "Taming Various Privilege Escalation in LLM-Based Agent Systems: A Mandatory Access Control Framework",
     "Multiple authors", "2026", "arXiv:2601.11893",
     "AI Agent Security",
     "privilege escalation, MAC framework, SEAgent, formal model, agent security policy",
     "Unread",
     "Formal model of privilege escalation in LLM agents with SEAgent MAC framework — informs detection rules for trust boundary violations in agentic codebases.",
     "https://arxiv.org/abs/2601.11893", ""),

    # ── AREA 10 — Logic Vulnerability Detection (IDOR, Access Control) ────────
    (52,
     "BacAlarm: Mining and Simulating Composite API Traffic to Prevent Broken Access Control Violations",
     "Multiple authors", "2025", "arXiv:2512.19997",
     "Logic Vuln Detection",
     "broken access control, REST API, OWASP API, LLM agent, RAG, anomaly detection",
     "Unread",
     "Broken access control detection in REST APIs via LLM-based agent traffic simulation and ensemble anomaly detection — core to detecting authorization gaps in Path B.",
     "https://arxiv.org/abs/2512.19997", ""),
    (53,
     "Rethinking Broken Object Level Authorization Attacks Under Zero Trust Principle",
     "Wu et al.", "2025", "arXiv:2507.02309",
     "Logic Vuln Detection",
     "BOLA, IDOR, zero trust, resource ID dataflow, authorization interval, 35 new CVEs",
     "Unread",
     "BOLAZ: zero-trust defense for BOLA/IDOR via resource ID data flow analysis; discovered 35 new vulnerabilities — directly validates Path B's IDOR detection design.",
     "https://arxiv.org/abs/2507.02309", ""),
    (54,
     "VULSOLVER: Vulnerability Detection via LLM-Driven Constraint Solving",
     "Li et al.", "2025", "arXiv:2509.00882",
     "Logic Vuln Detection",
     "LLM, constraint solving, call-chain analysis, 96.29% accuracy, semantic reasoning",
     "Unread",
     "SAST + LLM semantic reasoning with progressive constraint modeling achieves 96.29% accuracy — applicable to detecting contextually-wrong authorization logic.",
     "https://arxiv.org/abs/2509.00882", ""),
    (55,
     "SAVANT: Vulnerability Detection in Application Dependencies through Semantic-Guided Reachability Analysis",
     "Multiple authors", "2025", "arXiv:2506.17798",
     "Logic Vuln Detection",
     "semantic preprocessing, LLM, reachability analysis, library API, 83.8% precision",
     "Unread",
     "Semantic preprocessing + LLM context analysis for vulnerable API patterns; 83.8% precision — relevant to detecting missing authorization checks in library calls.",
     "https://arxiv.org/abs/2506.17798", ""),
    (56,
     "Argus: Reorchestrating Static Analysis via a Multi-Agent Ensemble for Full-Chain Security Vulnerability Detection",
     "Multiple authors", "2025", "arXiv:2604.06633",
     "Logic Vuln Detection",
     "multi-agent, SAST, false positive reduction, multi-hop reasoning, zero-day, full-chain",
     "Unread",
     "First multi-agent LLM framework for SAST with multi-hop reasoning — discovers zero-day vulnerabilities spanning multiple functions; applicable to detecting business logic flaws.",
     "https://arxiv.org/abs/2604.06633", ""),
    (57,
     "Benchmarking LLMs and LLM-based Agents in Practical Vulnerability Detection for Code Repositories",
     "Multiple authors", "2025", "ACL 2025 / arXiv:2503.03586",
     "Logic Vuln Detection",
     "JitVul benchmark, 879 CVEs, interprocedural, ReAct agents, authorization flaws, ACL",
     "Unread",
     "JitVul benchmark (879 CVEs) requires interprocedural analysis; ReAct agents outperform LLMs on auth flaws — suggests agentic approaches for Path B's multi-function context.",
     "https://arxiv.org/abs/2503.03586", ""),
    (58,
     "Prompting the Priorities: A First Look at Evaluating LLMs for Vulnerability Triage and Prioritization",
     "Multiple authors", "2025", "arXiv:2510.18508",
     "Logic Vuln Detection",
     "SSVC, triage, prioritization, 384 real-world CVEs, risk surface selection, LLM evaluation",
     "Unread",
     "LLM evaluation on SSVC triage framework with 384 real-world vulnerabilities — methodology for prioritizing high-risk code surfaces before deep Path B analysis.",
     "https://arxiv.org/abs/2510.18508", ""),
    (59,
     "I Can't Believe It's Not a Valid Exploit",
     "Multiple authors", "2026", "arXiv:2602.04165",
     "Logic Vuln Detection",
     "PoC-Gym, exploit generation, static analysis guidance, 21% improvement, validation limits",
     "Unread",
     "PoC-Gym framework shows static analysis guidance improves LLM exploit generation by 21% — informs heuristics for filtering false-positive authorization scenarios in Path B.",
     "https://arxiv.org/abs/2602.04165", ""),

    # ── AREA 11 — Local LLM Deployment & On-Device Security ──────────────────
    (60,
     "Mind the Gap: A Practical Attack on GGUF Quantization",
     "Egashira et al.", "2025", "ICML 2025 / arXiv:2505.23786",
     "Local LLM Deployment",
     "GGUF, quantization, backdoor injection, Ollama, llama.cpp, supply chain, ICML",
     "Unread",
     "Backdoor injection via GGUF quantization errors — essential threat model for ZeroTrust.sh's local LLM deployment: quantized models can exhibit hidden malicious behaviors.",
     "https://arxiv.org/abs/2505.23786", ""),
    (61,
     "Widening the Gap: Exploiting LLM Quantization via Outlier Injection",
     "Shi et al.", "2025", "arXiv:2605.15152",
     "Local LLM Deployment",
     "quantization attack, AWQ, GPTQ, GGUF, outlier injection, adversarial, supply chain",
     "Unread",
     "First quantization-conditioned attack affecting AWQ/GPTQ/GGUF — demonstrates supply-chain risks in locally-deployed quantized models; complements GGUF threat model.",
     "https://arxiv.org/abs/2605.15152", ""),
    (62,
     "A First Look At Efficient And Secure On-Device LLM Inference Against KV Leakage",
     "Yang et al.", "2024", "arXiv:2409.04040",
     "Local LLM Deployment",
     "on-device LLM, KV cache leakage, FHE, TEE, privacy, local inference security",
     "Unread",
     "Privacy attacks on on-device LLM inference via KV cache leakage — demonstrates how local inference can leak conversation history; relevant to ZeroTrust.sh's privacy-first design.",
     "https://arxiv.org/abs/2409.04040", ""),
    (63,
     "A Cost-Benefit Analysis of On-Premise Large Language Model Deployment: Breaking Even with Commercial LLM Services",
     "Pan et al.", "2025", "arXiv:2509.18101",
     "Local LLM Deployment",
     "on-premise, cost analysis, breakeven, Qwen, Llama, Mistral, cloud vs local, value proposition",
     "Unread",
     "Quantifies breakeven for on-premise LLM vs cloud APIs — supports ZeroTrust.sh's value proposition: local deployment becomes cost-competitive at scale.",
     "https://arxiv.org/abs/2509.18101", ""),

    # ── AREA 13 — Prompt Compression & Context Reduction ─────────────────────
    (69,
     "LLMLingua: Compressing Prompts for Accelerated Inference of Large Language Models",
     "Jiang et al.", "2023", "arXiv:2310.05736",
     "Prompt Compression",
     "prompt compression, token pruning, 20x compression, budget control, coarse-to-fine",
     "Unread",
     "Iterative token-level compression with budget control; up to 20x compression with minimal loss — directly applicable to compressing code context before LLM security scans.",
     "https://arxiv.org/abs/2310.05736", ""),
    (70,
     "LLMLingua-2: Data Distillation for Efficient and Faithful Task-Agnostic Prompt Compression",
     "Jiang et al.", "2024", "arXiv:2403.12968",
     "Prompt Compression",
     "prompt compression, data distillation, token classification, 3-6x faster, 1.6-2.9x latency",
     "Unread",
     "Compression as token classification via GPT-4 distillation; 3–6x faster than LLMLingua with 2–5x compression ratios — ideal for repeated scans of similar code patterns.",
     "https://arxiv.org/abs/2403.12968", ""),
    (71,
     "LongLLMLingua: Accelerating and Enhancing LLMs in Long Context Scenarios via Prompt Compression",
     "Jiang et al.", "2023", "arXiv:2310.06839",
     "Prompt Compression",
     "long context, question-aware compression, document reordering, 94% cost reduction, 4x fewer tokens",
     "Unread",
     "Question-aware compression for long-context scenarios; 94% cost reduction on LooGLE — critical for scanning large codebases and enriching prompts with vulnerability databases.",
     "https://arxiv.org/abs/2310.06839", ""),
    (72,
     "RECOMP: Improving Retrieval-Augmented LMs with Compression and Selective Augmentation",
     "Xu et al.", "2023", "arXiv:2310.04408",
     "Prompt Compression",
     "RAG, retrieval augmentation, extractive compression, abstractive compression, 6% compression rate",
     "Unread",
     "Compresses retrieved documents in RAG pipelines via extractive and abstractive methods — relevant when ZeroTrust.sh augments prompts with retrieved CVE or security knowledge.",
     "https://arxiv.org/abs/2310.04408", ""),
    (73,
     "Pruning the Unsurprising: Efficient LLM Reasoning via First-Token Surprisal",
     "Multiple authors", "2025", "arXiv:2508.05988",
     "Prompt Compression",
     "CoT compression, first-token surprisal, 23.5% token reduction, 43.5% latency reduction, ASAP",
     "Unread",
     "ASAP framework compresses chain-of-thought by 23.5% tokens and 43.5% latency — optimizes reasoning chains in vulnerability analysis without sacrificing accuracy.",
     "https://arxiv.org/abs/2508.05988", ""),

    # ── AREA 14 — Prompt Engineering & Automatic Optimization ─────────────────
    (74,
     "Large Language Models as Optimizers (OPRO)",
     "Yang et al.", "2023", "arXiv:2309.03409",
     "Prompt Engineering",
     "automatic prompt optimization, OPRO, LLM-based refinement, in-context learning, 50% improvement",
     "Unread",
     "Automatic prompt optimization via LLM-based iterative refinement; best prompts beat human-designed by up to 50% — applicable to auto-tuning ZeroTrust.sh security scanning prompts.",
     "https://arxiv.org/abs/2309.03409", ""),
    (75,
     "Chain of Draft: Thinking Faster by Writing Less",
     "Xu et al.", "2025", "arXiv:2502.18600",
     "Prompt Engineering",
     "chain of draft, CoT efficiency, 7.6% tokens, reasoning cost reduction, latency",
     "Unread",
     "Reduces CoT verbosity to 7.6% of tokens while matching CoT accuracy — practical cost/latency reduction for reasoning-heavy vulnerability analysis in Path B.",
     "https://arxiv.org/abs/2502.18600", ""),
    (76,
     "Which Examples to Annotate for In-Context Learning? Towards Effective and Efficient Selection",
     "Mavromatis et al.", "2023", "arXiv:2310.20046",
     "Prompt Engineering",
     "few-shot selection, in-context learning, budget constraint, uncertainty sampling, diversity, AdaICL",
     "Unread",
     "AdaICL selects maximally informative few-shot examples under token budget constraints — applicable to selecting the best security examples per scan without exceeding token limits.",
     "https://arxiv.org/abs/2310.20046", ""),
    (77,
     "DecoPrompt: Decoding Prompts Reduces Hallucinations when Large Language Models Meet False Premises",
     "Xu and Ma", "2024", "arXiv:2411.07457",
     "Prompt Engineering",
     "hallucination reduction, false premises, decoding prompts, inference-time, no retraining",
     "Unread",
     "Inference-time hallucination reduction by decoding the prompt itself — cost-effective false positive reduction for ZeroTrust.sh without model retraining.",
     "https://arxiv.org/abs/2411.07457", ""),

    # ── AREA 15 — Inference Speed & KV-Cache Optimization ─────────────────────
    (78,
     "Unlocking Efficiency in Large Language Model Inference: A Comprehensive Survey of Speculative Decoding",
     "Multiple authors", "2024", "arXiv:2401.07851",
     "Inference Speed",
     "speculative decoding, draft model, parallel verification, inference acceleration, survey",
     "Unread",
     "Comprehensive survey of speculative decoding (small draft model + large verifier) for LLM inference acceleration — core technique for optimizing local Ollama/llama.cpp execution speed.",
     "https://arxiv.org/abs/2401.07851", ""),
    (79,
     "Keep the Cost Down: A Review on Methods to Optimize LLM's KV-Cache Consumption",
     "Multiple authors", "2024", "arXiv:2407.18003",
     "Inference Speed",
     "KV-cache, PagedAttention, vLLM, token pruning, VRAM, memory optimization, survey",
     "Unread",
     "Survey of KV-cache optimization strategies (PagedAttention, distributed KV-cache, token pruning) — reduces VRAM footprint for quantized GGUF models in ZeroTrust.sh local inference.",
     "https://arxiv.org/abs/2407.18003", ""),
    (80,
     "FlashAttention: Fast and Memory-Efficient Exact Attention with IO-Awareness",
     "Dao et al.", "2022", "arXiv:2205.14135",
     "Inference Speed",
     "flash attention, IO-aware, O(N) memory, foundational, attention efficiency, transformer",
     "Unread",
     "Foundational IO-aware attention reducing memory from O(N²) to O(N) — enables processing longer code snippets with lower memory overhead during semantic analysis.",
     "https://arxiv.org/abs/2205.14135", ""),
    (81,
     "Why Low-Precision Transformer Training Fails: An Analysis on Flash Attention",
     "Multiple authors", "2025", "arXiv:2510.04212",
     "Inference Speed",
     "low-precision, GGUF, flash attention, quantization, performance analysis, precision limits",
     "Unread",
     "Analyzes Flash Attention behavior under low-precision quantization — critical for understanding performance characteristics of quantized models used in ZeroTrust.sh's local deployment.",
     "https://arxiv.org/abs/2510.04212", ""),
    (82,
     "Optimizing LLM Inference Throughput via Memory-aware and SLA-constrained Dynamic Batching",
     "Multiple authors", "2025", "arXiv:2503.05248",
     "Inference Speed",
     "dynamic batching, throughput, SLA, memory-aware, continuous batching, parallel scanning",
     "Unread",
     "Memory-aware dynamic batching with SLA constraints — optimizes continuous batching for ZeroTrust.sh's parallel Path A/B analysis phases to maximize throughput.",
     "https://arxiv.org/abs/2503.05248", ""),

    # ── AREA 16 — Structured Generation & Constrained Decoding ───────────────
    (83,
     "XGrammar: Flexible and Efficient Structured Generation Engine for Large Language Models",
     "Multiple authors", "2024", "arXiv:2411.15100",
     "Structured Generation",
     "grammar-constrained decoding, BNF, CFG, JSON, zero overhead, structured output, XGrammar",
     "Unread",
     "Grammar-constrained decoding supporting BNF/CFG with zero overhead — enables Path B LLM to output structured JSON vulnerability findings without post-processing verification tokens.",
     "https://arxiv.org/abs/2411.15100", ""),
    (84,
     "JSONSchemaBench: A Rigorous Benchmark of Structured Outputs for Language Models",
     "Multiple authors", "2025", "arXiv:2501.10868",
     "Structured Generation",
     "JSON schema, structured output, Guidance, Outlines, XGrammar, benchmark, compliance",
     "Unread",
     "Benchmarks constrained-decoding frameworks (Guidance, Outlines, XGrammar) on JSON Schema compliance — validates structured generation reliability for ZeroTrust.sh vulnerability report output.",
     "https://arxiv.org/abs/2501.10868", ""),
    (85,
     "SynCode: LLM Generation with Grammar Augmentation",
     "Multiple authors", "2024", "arXiv:2403.01632",
     "Structured Generation",
     "grammar augmentation, syntactic constraints, output schema, structured generation, alternative to XGrammar",
     "Unread",
     "Grammar-augmented generation enforcing output structure via syntactic constraints — alternative to XGrammar for ensuring ZeroTrust.sh semantic output conforms to fixed schema.",
     "https://arxiv.org/abs/2403.01632", ""),

    # ── AREA 17 — Context Window Management ──────────────────────────────────
    (86,
     "Beyond RAG vs. Long-Context: Learning Distraction-Aware Retrieval for Efficient Knowledge Grounding",
     "Multiple authors", "2024", "arXiv:2509.21865",
     "Context Management",
     "RAG, long context, distraction-aware, lost in the middle, retrieval, code snippet selection",
     "Unread",
     "Distraction-aware retrieval mitigating 'lost in the middle' effect — addresses how ZeroTrust.sh should select and contextualize code snippets rather than passing entire files to the LLM.",
     "https://arxiv.org/abs/2509.21865", ""),

    # ── AREA 12 — Exploit Automation & Proof-of-Concept Generation ────────────
    (64,
     "Automated Vulnerability Validation and Verification: A Large Language Model Approach",
     "Lotfi et al.", "2025", "arXiv:2509.24037",
     "Exploit Automation",
     "CVE exploitation, Docker, automated validation, 70% reproduction rate, offline LLM, PoE layer",
     "Unread",
     "End-to-end LLM pipeline for automated CVE exploitation with Docker orchestration; 70% CVE reproduction rate — directly applicable to ZeroTrust.sh's Approach 3 PoE layer design.",
     "https://arxiv.org/abs/2509.24037", ""),
    (65,
     "Patch-to-PoC: A Systematic Study of Agentic LLM Systems for Linux Kernel N-Day Reproduction",
     "Pu et al.", "2026", "arXiv:2602.07287",
     "Exploit Automation",
     "PoC generation, kernel exploit, KernelCTF, VM sandbox, 50%+ success, K-REPRO agent",
     "Unread",
     "Autonomous LLM-based PoC generation from kernel patches with 50%+ success on KernelCTF — K-REPRO agent architecture with VM management informs sandbox PoE design.",
     "https://arxiv.org/abs/2602.07287", ""),
    (66,
     "PentestAgent: Incorporating LLM Agents to Automated Penetration Testing",
     "Feng et al.", "2024", "arXiv:2411.05185",
     "Exploit Automation",
     "penetration testing, multi-agent, VulHub, Docker, reconnaissance, exploitation, red team",
     "Unread",
     "Multi-agent framework for autonomous pentest using Docker VulHub environments — demonstrates agent architecture for reconnaissance and exploitation applicable to ZeroTrust.sh red-team layer.",
     "https://arxiv.org/abs/2411.05185", ""),
    (67,
     "Directed Greybox Fuzzing via Large Language Model",
     "Xu et al.", "2025", "arXiv:2505.03425",
     "Exploit Automation",
     "LLM fuzzing, path constraint, code generation, 9 new CVEs, 17/20 trigger rate, HGFuzzer",
     "Unread",
     "HGFuzzer uses LLMs to transform path constraints into code generation tasks; discovered 9 new CVEs with 17/20 trigger rate — alternative exploit discovery layer for Approach 3.",
     "https://arxiv.org/abs/2505.03425", ""),
    (68,
     "PoCGen: Generating Proof-of-Concept Exploits for Vulnerabilities in NPM Packages",
     "Shen et al.", "2025", "arXiv:2506.04962",
     "Exploit Automation",
     "PoC generation, npm, JavaScript, Node.js, 77% success rate, CVE disclosure, ecosystem-specific",
     "Unread",
     "Autonomous LLM-based PoC for npm vulnerabilities (77% on SecBench.js); 6 PoCs in official CVE disclosures — demonstrates ecosystem-specific PoE applicability for JavaScript targets.",
     "https://arxiv.org/abs/2506.04962", ""),
]

# ── Write paper rows ──────────────────────────────────────────────────────────
current_row = 5
last_paper_row = current_row  # track last row for auto-filter range

for entry in papers:
    (num, title, authors, year, venue, category, tags,
     read_status, relevance, url, lit_notes) = entry

    is_even = (current_row % 2 == 0)
    base_fill_color = LIGHT_BLUE if is_even else LIGHT_GRAY

    # Read Status cell gets its own color badge
    status_color_map = {
        "Unread":   STATUS_UNREAD,
        "Reading":  STATUS_READING,
        "Read":     STATUS_READ,
        "Reviewed": STATUS_REVIEWED,
    }

    values = [num, title, authors, year, venue, category, tags,
              read_status, relevance, url, lit_notes]

    for ci, value in enumerate(values, start=1):
        cell = ws2.cell(row=current_row, column=ci, value=value)
        cell_fill = base_fill_color
        if ci == 8:   # Read Status column
            cell_fill = status_color_map.get(read_status, STATUS_UNREAD)
        if ci == 11:  # Literature Review Notes — slightly warmer tint
            cell_fill = "FFFEF5" if not is_even else "F0F8FF"
        cell.fill = fill(cell_fill)
        cell.border = border()
        cell.alignment = align("left", "center", wrap=True)
        cell.font = font(size=10)

        # Column-specific overrides
        if ci == 1:   # #
            cell.alignment = align("center", "center", wrap=False)
            cell.font = font(bold=True, size=10)
        if ci == 4:   # Year
            cell.alignment = align("center", "center", wrap=False)
        if ci == 8:   # Read Status
            cell.alignment = align("center", "center", wrap=False)
            cell.font = font(bold=True, size=9)
        if ci == 10:  # URL — smaller font
            cell.font = Font(size=9, color="2E5FA3", name="Calibri",
                             underline="single")
        if ci == 11:  # Literature Review Notes — italic placeholder hint
            if not value:
                cell.font = Font(size=9, color="AAAAAA", italic=True, name="Calibri")
                cell.value = "Write your notes here…"

    ws2.row_dimensions[current_row].height = 60
    last_paper_row = current_row
    current_row += 1

# ── Auto-filter on header row (enables column sort + filter dropdowns) ────────
ws2.auto_filter.ref = f"A4:{LAST_COL}{last_paper_row}"

# ── Data validation: Read Status dropdown (column H) ─────────────────────────
dv = DataValidation(
    type="list",
    formula1='"Unread,Reading,Read,Reviewed"',
    showDropDown=False,      # False = show the dropdown arrow in Excel
    showErrorMessage=True,
    errorTitle="Invalid value",
    error='Choose: Unread, Reading, Read, or Reviewed',
)
ws2.add_data_validation(dv)
dv.add(f"H5:H{last_paper_row}")

# ── Legend row ────────────────────────────────────────────────────────────────
legend2_row = last_paper_row + 2
ws2.merge_cells(f"A{legend2_row}:{LAST_COL}{legend2_row}")
leg2 = ws2[f"A{legend2_row}"]
leg2.value = (
    "Read Status:  Unread (gray) — not yet read  ·  Reading (amber) — in progress  "
    "·  Read (blue) — finished  ·  Reviewed (green) — literature review written  "
    "·  Filter tip: click ▾ on Category or Tags header to filter by research area"
)
leg2.font = Font(italic=True, color="666666", size=9, name="Calibri")
leg2.alignment = align("left", "center")

ws2.freeze_panes = "B5"   # freeze row 4 header + column A (#)

# ── Save ──────────────────────────────────────────────────────────────────────
out = "docs/execution-overview.xlsx"
wb.save(out)
print(f"Saved: {out}  ({last_paper_row - 4} papers)")
